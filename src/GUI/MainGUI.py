# -*- coding: utf-8 -*-
# GUI 主脚本
# 作者：LF
# 创建时间：2026-08-23
# 功能：基于 PyQt6 的邮件助手图形界面。
#       主窗口提供「普通模式」「高级模式」两个入口按钮，点击进入对应发信页面。
#       普通模式发信页仿照 QQ 邮箱网页版写邮件布局：
#         顶部操作栏（发送 / 预览 / 附件 ▼ / 发信设置）
#         → 收件人（可展开抄送/密送）→ 主题
#         → 两行富文本工具栏
#         → 正文大输入区 → 左下发件人选择
#       界面使用 PyQt6 默认样式（Windows 原生控件外观），不做额外美化。

import os
import re
import sys
import time
import random
from datetime import datetime, timedelta

# 日志系统：记录 GUI 关键流程（页面切换、发送、设置修改等）
from logger import getLogger
log = getLogger(__name__)

from PyQt6.QtCore import Qt, QThread, QTimer, QUrl, QRect, QRectF, pyqtSignal
from PyQt6.QtGui import (
    QAction, QColor, QFont, QImage, QPainter, QPen,
    QTextCharFormat, QTextCursor, QTextDocument, QTextImageFormat, QTextListFormat,
)
from PyQt6.QtWidgets import (
    QMainWindow, QStackedWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTextEdit, QFontComboBox,
    QTextBrowser, QComboBox, QListWidget, QListWidgetItem, QFileDialog, QColorDialog, QMessageBox,
    QAbstractItemView, QMenu, QToolButton, QSizePolicy, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame, QDialog, QTabWidget,
    QDateEdit, QRadioButton, QGroupBox, QScrollArea, QSplitter, QButtonGroup, QSpinBox,
    QStyle, QStyleOptionFrame, QFormLayout,
)

# 跨文件调用说明：从 src/Email/NormalMode.py 导入 NormalMode 类，
# 由它统一编排参数并对接 LEmail 完成邮件发送（GUI 只负责收集界面输入，
# 不直接操作 LEmail，保证发送逻辑复用普通模式的上游处理）。
# GUI 与 Email 是同级顶层包（src 非包），跨包须用绝对导入；
# 其 sys.path 由启动脚本（main.py / runTest.py）经 PathManager.source_root()
# 统一注入，开发与打包环境均能正确定位。
from Email.NormalMode import NormalMode
# 复用本地配置加载的发信账号作为默认发件邮箱（config.local，不入库）
from Email.MainEmail import SMTP_USERNAME
from database import (
    getDb,
    CATEGORY_DRAFT,
    CATEGORY_SENT,
    CATEGORY_DELETED,
)
# UI 按钮响应之后的业务逻辑收敛在 activity 包中，界面层仅负责调用
from activity.recipient_bulk import (
    parseRecipientFile,
    mergeRecipients,
    sortRecipients,
    filterRecipients,
    splitRecipients,
    validateRecipients,
)
from activity.mail_query import queryEmails

# 设置模块：option 包位于项目根（非 src），先注入项目根到 sys.path 再导入设置管理器单例
if os.path.dirname(os.path.dirname(os.path.abspath(__file__))) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from option.settings import settings

# 发件人昵称默认值（发件人昵称输入框留空时使用）；来源改为用户设置 default_from_name
DEFAULT_FROM_NAME = settings.get('default_from_name')
# SMTP 默认发件邮箱地址：复用本地 config 中加载的发信账号；缺失时用示例占位
DEFAULT_FROM_EMAIL = SMTP_USERNAME or 'your-email@example.com'


def refreshRuntimeDefaults():
    """设置保存后刷新受设置影响的运行时默认值

    设置界面写入 default_from_name 后调用本函数，用 global 更新模块级
    DEFAULT_FROM_NAME，使后续发送在昵称留空时用到新的默认昵称。
    """
    global DEFAULT_FROM_NAME
    DEFAULT_FROM_NAME = settings.get('default_from_name')

# 字号下拉框候选字号列表（单位：磅）
FONT_SIZES = ['8', '9', '10', '11', '12', '14', '16', '18', '20', '24', '28', '32', '36', '48', '72']

# 高级模式「配置页」全局样式表（QSS）：控制分组区块、进度指示、变量面板、主按钮与黄字提示
CONFIG_STYLE = """
QFrame#configSection { background:#fafbfc; border:1px solid #e3e6ea; border-radius:8px; }
QLabel#sectionTitle { color:#1f2329; font-weight:600; font-size:14px; }
QLabel#sectionHint { color:#8a919b; font-size:12px; }
QWidget#varPanel { background:#f6f8fb; border:1px solid #e0e4ea; border-radius:8px; }
QLabel#varTitle { color:#1f2329; font-weight:600; font-size:14px; }
QLabel#varHint, QLabel#varNote { color:#8a919b; font-size:12px; }
QPushButton#varMarkBtn { background:#2f6fed; color:#ffffff; font-weight:600;
    border:none; border-radius:4px; padding:6px 12px; }
QPushButton#varMarkBtn:hover { background:#2859c9; }
QPushButton#varMarkBtn:pressed { background:#1f4bad; }
QListWidget#varList { background:#ffffff; border:1px solid #e3e6ea;
    border-radius:6px; padding:2px; }
QLabel#yellowTip { color:#b06a00; background:#fff8e1; border:1px solid #f3d185;
    border-radius:4px; padding:6px 8px; }
QPushButton#primaryBtn { background:#2f6fed; color:#ffffff; font-weight:600;
    border:none; border-radius:4px; }
QPushButton#primaryBtn:hover { background:#2859c9; }
QPushButton#primaryBtn:pressed { background:#1f4bad; }
QLabel#stepActive { background:#2f6fed; color:#ffffff; font-weight:600;
    border-radius:12px; padding:6px 12px; }
QLabel#stepDone { background:#e8efff; color:#2f6fed; font-weight:500;
    border-radius:12px; padding:6px 12px; }
QLabel#stepTodo { background:#f1f3f5; color:#9aa1ab; font-weight:500;
    border-radius:12px; padding:6px 12px; }
QLabel#loadHint { color:#b8860b; font-weight:600; padding:2px 0; }
QLabel#loadRowName { color:#1f2329; font-weight:600; min-width:96px; }
QLabel#loadPathEmpty { color:#c0c4cc; }
QLabel#loadPathSet { color:#2f6fed; font-size:11pt; }
QListWidget#loadList { background:#ffffff; border:1px solid #e3e6ea;
    border-radius:10px; padding:4px; }
/* 负载配给页：标记选择区 */
QLabel#loadSectionTitle { color:#1f2329; font-weight:700; font-size:15px; }
QLabel#loadSubTitle { color:#1f2329; font-weight:600; font-size:13px; }
QLabel#loadStatusOk { color:#18794e; font-weight:600; }
QLabel#loadStatusPending { color:#b06a00; font-weight:600; }
QComboBox#loadMarkCombo { background:#ffffff; border:1px solid #d0d3d9;
    border-radius:6px; padding:4px 8px; min-height:28px; }
/* 负载配给页：Payload设置区 */
QFrame#loadPayloadFrame { background:#fafbfc; border:1px solid #e3e6ea;
    border-radius:8px; }
QPushButton#loadActionBtn { background:#ffffff; border:1px solid #d0d3d9;
    border-radius:5px; padding:6px 10px; color:#1f2329; }
QPushButton#loadActionBtn:hover { background:#f1f3f5; border-color:#2f6fed; color:#2f6fed; }
QPushButton#loadActionBtn:disabled { color:#c0c4cc; border-color:#e8eaed; background:#fafbfc; }
QPushButton#loadPrimaryBtn { background:#2f6fed; color:#ffffff; font-weight:600;
    border:none; border-radius:5px; padding:6px 12px; }
QPushButton#loadPrimaryBtn:hover { background:#2859c9; }
QPushButton#loadPrimaryBtn:disabled { background:#a8c0f5; color:#ffffff; }
QListWidget#loadPreviewList { background:#ffffff; border:1px solid #e3e6ea;
    border-radius:6px; padding:2px; font-family:'Consolas','Courier New',monospace;
    font-size:12px; }
QLabel#loadFileInfo { color:#595f69; font-size:12px; padding:4px 0; }
QLabel#loadFileInfoEmpty { color:#c0c4cc; font-size:12px; padding:4px 0; }
/* 负载配给页：绑定规则表 */
QTableWidget#loadBindingTable { background:#ffffff; gridline-color:#e8eaed;
    border:1px solid #e3e6ea; border-radius:6px; }
QTableWidget#loadBindingTable::item { padding:4px 6px; }
QHeaderView::section { background:#f6f8fb; color:#1f2329; font-weight:600;
    padding:6px; border:1px solid #e3e6ea; }
QPushButton#loadDelBtn { background:transparent; color:#d03050; border:none;
    padding:2px 6px; font-size:12px; }
QPushButton#loadDelBtn:hover { text-decoration:underline; }
"""

# 主窗口左侧菜单：(id, 显示名)，用于左栏列表 + 选中态切换
# 说明：本项目只提供「发信」服务，不收邮件，故无「收件箱/星标/群邮件/记事本/垃圾箱」。
NAV_MENU_ITEMS = [
    ('drafts',    '草稿箱'),
    ('sent',      '已发送'),
    ('deleted',   '已删除'),
]

# 左侧菜单 id 与数据库分类常量映射（列表页读取 / 增写使用）
MENU_TO_CATEGORY = {
    'drafts':  CATEGORY_DRAFT,
    'sent':    CATEGORY_SENT,
    'deleted': CATEGORY_DELETED,
}

# 左栏菜单中可对应到邮件列表页的 id（其分类均有真实列表展示）
LIST_PAGE_MENUS = {'drafts', 'sent', 'deleted'}

# 「已删除」邮件保留天数：到期自动永久删除（默认 30 天）
DELETE_RETENTION_DAYS = 30


def parseEmails(text):
    """将逗号/分号/空白分隔的邮箱串拆分为去空后的列表

    参数：
        text<str>：用户输入的邮箱地址串（可含多个，用逗号/分号/空白分隔）

    返回：
        list：拆分去空后的邮箱地址列表
    """
    # 正则按逗号、分号、中文逗号/分号、空白拆分，并过滤空项
    return [part for part in re.split(r'[,;，；\s]+', text.strip()) if part]


class LoadingOverlay(QWidget):
    """发送中遮罩：半透明覆盖整个窗口，中心绘制旋转加载圆圈，下方显示提示文字

    通过 QTimer 驱动旋转角度，由于实际发送在 SendWorker 线程中执行，
    主线程事件循环保持空闲，转圈动画可正常刷新。
    """

    def __init__(self, parent, message='正在发送中...'):
        super().__init__(parent)
        # 遮罩背景半透明（设置为可样式化背景以便背景色生效）
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet('background-color: rgba(0, 0, 0, 80);')
        # 提示文字与加载圆圈相关参数
        self._message = message
        self._angle = 0
        # 定时器：定时更新旋转角度并触发重绘
        self._timer = QTimer(self)
        self._timer.setInterval(50)
        self._timer.timeout.connect(self._rotate)
        # 默认隐藏
        self.hide()

    def _rotate(self):
        """定时器回调：旋转角度自增并触发重绘"""
        self._angle = (self._angle + 12) % 360
        self.update()

    def showLoading(self):
        """显示遮罩并覆盖父窗口全部区域，非阻塞式启动动画"""
        parent = self.parentWidget()
        if parent is not None:
            # 覆盖整个父窗口（顶层窗口）区域
            self.setGeometry(parent.rect())
            self.raise_()
        self.show()
        self._timer.start()

    def hideLoading(self):
        """隐藏遮罩并停止动画"""
        self._timer.stop()
        self.hide()

    def paintEvent(self, event):
        """自绘：半透明背景 + 中心旋转圆弧加载圈 + 下方文字"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # 1. 绘制半透明遮罩背景
        painter.fillRect(self.rect(), QColor(0, 0, 0, 80))
        center = self.rect().center()
        # 2. 绘制旋转加载圆环（缺口圆弧，模拟加载转圈）
        radius = 32
        ring_width = 6
        painter.setPen(QPen(QColor(255, 255, 255), ring_width,
                            Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        painter.drawArc(center.x() - radius, center.y() - radius,
                        radius * 2, radius * 2, self._angle * 16, 270 * 16)
        # 3. 在圆圈下方绘制提示文字
        font = painter.font()
        font.setPointSize(13)
        painter.setFont(font)
        painter.setPen(QColor(255, 255, 255))
        painter.drawText(center.x(), center.y() + radius + 30, self._message)
        painter.end()


class SendWorker(QThread):
    """邮件发送工作线程：在后台执行 NormalMode 发送，避免阻塞 GUI 动画主线程

    通过信号将发送结果传回主线程。
    """

    # 发送完成信号：参数为是否发送成功
    finished_signal = pyqtSignal(bool)

    def __init__(self, normal_mode):
        super().__init__()
        # 持有普通模式对象（内含已配置好的 LEmail）
        self.normal_mode = normal_mode

    def run(self):
        """线程体：调用普通模式发送接口并发出结果信号"""
        result = self.normal_mode.send()
        self.finished_signal.emit(result)


class BatchSendWorker(QThread):
    """批量邮件发送工作线程：高级模式第5步，逐个收件人单独发送

    为规避收件方垃圾邮件流控对整批发信的影响，每封之间随机休息
    SEND_INTERVAL_CANDIDATES（0.2~0.5 秒）再发下一封，避免固定节奏被识别为群发。
    每封完成后通过信号把结果回传主线程，供发送页实时更新状态。
    """

    # 相邻两封发送的随机间隔候选（秒）：避开固定频率，降低被判垃圾邮件/IP封禁概率
    SEND_INTERVAL_CANDIDATES = (0.2, 0.3, 0.4, 0.5)

    # 单封发送完成信号：行号, 是否成功, 结果说明
    mail_done = pyqtSignal(int, bool, str)
    # 单封开始发送信号：行号（供界面把该行状态改为「正在发送…」）
    mail_started = pyqtSignal(int)
    # 全部发送完成信号：成功数, 失败数
    all_done = pyqtSignal(int, int)

    def __init__(self, send_items, interval_candidates=SEND_INTERVAL_CANDIDATES):
        """初始化批量发送线程

        参数：
            send_items<list>：待发送项，每项为 dict，含构造单封 NormalMode 所需的字段
            interval_candidates<tuple>：随机间隔候选集合
        """
        super().__init__()
        self.send_items = list(send_items)
        self.interval_candidates = interval_candidates

    def _buildAndSendOne(self, item):
        """构造单封 NormalMode 并发送，返回 (是否成功, 说明)

        参数：
            item<dict>：单封发送配置（由主线程传入）
        返回：
            tuple<bool, str>
        """
        from src.Email.NormalMode import NormalMode
        try:
            nm = NormalMode()
            kwargs = dict(
                n_name=item.get('n_name'),
                to_list=item['to_list'],
                email_title=item.get('email_title'),
                html_text=item.get('html_text'),
                attachment_paths=item.get('attachment_paths'),
                inline_image_paths=item.get('inline_image_paths'),
                cc_list=item.get('cc_list'),
                bcc_list=item.get('bcc_list'),
                reply_to=item.get('reply_to'),
                return_email=item.get('return_email'),
            )
            if not nm.setConfig(**kwargs):
                return False, '邮件配置失败'
            over_limits, limit_msg = nm.checkLimits()
            if over_limits:
                return False, limit_msg
            over_size, _, _ = nm.checkSize()
            if over_size:
                return False, '邮件总大小超过上限，请改用网盘链接'
            ok = nm.send()
            return (True, '发送成功') if ok else (False, '发送失败')
        except Exception as e:   # 网络/配置等异常统一转为失败
            return False, '异常：%s' % e

    def run(self):
        """线程体：逐封发送并按随机间隔节流，全部结束后发完成信号"""
        ok_cnt = 0
        fail_cnt = 0
        count = len(self.send_items)
        for idx, item in enumerate(self.send_items):
            self.mail_started.emit(idx)       # 通知界面该行开始发送
            success, msg = self._buildAndSendOne(item)
            if success:
                ok_cnt += 1
            else:
                fail_cnt += 1
            self.mail_done.emit(idx, success, msg)
            # 非末尾一封：随机休息，避免固定节奏被判为批量群发
            if idx < count - 1:
                time.sleep(random.choice(self.interval_candidates))
        self.all_done.emit(ok_cnt, fail_cnt)


class SearchDialog(QDialog):
    """邮件查询对话框：按 主题 / 收件人 / 发件人 / 时间范围 组合搜索

    提供四个筛选条件（均可留空=不限）：
      主题（标题模糊）、收件人、发件人、发送日期区间。
    暂不支持按正文内容搜索。确认后把条件回传给调用方执行查询。
    """

    def __init__(self, category_label, parent=None):
        """初始化查询对话框

        参数：
            category_label<str>：当前分类的中文名（仅提示用途）
            parent<QWidget|None>：父窗口
        """
        super().__init__(parent)
        self.setWindowTitle('查询邮件')
        self.setMinimumWidth(420)
        v = QVBoxLayout(self)
        v.setSpacing(10)

        # 提示当前查询范围（当前分类）
        tip = QLabel('当前范围：%s（查询主题/收件人/发件人/发送日期，不含正文）' % category_label)
        tip.setStyleSheet('color:#666;')
        v.addWidget(tip)

        # 主题
        self.title_edit = QLineEdit()
        v.addLayout(self._labeledEdit('主题', self.title_edit))
        # 收件人
        self.recipient_edit = QLineEdit()
        v.addLayout(self._labeledEdit('收件人', self.recipient_edit))
        # 发件人
        self.sender_edit = QLineEdit()
        v.addLayout(self._labeledEdit('发件人', self.sender_edit))

        # 时间范围行（由复选框开启，关闭时不参与筛选）
        time_row = QHBoxLayout()
        self.date_check = QCheckBox('限定发送日期')
        self.date_check.toggled.connect(lambda on: self._toggleTimeRange(on))
        time_row.addWidget(self.date_check)
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat('yyyy-MM-dd')
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat('yyyy-MM-dd')
        time_row.addWidget(self.date_from)
        time_row.addWidget(QLabel(' 至 '))
        time_row.addWidget(self.date_to)
        time_row.addStretch()
        v.addLayout(time_row)
        # 默认关闭时间范围（避免默认日期把结果筛空）
        self.date_check.setChecked(False)
        self._toggleTimeRange(False)

        # 按钮
        btns = QHBoxLayout()
        btns.addStretch()
        clear_btn = QPushButton('清空')
        clear_btn.clicked.connect(self.clearInputs)
        search_btn = QPushButton('查询')
        search_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;border:0;'
                                 'padding:6px 20px;border-radius:4px;}')
        search_btn.clicked.connect(self.accept)
        btns.addWidget(clear_btn)
        btns.addWidget(search_btn)
        v.addLayout(btns)

    def _labeledEdit(self, label, edit):
        """构造一行「标签 + 输入框」的水平布局

        参数：
            label<str>：左侧标签文字
            edit<QLineEdit>：目标输入框（作为成员传回，取值/清空直接用成员）

        返回：
            QHBoxLayout：标签(定宽) + 输入框 的一行布局
        """
        row = QHBoxLayout()
        lab = QLabel(label)
        lab.setFixedWidth(70)
        row.addWidget(lab)
        row.addWidget(edit)
        return row

    def clearInputs(self):
        """清空全部查询条件"""
        self.title_edit.clear()
        self.recipient_edit.clear()
        self.sender_edit.clear()
        # 时间范围恢复默认关闭（避免残留过滤条件）
        self.date_check.setChecked(False)

    def _toggleTimeRange(self, enabled):
        """勾选「限定发送日期」时启用日期控件，否则禁用

        参数：
            enabled<bool>：是否启用日期范围筛选
        """
        self.date_from.setEnabled(enabled)
        self.date_to.setEnabled(enabled)

    def collectConditions(self):
        """收集并返回查询条件（供 MainWindow 执行搜索用）

        返回：
            dict：含 title / recipient / sender / date_from / date_to
                  （date_from/date_to 仅在开启日期限定时有值，否则为 None）
        """
        ret = {
            'title': self.title_edit.text().strip(),
            'recipient': self.recipient_edit.text().strip(),
            'sender': self.sender_edit.text().strip(),
            'date_from': None,
            'date_to': None,
        }
        # 勾选日期限定才返回日期区间；否则不参与筛选，避免默认日期筛空结果
        if self.date_check.isChecked():
            ret['date_from'] = self.date_from.date().toString('yyyy-MM-dd')
            ret['date_to'] = self.date_to.date().toString('yyyy-MM-dd')
        return ret


class SettingsDialog(QDialog):
    """应用设置页：铺满窗口，编辑可修改的用户级设置项，保存后立即生效

    当前开放给用户的设置项仅两类（高风险的连接/清理参数保持代码写死，不开放）：
      - 默认发件人昵称（default_from_name）：发件人昵称留空时使用的默认名称
      - 预览批量生成邮件数（preview_batch_size）：高级模式预览页「继续生成」每批数量
    顶部提供「返回主页」按钮，关闭本页回到主窗口。
    """

    def __init__(self, parent=None):
        """初始化设置页

        参数：
            parent<QWidget|None>：父窗口（用于铺满窗口并定位）
        """
        super().__init__(parent)
        self.setWindowTitle('设置')
        self.setModal(True)
        # 铺满父窗口，作为扁平化的整页设置界面
        if parent is not None:
            self.resize(parent.size())
        self.setStyleSheet(CONFIG_STYLE)
        self._buildUi()
        self._loadValues()

    def _buildUi(self):
        """构建设置页布局：顶部标题栏 + 返回主页按钮 + 设置表单 + 底部保存按钮"""
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(14)

        # ---- 顶部：标题 + 返回主页按钮 ----
        top = QHBoxLayout()
        title = QLabel('设置')
        tf = title.font()
        tf.setPointSize(16)
        tf.setBold(True)
        title.setFont(tf)
        top.addWidget(title)
        top.addStretch()
        back_btn = QPushButton('返回主页')
        back_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;'
                               'border:0;padding:6px 16px;border-radius:4px;}'
                               'QPushButton:hover{background:#1567d0;}')
        back_btn.clicked.connect(self.goBack)
        top.addWidget(back_btn)
        root.addLayout(top)

        # ---- 设置表单区 ----
        form = QFormLayout()
        form.setSpacing(16)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        # 默认发件人昵称
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText('发件人昵称留空时的默认显示名称')
        self.name_edit.setStyleSheet('QLineEdit{padding:6px 8px;border:1px solid #d0d7e2;'
                                     'border-radius:4px;background:#fff;}')
        form.addRow('默认发件人昵称：', self.name_edit)
        # 预览批量生成邮件数
        self.batch_spin = QSpinBox()
        self.batch_spin.setRange(1, 100)
        self.batch_spin.setSuffix(' 封')
        self.batch_spin.setStyleSheet('QSpinBox{padding:4px 8px;border:1px solid #d0d7e2;'
                                      'border-radius:4px;background:#fff;}')
        form.addRow('预览批量生成邮件数：', self.batch_spin)
        hint = QLabel('仅影响高级模式预览页「继续生成」的每批数量，数值越大生成越快但可能卡顿。')
        hint.setStyleSheet('color:#888;font-size:12px;')
        hint.setWordWrap(True)
        form.addRow('', hint)
        # CSV/TXT 收件人数据文件默认编码
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(['UTF-8', 'GBK', 'GB2312', 'GB18030', 'BIG5'])
        self.encoding_combo.setStyleSheet('QComboBox{padding:4px 8px;border:1px solid #d0d7e2;'
                                          'border-radius:4px;background:#fff;}')
        form.addRow('CSV/TXT 文件编码：', self.encoding_combo)
        enc_hint = QLabel('导入收件人 CSV/TXT 时的默认编码；中文乱码时切换为 GBK。')
        enc_hint.setStyleSheet('color:#888;font-size:12px;')
        form.addRow('', enc_hint)
        # 邮件发信追踪开关（功能预留）
        self.tracking_check = QCheckBox('开启邮件发信追踪（统计打开率）')
        self.tracking_check.setToolTip('在邮件头嵌入追踪标记以统计打开率。当前版本仅预留此开关，'
                                       '实际追踪逻辑后续版本生效。')
        form.addRow('发信追踪：', self.tracking_check)
        track_hint = QLabel('当前为功能预留位：保存后不会立即发送追踪标记，待后续版本接通用。')
        track_hint.setStyleSheet('color:#888;font-size:12px;')
        form.addRow('', track_hint)
        # 日志保存路径
        log_row = QHBoxLayout()
        self.log_edit = QLineEdit()
        self.log_edit.setStyleSheet('QLineEdit{padding:6px 8px;border:1px solid #d0d7e2;'
                                    'border-radius:4px;background:#fff;}')
        log_row.addWidget(self.log_edit)
        browse_btn = QPushButton('浏览…')
        browse_btn.setStyleSheet('QPushButton{background:#f0f2f5;color:#333;border:1px solid #d0d7e2;'
                                 'padding:5px 12px;border-radius:4px;}'
                                 'QPushButton:hover{background:#e3e7ec;}')
        browse_btn.clicked.connect(self._browseLogPath)
        log_row.addWidget(browse_btn)
        form.addRow('日志保存路径：', log_row)
        log_hint = QLabel('日志模块保存目录；默认开发为项目根 logs，打包后为安装目录 logs。')
        log_hint.setStyleSheet('color:#888;font-size:12px;')
        form.addRow('', log_hint)
        root.addLayout(form)

        root.addStretch()
        # ---- 底部保存按钮 ----
        save_btn = QPushButton('保存设置')
        save_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;'
                               'border:0;padding:8px 28px;border-radius:4px;font-weight:bold;}'
                               'QPushButton:hover{background:#1567d0;}')
        save_btn.clicked.connect(self.saveChanges)
        save_row = QHBoxLayout()
        save_row.addStretch()
        save_row.addWidget(save_btn)
        root.addLayout(save_row)

    def _loadValues(self):
        """把当前设置值回填到各控件"""
        self.name_edit.setText(str(settings.get('default_from_name')))
        self.batch_spin.setValue(int(settings.get('preview_batch_size')))
        # CSV/TXT 文件编码：期望值对应下拉项时选中该项，否则回退首个（UTF-8）
        enc = str(settings.get('data_file_encoding'))
        idx = self.encoding_combo.findText(enc)
        self.encoding_combo.setCurrentIndex(idx if idx >= 0 else 0)
        # 发信追踪开关 + 日志路径
        self.tracking_check.setChecked(bool(settings.get('mail_tracking_enabled')))
        self.log_edit.setText(str(settings.get('log_path') or ''))

    def saveChanges(self):
        """读取界面输入写回设置并落盘，随后刷新运行时默认值以立即生效"""
        name = self.name_edit.text().strip()
        # 昵称留空时回退到内置默认值，避免保存空昵称
        settings.set('default_from_name',
                     name or settings.DEFAULTS['default_from_name'])
        settings.set('preview_batch_size', self.batch_spin.value())
        # 新增三项：文件编码既服务于收件人解析；追踪开关为后续功能预留；日志路径供日志模块使用
        settings.set('data_file_encoding', self.encoding_combo.currentText())
        settings.set('mail_tracking_enabled', self.tracking_check.isChecked())
        settings.set('log_path', self.log_edit.text().strip())
        settings.save()
        # 让默认发件人昵称在后续发送时立即生效
        refreshRuntimeDefaults()
        log.info('设置已保存：昵称=%s 批数=%d 编码=%s 追踪=%s 日志路径=%s',
                 settings.get('default_from_name'),
                 settings.get('preview_batch_size'),
                 settings.get('data_file_encoding'),
                 settings.get('mail_tracking_enabled'),
                 settings.get('log_path'))
        QMessageBox.information(self, '保存成功', '设置已保存并立即生效。')

    def _browseLogPath(self):
        """打开目录选择框，把所选目录写入日志路径输入框

        首次浏览时以当前输入值或用户主目录作为起始目录。
        """
        start = self.log_edit.text().strip() or os.path.expanduser('~')
        folder = QFileDialog.getExistingDirectory(self, '选择日志保存目录', start)
        if folder:
            self.log_edit.setText(folder)

    def goBack(self):
        """返回主页：关闭设置页回到主窗口"""
        self.reject()


class MainWindow(QMainWindow):
    """仿 QQ 邮箱网页版主窗口

    结构：[顶部栏: 标题 + 搜索框] | [左栏: 写信按钮 + 导航菜单] | [右栏: 列表/写信页堆叠]
    流程：点击「写信」→ 跳到 ComposeModePage（普通模式 / 高级模式 两按钮）
          → 选择模式 → 跳到具体发信页 → 发送成功后自动返回列表主页
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle('Lhack 邮箱助手')
        self.resize(1120, 720)
        # 全局数据库引用（邮件记录读写统一走此实例）
        self.db = getDb()
        log.info('主窗口初始化')

        # 中心容器：一个 QWidget 承载三栏布局；右栏内容用 QStackedWidget 管理多页切换
        center = QWidget()
        self.setCentralWidget(center)
        root = QVBoxLayout(center)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # 1) 顶部栏：左侧标题、右侧搜索框（查询用）
        root.addWidget(self.buildTopHeader())

        # 2) 主体：左栏（导航） + 右栏（内容堆叠）
        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0)
        body.setSpacing(0)
        self.left_pane = self.buildLeftPane()
        body.addWidget(self.left_pane, 0)   # 左栏：固定宽
        # 右栏堆叠容器（4 页：列表主页、模式选择页、普通发信页、高级模式占位页）
        self.stack = QStackedWidget()
        body.addWidget(self.stack, 1)             # 右栏：占剩余空间
        root.addLayout(body, 1)

        # ---------- 堆叠页装配 ----------
        # 邮件列表主页（切换左侧菜单时复用同一页面，只更新标题与占位行）
        self.list_page = MailListPage(
            db=self.db,
            on_click_compose=self.showComposeModePage,
            on_click_view=self.onRequestView,
            on_click_edit=self.onRequestEdit,
            on_click_delete=self.onRequestDelete,
            on_open_task=self.onRequestOpenTask,
            on_click_task_delete=self.onRequestTaskDelete,
            on_click_task_restore=self.onRequestTaskRestore,
            on_back_task=self.onTaskListBack,
        )
        self.stack.addWidget(self.list_page)

        # 模式选择页（点写信后出现，展示普通/高级两个按钮）
        self.mode_page = ComposeModePage(
            on_click_normal=self.showNormalPage,
            on_click_high=self.showHighModePlaceholder,
            on_cancel=self.showListPage,
        )
        self.stack.addWidget(self.mode_page)

        # 普通模式发信页：返回动作切回列表主页；发送成功也自动回列表主页
        self.normal_page = NormalPage(db=self.db, on_back=self.showListPage)
        # 注入「恢复」回调（查看已删除时点「恢复」恢复回草稿箱）
        self.normal_page.setRestoreCallback(self.onRequestRestore)
        self.stack.addWidget(self.normal_page)

        # 高级模式发信页：支持在主题/正文中标记变量
        self.high_page = AdvancedPage(db=self.db, on_back=self.showListPage)
        self.stack.addWidget(self.high_page)

        # 默认进入：列表主页 + 左栏默认选中「草稿箱」（本项目只发不收）
        self.setActiveNav('drafts')
        self.stack.setCurrentWidget(self.list_page)
        # 启动时清理已过期的「已删除」邮件（默认保留 30 天）
        self.db.deleteExpired(CATEGORY_DELETED, DELETE_RETENTION_DAYS)

    # ---------------- 顶部栏 ----------------
    def buildTopHeader(self):
        """顶部栏：左侧「Lhack 邮箱」标题、右侧查询输入框（UI 占位，搜索逻辑不做）"""
        bar = QFrame()
        bar.setFrameShape(QFrame.Shape.NoFrame)
        # 顶部栏轻背景（QQ邮箱深蓝风格）
        bar.setStyleSheet('QFrame{background:#1b7bf2; color:#fff;}'
                          'QLineEdit{background:#4a9af6; color:#fff; border:1px solid #3f8eee;'
                          'padding:4px 8px; border-radius:4px;}'
                          'QLineEdit::placeholder{color:#d9e8ff;}')
        h = QHBoxLayout(bar)
        h.setContentsMargins(16, 10, 16, 10)
        h.setSpacing(16)

        title = QLabel('Lhack 邮箱')
        tf = title.font(); tf.setPointSize(13); tf.setBold(True)
        title.setFont(tf)
        title.setStyleSheet('color:#fff;')
        h.addWidget(title)
        h.addStretch()

        # 搜索框：回车触发查询
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText('搜索邮件（需点击查询）')
        self.search_edit.setMinimumWidth(280)
        self.search_edit.returnPressed.connect(self.onSearch)
        self.search_btn = QPushButton('查询')
        self.search_btn.setStyleSheet('QPushButton{background:#fff;color:#1b7bf2;border:0;'
                                      'padding:4px 12px;border-radius:4px;font-weight:bold;}'
                                      'QPushButton:hover{background:#e6efff;}')
        self.search_btn.clicked.connect(self.onSearch)
        h.addWidget(self.search_edit)
        h.addWidget(self.search_btn)
        return bar

    def onSearch(self):
        """查询功能实装：弹出查询对话框，按主题/收件人/发件人/时间范围搜索并展示结果

        查询范围限定在当前选中的分类（草稿箱/已发送/已删除）；检索结果直接渲染进列表页。
        对话框内条件全部留空时同样执行（等价于列出该分类全部邮件）。
        """
        # 当前分类中文名（用于对话框提示范围）
        current_label = self._menuName(self._current_menu_id) or '全部'
        dlg = SearchDialog(current_label, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        cond = dlg.collectConditions()
        category = MENU_TO_CATEGORY.get(self._current_menu_id)
        # 查询业务委托 activity 层执行（按当前分类 + 传入条件过滤）
        records = queryEmails(self.db, category, cond)
        # 用查询结果渲染列表页（标题后追加「搜索结果」说明当前处于检索模式）
        self.list_page.setTitle('%s（搜索结果 %d 条）' % (current_label, len(records)))
        self.list_page.showRecords(records)
        self.list_page.showTable(bool(records))
        # 处于检索模式时收藏线的「4:剩余暂存」对非已删除品类仍由 showRecords 归为 '-'
        self.stack.setCurrentWidget(self.list_page)

    # ---------------- 左栏 ----------------
    def buildLeftPane(self):
        """左栏：写信大按钮 + 导航列表（收件箱、星标、草稿箱、已发送等）"""
        pane = QFrame()
        pane.setFixedWidth(210)
        pane.setStyleSheet('QFrame{background:#f4f7fb; border-right:1px solid #e1e6ef;}'
                           'QPushButton#compose_btn{background:#1b7bf2; color:#fff;'
                           'border:0; padding:10px 16px; border-radius:6px; font-weight:bold;}'
                           'QPushButton#compose_btn:hover{background:#1567d0;}'
                           'QListWidget{background:transparent; border:0; padding:6px 4px;}'
                           'QListWidget::item{height:30px; padding-left:10px; border-radius:4px;}'
                           'QListWidget::item:selected{background:#dde9fb; color:#1b7bf2;'
                           'font-weight:bold;}')
        v = QVBoxLayout(pane)
        v.setContentsMargins(12, 16, 12, 12)
        v.setSpacing(8)

        # 写信按钮：点击跳到模式选择页（普通/高级）
        compose_btn = QPushButton('＋ 写 信')
        compose_btn.setObjectName('compose_btn')
        compose_btn.setMinimumHeight(40)
        compose_btn.clicked.connect(self.showComposeModePage)
        v.addWidget(compose_btn)

        v.addSpacing(6)

        # 导航菜单
        self.nav_list = QListWidget()
        for mid, mname in NAV_MENU_ITEMS:
            item = QListWidgetItem(mname)
            item.setData(Qt.ItemDataRole.UserRole, mid)
            self.nav_list.addItem(item)
        self.nav_list.setCurrentRow(0)
        self.nav_list.itemClicked.connect(self.onNavItemClicked)
        v.addWidget(self.nav_list, 1)

        v.addStretch()
        # ---- 窗口左下角：设置齿轮入口（点击打开铺满窗口的设置页）----
        gear_btn = QPushButton('⚙ 设置')
        gear_btn.setStyleSheet('QPushButton{background:transparent;color:#5f6b7a;'
                               'border:0;text-align:left;padding:6px 8px;}'
                               'QPushButton:hover{background:#e6efff;color:#1b7bf2;}')
        gear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        gear_btn.clicked.connect(self.openSettings)
        v.addWidget(gear_btn)
        return pane

    def openSettings(self):
        """打开铺满窗口的设置页（由左下角齿轮入口触发）"""
        SettingsDialog(self).exec()

    def setActiveNav(self, menu_id):
        """按 id 设定当前选中的菜单项（高亮并同步列表页标题/数据）"""
        self._current_menu_id = menu_id
        for i in range(self.nav_list.count()):
            item = self.nav_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == menu_id:
                self.nav_list.setCurrentRow(i)
                break
        self._refreshListPageFor(menu_id)

    def onNavItemClicked(self, item):
        """左栏菜单点击：切换到列表主页并按当前菜单刷新标题/占位数据"""
        menu_id = item.data(Qt.ItemDataRole.UserRole)
        self.setActiveNav(menu_id)
        # 点菜单应从写信页回到列表主页
        self.stack.setCurrentWidget(self.list_page)

    def _menuName(self, menu_id):
        """根据菜单 id 取中文名称（用于列表页标题显示）"""
        for mid, mname in NAV_MENU_ITEMS:
            if mid == menu_id:
                return mname
        return ''

    def _refreshListPageFor(self, menu_id):
        """刷新 MailListPage：从数据库读取当前分类的真实邮件并展示

        参数：
            menu_id<str>：左侧菜单 id（drafts/sent/deleted）
        """
        title = self._menuName(menu_id) or '列表'
        # 依据当前菜单控制列表工具条按钮显隐（已删除隐藏删除/再次编辑）
        self.list_page.setCurrentMenu(menu_id)
        if menu_id in LIST_PAGE_MENUS:
            # 从数据库读取该分类下全部记录（时间倒序）
            category = MENU_TO_CATEGORY[menu_id]
            # 正在查看「已删除」时顺带清理过期记录（超过保留期的自动永久删除），
            # 同时清理过期的已删除任务（含其下邮件）
            if category == CATEGORY_DELETED:
                self.db.deleteExpired(CATEGORY_DELETED, DELETE_RETENTION_DAYS)
                self._cleanupExpiredTasks()
            # 高级模式邮件以任务文件夹展示：独立邮件（task_id=0）与任务记录分别取
            records = [m for m in self.db.getMails(category) if not m.get('task_id')]
            tasks = self.db.getTasks(category)
            self.list_page.setTitle(title)
            self.list_page.showCategoryData(records, tasks)
            self.list_page.showTable(bool(records) or bool(tasks))
        else:
            # 非列表菜单（本项目暂无）：显示空白占位提示
            self.list_page.setTitle(title)
            self.list_page.showRecords([])
            self.list_page.showTable(False, '该分类暂无邮件（占位）。')

    # ---------------- 列表：删除 / 再次编辑 实装 ----------------
    def onRequestView(self, mail_id):
        """查看：双击进入只读查看（不提供编辑/发送/删除），已删除可恢复

        参数：
            mail_id<int>：待查看的邮件记录 id
        """
        record = self.db.getMail(mail_id)
        if not record:
            QMessageBox.warning(self, '提示', '记录不存在或已被删除。')
            return
        # 以只读模式回填发信页（fillFromRecord 会先清空并记录来源分类）
        self.normal_page.fillFromRecord(record)
        # 记录当前查看的记录 id（供「恢复」回调回传）
        self.normal_page.view_record_id = mail_id
        # 已删除：进入只读查看模式（显示恢复）；其余只读查看
        self.normal_page.setViewMode(restorable=record['category'] == CATEGORY_DELETED)
        self.stack.setCurrentWidget(self.normal_page)

    def onRequestEdit(self, mail_id):
        """再次编辑：从数据库取回邮件记录，以可编辑模式回填发信页

        草稿箱/已发送可再次编辑；已删除不允许编辑（无此入口）。
        参数：
            mail_id<int>：待编辑的邮件记录 id
        """
        record = self.db.getMail(mail_id)
        if not record:
            QMessageBox.warning(self, '提示', '记录不存在或已被删除。')
            return
        # 已删除的邮件只允许查看/恢复，不允许再次编辑、发送
        if record['category'] == CATEGORY_DELETED:
            QMessageBox.information(self, '提示', '已删除的邮件不支持再次编辑，可恢复后编辑。')
            return
        # 以可编辑模式回填发信页（再次编辑：可发送；草稿箱不提供存草稿）
        self.normal_page.fillFromRecord(record)
        self.normal_page.setEditMode()
        self.normal_page.rememberRecordId(mail_id)
        self.stack.setCurrentWidget(self.normal_page)

    def onRequestRestore(self, mail_id):
        """恢复：把「已删除」记录恢复回「草稿箱」

        参数：
            mail_id<int>：待恢复记录 id
        """
        if mail_id is None:
            return
        self.db.moveToCategory(mail_id, CATEGORY_DRAFT)
        # 刷新当前（已删除）列表并返回
        self._refreshListPageFor(self._current_menu_id)
        QMessageBox.information(self, '恢复成功', '邮件已恢复到草稿箱。')
        self.showListPage()

    def onRequestDelete(self, mail_ids, permanent=False):
        """删除：常规删除将记录移入「已删除」；彻底删除(永久)直接物理删除

        参数：
            mail_ids<list<int>>：待删除的记录 id 列表
            permanent<bool>：True=彻底删除(永久删除，物理删库)；False=移到已删除
        """
        if not mail_ids:
            return
        if permanent:
            # 彻底删除：物理删除数据库记录（不可恢复）
            self.db.deleteMails(mail_ids)
            self._refreshListPageFor(self._current_menu_id)
            QMessageBox.information(self, '彻底删除', '已永久删除 %d 封邮件。' % len(mail_ids))
            return
        # 常规删除：草稿/已发送 → 移入「已删除」
        for mid in mail_ids:
            self.db.moveToCategory(mid, CATEGORY_DELETED)
        # 刷新当前列表
        self._refreshListPageFor(self._current_menu_id)
        QMessageBox.information(self, '删除成功', '已删除 %d 封邮件。' % len(mail_ids))

    # ---------------- 高级模式任务：打开/详情/删除/恢复/返回 ----------------
    def onRequestOpenTask(self, task_id):
        """打开任务文件夹：双击/右键任务按分类路由
        - 草稿箱任务 → 恢复编辑（进入高级模式继续编辑）
        - 已发送/已删除任务 → 进入任务详情视图（列出该任务全部邮件）
        参数：
            task_id<int>：任务 id
        """
        task = self.db.getTask(task_id)
        if not task:
            QMessageBox.warning(self, '提示', '任务不存在或已被删除。')
            return
        if task['category'] == CATEGORY_DRAFT:
            self._restoreDraftTaskEditing(task_id)
        else:
            self._showTaskDetail(task_id)

    def _showTaskDetail(self, task_id):
        """进入任务详情视图：列出该任务下全部邮件（用于查看已发送/已删除任务）"""
        task = self.db.getTask(task_id)
        if not task:
            QMessageBox.warning(self, '提示', '任务不存在或已被删除。')
            return
        mails = self.db.getMailsByTask(task_id)
        self.list_page.showTaskDetail(task, mails)
        self.list_page.showTable(bool(mails), '该任务暂无邮件记录。')
        # 详情视图仍处于列表浏览场景：显示左侧菜单与顶部搜索
        self.left_pane.setVisible(True)
        self._setSearchBarVisible(True)
        self.stack.setCurrentWidget(self.list_page)

    def _restoreDraftTaskEditing(self, task_id):
        """恢复草稿任务为高级模式可编辑状态：还原全部配置/负载/预览，可继续编辑或发送"""
        task = self.db.getTask(task_id)
        if not task:
            QMessageBox.warning(self, '提示', '任务不存在或已被删除。')
            return
        # 还原序列化的全部配置（含输入控件/附件/负载绑定/预览邮件），并回到配置页
        self.high_page._restoreTaskConfig(task.get('config') or {})
        # 回填任务名，并记住当前编辑的任务 id（此后存草稿/发送更新到该任务而非新建）
        if hasattr(self.high_page, 'task_name_edit'):
            self.high_page.task_name_edit.setText(task.get('name') or '')
        self.high_page._current_task_id = task_id
        # 隐藏左侧菜单与顶部搜索，获得更大的高级模式编辑空间
        self.left_pane.setVisible(False)
        self._setSearchBarVisible(False)
        self.stack.setCurrentWidget(self.high_page)

    def onRequestTaskDelete(self, task_ids, permanent=False):
        """删除任务：常规删除移入「已删除」，彻底删除(永久)连带其下邮件物理删除
        参数：
            task_ids<list<int>>：任务 id 列表
            permanent<bool>：True=彻底删除；False=移入已删除
        """
        if not task_ids:
            return
        for tid in task_ids:
            if permanent:
                self.db.deleteTask(tid)
            else:
                self.db.moveTask(tid, CATEGORY_DELETED)
        self._refreshListPageFor(self._current_menu_id)
        verb = '永久删除' if permanent else '移入已删除'
        QMessageBox.information(self, '任务管理', '已%s %d 个高级模式任务。' % (verb, len(task_ids)))

    def onRequestTaskRestore(self, task_id):
        """恢复已删除任务回「草稿箱」（连带恢复其下全部邮件）"""
        if task_id is None:
            return
        self.db.moveTask(task_id, CATEGORY_DRAFT)
        QMessageBox.information(self, '恢复成功', '高级模式任务已恢复到草稿箱。')
        self.onTaskListBack()

    def onTaskListBack(self):
        """任务详情视图「返回列表」：重建当前分类视图并切回列表主页"""
        self._refreshListPageFor(self._current_menu_id)
        self.stack.setCurrentWidget(self.list_page)

    def _cleanupExpiredTasks(self):
        """清理「已删除」中保留超过 DELETE_RETENTION_DAYS 天的任务（连带其下邮件）
        在进入已删除列表页时调用，保证任务与邮件一致的 30 天清理机制。
        """
        now = datetime.now()
        for task in self.db.getTasks(CATEGORY_DELETED):
            # 以进入已删除时间(deleted_at)为准，无则退化用完成/保存时间(send_time)
            anchor = (task.get('deleted_at') or '').strip() or (task.get('send_time') or '').strip()
            if not anchor:
                continue
            try:
                anchor_dt = datetime.strptime(anchor, '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                continue
            if (now - anchor_dt).total_seconds() > DELETE_RETENTION_DAYS * 24 * 3600:
                self.db.deleteTask(task['id'])

    # ---------------- 堆叠页切换 ----------------
    def showListPage(self):
        """返回列表主页（发信完成或取消/返回后调用）

        返回时自动刷新当前菜单的列表数据，确保发送成功/存草稿等页面操作
        的改动立即可见，无需用户重新进入该分类。
        """
        # 恢复显示左侧菜单栏（从高级模式返回时）
        self.left_pane.setVisible(True)
        # 列表页为浏览/查询场景：显示顶部搜索框与查询按钮
        self._setSearchBarVisible(True)
        # 回到列表前先从数据库重读当前分类的数据（发送/存草稿后内容已变化）
        self._refreshListPageFor(self._current_menu_id)
        self.stack.setCurrentWidget(self.list_page)

    def showComposeModePage(self):
        """点「写信」→ 显示模式选择页（普通模式/高级模式）"""
        # 模式选择页可回到列表，保留搜索框
        self._setSearchBarVisible(True)
        self.stack.setCurrentWidget(self.mode_page)

    def showNormalPage(self):
        """选择模式 → 普通模式发信页（新建写信：重置为空白的写信状态）"""
        # 恢复显示左侧菜单栏（从高级模式切换时）
        self.left_pane.setVisible(True)
        # 发信页为写信场景：隐藏顶部搜索框与查询按钮，避免干扰写信
        self._setSearchBarVisible(False)
        # 新写信：先清空上次编辑残留（主题/正文/收件人等），再进入写信模式
        self.normal_page.clearForm()
        self.normal_page.setWriteMode()
        # 写信新邮件：清空已编辑记录标记，使「存草稿/发送」走新建逻辑
        self.normal_page.rememberRecordId(None)
        self.stack.setCurrentWidget(self.normal_page)

    def showHighModePlaceholder(self):
        """选择模式 → 高级模式页面，隐藏左侧菜单栏以获得更大编辑空间"""
        self.left_pane.setVisible(False)
        # 高级发信页同样隐藏顶部搜索框与查询按钮
        self._setSearchBarVisible(False)
        # 从模式选择页重新进入高级模式视为「开启新一轮发送任务」，
        # 重置负载初始化标志，使负载配给页在首次进入时重新构建空状态
        if hasattr(self.high_page, '_load_state_initialized'):
            self.high_page._load_state_initialized = False
        # 开启新任务：重置任务名（自增默认名称）并清空当前任务 id（当前编辑状态非草稿恢复而来）
        if hasattr(self.high_page, 'startNewTask'):
            self.high_page.startNewTask()
        self.stack.setCurrentWidget(self.high_page)

    def _setSearchBarVisible(self, visible):
        """统一控制主窗口顶部搜索框与查询按钮的显示/隐藏

        列表页与模式选择页需要搜索功能，显示其上的搜索框；
        进入发信页（普通/高级）后隐藏，避免与写信操作冲突。
        参数：
            visible<bool>：True 显示，False 隐藏
        """
        if not hasattr(self, 'search_edit'):
            return
        self.search_edit.setVisible(visible)
        if hasattr(self, 'search_btn'):
            self.search_btn.setVisible(visible)


class ComposeModePage(QWidget):
    """写信模式选择页：「普通模式 / 高级模式」两按钮

    用户在主窗口点「写信」先跳到本页，选完模式后才进入具体发信页。
    取消按钮可直接返回列表主页。
    """

    def __init__(self, on_click_normal, on_click_high, on_cancel):
        super().__init__()
        self.on_click_normal = on_click_normal
        self.on_click_high = on_click_high
        self.on_cancel = on_cancel
        self.buildUi()

    def buildUi(self):
        """居中布局：标题 + 两模式卡片大按钮 + 取消按钮"""
        outer = QVBoxLayout(self)
        outer.addStretch()
        wrap = QVBoxLayout()
        outer.addLayout(wrap)
        wrap.setSpacing(18)

        # 标题说明
        title = QLabel('选择写信模式')
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tf = title.font(); tf.setPointSize(18); tf.setBold(True); title.setFont(tf)
        tip = QLabel('普通模式：一对一常规写信。\n高级模式：一次性配置，向多人发送差异化内容与附件（开发中）。')
        tip.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tip.setStyleSheet('color:#666;')
        wrap.addWidget(title); wrap.addWidget(tip)

        # 两模式按钮行
        row = QHBoxLayout(); row.setSpacing(24)
        row.addStretch()
        normal_btn = QPushButton('普通模式')
        normal_btn.setMinimumSize(200, 96)
        normal_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;border:0;'
                                 'border-radius:8px;font-size:16px;font-weight:bold;}'
                                 'QPushButton:hover{background:#1567d0;}')
        normal_btn.clicked.connect(self.on_click_normal)
        high_btn = QPushButton('高级模式')
        high_btn.setMinimumSize(200, 96)
        high_btn.setStyleSheet('QPushButton{background:#fff;color:#1b7bf2;border:2px solid #1b7bf2;'
                               'border-radius:8px;font-size:16px;font-weight:bold;}'
                               'QPushButton:hover{background:#e6efff;}')
        high_btn.clicked.connect(self.on_click_high)
        row.addWidget(normal_btn); row.addWidget(high_btn)
        row.addStretch()
        wrap.addLayout(row)

        # 取消返回
        cancel_row = QHBoxLayout(); cancel_row.addStretch()
        cancel_btn = QPushButton('返回')
        cancel_btn.setMinimumWidth(120)
        cancel_btn.clicked.connect(self.on_cancel)
        cancel_row.addWidget(cancel_btn)
        cancel_row.addStretch()
        wrap.addLayout(cancel_row)
        outer.addStretch()


class HighModePlaceholderPage(QWidget):
    """高级模式占位页：UI 提示开发中，提供返回按钮

    本类只提供 UI；功能逻辑（批量差异化发送）后续补。
    """

    def __init__(self, on_back):
        super().__init__()
        self.on_back = on_back
        v = QVBoxLayout(self); v.addStretch()
        tip = QLabel('高级模式（批量差异化发送）\n开发中，敬请期待。')
        tip.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tf = tip.font(); tf.setPointSize(18); tip.setFont(tf)
        tip.setStyleSheet('color:#555;')
        v.addWidget(tip); v.addSpacing(24)
        row = QHBoxLayout(); row.addStretch()
        btn = QPushButton('返回邮件列表')
        btn.setMinimumWidth(140); btn.setMinimumHeight(36)
        btn.clicked.connect(self.on_back)
        row.addWidget(btn); row.addStretch(); v.addLayout(row); v.addStretch()


class MarqueeLabel(QWidget):
    """跑马灯提示条：显示一行循环滚动的提示文字（如已删除保留提示）

    用法：newMarquee = MarqueeLabel(widget) → setText(...) → 通过 show/hide 控制显隐。
    文字超过可视宽度时自动左移循环滚动；不足则静止居中显示。
    """

    # 滚动速度：每 40ms 移动 1 像素
    SCROLL_INTERVAL_MS = 40
    SCROLL_STEP_PX = 1

    def __init__(self, parent=None):
        super().__init__(parent)
        # 提示文字（存放待滚动内容）
        self.text = ''
        # 当前滚动位移（负数表示向左偏移）
        self.offset = 0
        # 是否处于滚动状态（文字宽度 > 可视宽度时为真）
        self.scrolling = False
        # 定时器：驱动滚动动画
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.advanceScroll)
        # 默认样式：浅黄提示底色 + 深灰文字
        self.setStyleSheet(
            'QWidget{background:#fff8e1; border:1px solid #ffe082; border-radius:4px;}'
        )

    def setText(self, text):
        """设置跑马灯内容并重算滚动状态

        参数：
            text<str>：要展示的提示文字
        """
        self.text = text or ''
        # 文字变化后重置滚动位置，便于重头开始
        self.offset = 0
        self._updateScrolling()
        self.update()

    def refreshMetrics(self):
        """在窗口大小变化时重算是否需要滚动（防止宽度变化后状态失效）"""
        self._updateScrolling()
        self.update()

    def _updateScrolling(self):
        """根据可视宽度与文字宽度判断是否开启滚动，并启停定时器"""
        fm = self.fontMetrics()
        text_width = fm.horizontalAdvance(self.text) if self.text else 0
        view_width = self.width()
        # 文字宽度大于可视宽度才需要滚动
        needs = text_width > view_width > 0
        if needs and not self.scrolling:
            self.scrolling = True
            self.timer.start(self.SCROLL_INTERVAL_MS)
        elif not needs and self.scrolling:
            self.scrolling = False
            self.timer.stop()
            self.offset = 0

    def advanceScroll(self):
        """定时器回调：每次向左推进一点，滚完全程后回到起点重来"""
        fm = self.fontMetrics()
        text_width = fm.horizontalAdvance(self.text)
        view_width = self.width()
        # 向左累计位移；到达「文字完全滚出（-text_width）」后复位
        self.offset -= self.SCROLL_STEP_PX
        min_offset = -text_width
        if self.offset <= min_offset:
            self.offset = 0
        # 为避免空白停顿，右端距可视区留一定间距
        _ = view_width
        self.update()

    def paintEvent(self, event):
        """自绘滚动文字：按当前 offset 平移绘制，超出可视区部分被裁剪"""
        painter = QPainter(self)
        # 定义文字颜色
        col = QColor('#8a6d3b')
        painter.setPen(col)
        fm = self.fontMetrics()
        view_width = self.width()
        height = self.height()
        text_width = fm.horizontalAdvance(self.text) if self.text else 0
        ty = (height - fm.ascent() - fm.descent()) / 2 + fm.ascent()
        if not self.scrolling:
            # 文字不超宽：居中静止显示
            tx = (view_width - text_width) / 2.0
        else:
            # 滚动：把文字沿 X 轴平移（负 offset 向左滚），再叠加一个周期的副本实现无缝循环
            tx = float(self.offset)
            painter.drawText(QRectF(tx + view_width, ty - fm.ascent(),
                                    text_width, fm.height()), self.text)
        # 绘制主文本（滚动时加副本实现无缝循环）
        painter.drawText(QRectF(tx, ty - fm.ascent(), text_width, fm.height()), self.text)
        painter.end()

    def resizeEvent(self, event):
        """窗口尺寸变化后重新评估滚动状态"""
        super().resizeEvent(event)
        self.refreshMetrics()


class MailListPage(QWidget):
    """邮件列表页：标题 + 工具条（写信 / 转发 / 更多 / 删除 / 再次编辑） + 列表表格

    本页只做 UI 展示与动作信号触发，不含真实数据查询/删除/再次编辑逻辑。
    工具栏/右键/列表均发出信号，由 MainWindow 接收并做占位弹窗。
    """

    def __init__(self, db, on_click_compose, on_click_view, on_click_edit, on_click_delete,
                 on_open_task=None, on_click_task_delete=None, on_click_task_restore=None,
                 on_back_task=None):
        super().__init__()
        # 数据库引用（用于删除/恢复时移动分类，列表展示数据由 MainWindow 从库读取后传入）
        self.db = db
        # 普通邮件相关回调：写信、查看、再次编辑、删除
        self.on_click_compose = on_click_compose
        self.on_click_view = on_click_view
        self.on_click_edit = on_click_edit
        self.on_click_delete = on_click_delete
        # 高级模式任务相关回调：打开任务(双击)、删除任务(移入已删除/彻底删除)、恢复任务、返回分类视图
        self.on_open_task = on_open_task
        self.on_click_task_delete = on_click_task_delete
        self.on_click_task_restore = on_click_task_restore
        self.on_back_task = on_back_task
        # 当前列表的行数据：每行一个 dict，形如 {'type':'mail','mail':rec} 或 {'type':'task','task':rec}
        self.rows = []
        # 当前所处的视图：None=分类视图（邮件+任务混排）；int=某任务详情视图（只显示该任务邮件）
        self.current_task_id = None
        # 任务详情视图的任务名（用于标题与返回
        self.current_task_name = ''
        self.buildUi()

    def buildUi(self):
        """列表页整体布局：标题 → 工具条 → 表格列表（或占位提示）"""
        v = QVBoxLayout(self)
        v.setContentsMargins(16, 12, 16, 16)
        v.setSpacing(8)

        # 1) 跑马灯提示条（仅「已删除」页面显示，提示保留 30 天自动清理）
        self.marquee = MarqueeLabel(self)
        self.marquee.setText('已删除邮件可在本页暂存 30 天，30 天到期后将自动清理，请及时处理。')
        self.marquee.setFixedHeight(26)
        # 默认隐藏（setCurrentMenu 根据菜单决定是否显示）
        self.marquee.hide()
        v.addWidget(self.marquee)

        # 2) 标题 + 统计文字
        self.title_label = QLabel('草稿箱')
        tf = self.title_label.font(); tf.setPointSize(16); tf.setBold(True)
        self.title_label.setFont(tf)
        self.count_label = QLabel('共 0 封邮件')
        self.count_label.setStyleSheet('color:#666;')
        title_row = QHBoxLayout()
        # 返回按钮：仅在「任务详情视图」显示，点击回到分类视图
        self.task_back_btn = QPushButton('‹ 返回列表')
        self.task_back_btn.setStyleSheet('QPushButton{background:#e6efff;color:#1b7bf2;border:0;'
                                         'padding:4px 12px;border-radius:4px;font-weight:bold;}'
                                         'QPushButton:hover{background:#d5e6ff;}')
        self.task_back_btn.clicked.connect(self._onBackTask)
        self.task_back_btn.hide()
        title_row.addWidget(self.task_back_btn)
        title_row.addWidget(self.title_label); title_row.addSpacing(8)
        title_row.addWidget(self.count_label); title_row.addStretch()
        v.addLayout(title_row)

        # 2) 工具条：写信/转发/更多/删除/再次编辑
        v.addWidget(self.buildToolbar())

        # 3) 表格 / 占位提示容器
        self.table_container = QStackedWidget()
        v.addWidget(self.table_container, 1)

        # A：邮件表格（0 复选框 / 1 主题 / 2 收件人 / 3 时间 / 4 剩余暂存(仅已删除)）
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(['', '主题', '收件人/发件人', '时间', '剩余暂存'])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 32)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(2, 320)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(3, 160)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(4, 150)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.onTableContextMenu)
        self.table.doubleClicked.connect(self.onTableDoubleClickEdit)
        self.table_container.addWidget(self.table)

        # B：占位提示
        self.placeholder = QLabel('该分类暂无邮件（占位）。')
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setStyleSheet('color:#888; font-size:14px;')
        self.table_container.addWidget(self.placeholder)

    def buildToolbar(self):
        """构建列表工具条（写信/转发/更多/删除/再次编辑）"""
        tb = QFrame()
        tb.setStyleSheet('QFrame{background:#fff; border-bottom:1px solid #e1e6ef;}'
                         'QPushButton{padding:5px 14px; border:1px solid #d0d7e2;'
                         'border-radius:4px; background:#fff;}'
                         'QPushButton:hover{background:#f0f5fd;}')
        h = QHBoxLayout(tb)
        h.setContentsMargins(4, 6, 4, 6)
        h.setSpacing(6)

        def mk_button(text, slot, color='#333', bg='#fff'):
            b = QPushButton(text)
            if bg != '#fff' or color != '#333':
                b.setStyleSheet('QPushButton{background:%s;color:%s;border:0;border-radius:4px;'
                                'padding:5px 14px;} QPushButton:hover{opacity:0.9;}' % (bg, color))
            b.clicked.connect(slot)
            return b

        h.addWidget(mk_button('写 信',      self.on_click_compose, '#fff', '#1b7bf2'))
        # 转发/更多按钮存为成员：只在非「已删除」页面显示（已删除页面改为「彻底删除」）
        self.forward_tool_btn = mk_button('转 发(占)', self._placeholderForward)
        h.addWidget(self.forward_tool_btn)
        self.more_tool_btn = mk_button('更 多(占)', self._placeholderMore)
        h.addWidget(self.more_tool_btn)
        h.addSpacing(12)
        # 删除按钮存为成员，供已删除分类下隐藏（已删除只查看+恢复+彻底删除）
        self.delete_tool_btn = mk_button('删 除', self.onToolbarDelete, '#c0392b', '#fdecea')
        h.addWidget(self.delete_tool_btn)
        # 再次编辑按钮存为成员，供已删除分类下隐藏（已删除不支持再次编辑）
        self.edit_tool_btn = mk_button('再次编辑', self.onToolbarEdit, '#1b7bf2', '#e6efff')
        h.addWidget(self.edit_tool_btn)
        # 彻底删除按钮（永久删除）：仅已删除页面显示，删除前弹窗确认
        self.permanent_delete_btn = mk_button('彻底删除', self.onPermanentDelete, '#c0392b', '#fdecea')
        h.addWidget(self.permanent_delete_btn)
        h.addStretch()
        return tb

    def setCurrentMenu(self, menu_id):
        """记录当前菜单并控制工具条按钮显隐：
        已删除→隐藏 转发/更多/删除/再次编辑，显示「彻底删除」；其余反向。
        参数：
            menu_id<str>：当前左侧菜单 id
        """
        # 判断是否已删除页面
        is_deleted = (menu_id == 'deleted')
        # 写信总是显示；转发/更多在已删除隐藏
        self.forward_tool_btn.setVisible(not is_deleted)
        self.more_tool_btn.setVisible(not is_deleted)
        # 常规「删除」「再次编辑」在已删除隐藏
        self.edit_tool_btn.setVisible(not is_deleted)
        self.delete_tool_btn.setVisible(not is_deleted)
        # 「彻底删除」仅已删除显示
        self.permanent_delete_btn.setVisible(is_deleted)
        # 跑马灯提示条：仅已删除页面显示（提示 30 天自动清理）
        self.marquee.setVisible(is_deleted)

    def setTitle(self, title):
        """设置列表页顶部标题（由 MainWindow 在切换菜单时调用）"""
        self.title_label.setText(title)

    def showRecords(self, records):
        """用数据库邮件记录填充表格（仅邮件，不含任务；供搜索/普通场景展示）

        参数：
            records<list<dict>>：一条条邮件记录，需含 id/title/recipient/send_time
        """
        # 组装为统一的行结构（全部为邮件行），进入分类视图
        self.current_task_id = None
        self.task_back_btn.hide()
        self.rows = [{'type': 'mail', 'mail': rec} for rec in records]
        self._renderRows()

    def showCategoryData(self, mails, tasks):
        """用「独立邮件 + 任务文件夹」混排填充分类视图（草稿箱/已发送/已删除通用）

        高级模式产生的邮件以「任务文件夹」形式展示，避免邮件过多造成视觉阻碍；
        普通邮件直接展示为邮件行。两者按保存时间倒序混排。

        参数：
            mails<list<dict>>：该分类下独立（非任务内）的邮件记录
            tasks<list<dict>>：该分类下的高级模式任务记录
        """
        self.current_task_id = None
        self.task_back_btn.hide()
        self.rows = []
        # 普通邮件行
        for rec in mails:
            self.rows.append({'type': 'mail', 'mail': rec})
        # 高级模式任务文件夹行（展开后显示为该任务内全部邮件）
        for t in tasks:
            self.rows.append({'type': 'task', 'task': t})
        # 按保存/发送时间倒序混排（无时间者排后）
        self.rows.sort(key=lambda r: r['mail'].get('send_time', '')
                       if r['type'] == 'mail' else r['task'].get('send_time', ''),
                       reverse=True)
        self._renderRows()

    def showTaskDetail(self, task, mails):
        """进入某任务详情视图：标题显示任务名，表格只列出该任务下全部邮件

        参数：
            task<dict>：任务记录（含 id/name/mail_count/send_time）
            mails<list<dict>>：该任务下的邮件记录
        """
        self.current_task_id = task.get('id')
        self.current_task_name = task.get('name', '')
        # 显示返回按钮 + 任务名标题
        self.task_back_btn.show()
        self.title_label.setText(task.get('name', '高级模式任务'))
        self.rows = [{'type': 'mail', 'mail': rec} for rec in mails]
        self._renderRows()

    def _onBackTask(self):
        """「返回列表」点击：回调 MainWindow 重建分类视图"""
        if self.on_back_task:
            self.on_back_task()
        else:
            # 兜底：无回调时清掉任务详情状态并清空
            self.current_task_id = None
            self.task_back_btn.hide()
            self.rows = []
            self._renderRows()

    def _renderRows(self):
        """按 self.rows 渲染表格：支持邮件行与任务文件夹行两种类型

        邮件行：0 复选框 / 1 主题 / 2 收件人 / 3 时间 / 4 剩余暂存
        任务文件夹行：1「📁 任务名」/ 2「xx 封邮件」/ 3 时间 / 4 剩余暂存（已删除任务）
        """
        self.table.setRowCount(len(self.rows))
        for r_idx, row in enumerate(self.rows):
            # 第 0 列：复选框（用于批量删除/彻底删除）
            check_item = QTableWidgetItem()
            check_item.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            check_item.setCheckState(Qt.CheckState.Unchecked)
            self.table.setItem(r_idx, 0, check_item)

            if row['type'] == 'task':
                # 任务文件夹行：列 1 显示 📁任务名，列 2 显示邮件数，列 3 显示时间
                task = row['task']
                name_item = QTableWidgetItem('📁 ' + task.get('name', '高级模式任务'))
                f = name_item.font(); f.setBold(True); name_item.setFont(f)
                name_item.setForeground(QColor('#1b7bf2'))
                self.table.setItem(r_idx, 1, name_item)
                self.table.setItem(r_idx, 2, QTableWidgetItem('%d 封邮件' % task.get('mail_count', 0)))
                self.table.setItem(r_idx, 3, QTableWidgetItem(str(task.get('send_time', ''))))
                # 已删除任务行显示剩余暂存（复用邮件的时间计算）
                is_del = task.get('category') == CATEGORY_DELETED
                remain_text = self.formatRemainText(task) if is_del else '-'
                remain_item = QTableWidgetItem(remain_text)
                remain_item.setForeground(QColor('#c0392b') if is_del and self._isAlmostGone(task)
                                          else QColor('#333'))
                self.table.setItem(r_idx, 4, remain_item)
                continue

            # 邮件行（与历史逻辑一致）
            rec = row['mail']
            # 1:主题 2:收件人 3:时间
            row_vals = (rec.get('title', ''), rec.get('recipient', ''), rec.get('send_time', ''))
            for c_idx, val in enumerate(row_vals, start=1):
                item = QTableWidgetItem(str(val))
                if c_idx == 1:
                    f = item.font(); f.setBold(True); item.setFont(f)
                self.table.setItem(r_idx, c_idx, item)
            # 4:剩余暂存（仅已删除记录展示）
            is_deleted_rec = rec.get('category') == CATEGORY_DELETED
            remain_text = self.formatRemainText(rec) if is_deleted_rec else '-'
            remain_item = QTableWidgetItem(remain_text)
            remain_item.setForeground(QColor('#c0392b') if is_deleted_rec and self._isAlmostGone(rec) else QColor('#333'))
            self.table.setItem(r_idx, 4, remain_item)
        # 统计：任务详情视图显示「共 N 封邮件」；分类视图同时统计任务数量
        if self.current_task_id is not None:
            self.count_label.setText('共 %d 封邮件' % len(self.rows))
        else:
            mail_cnt = sum(1 for r in self.rows if r['type'] == 'mail')
            task_cnt = sum(1 for r in self.rows if r['type'] == 'task')
            base = '共 %d 封邮件' % mail_cnt
            if task_cnt:
                base += '，%d 个高级任务' % task_cnt
            self.count_label.setText(base)

    def formatRemainText(self, rec):
        """计算单条邮件「剩余暂存」的可读文本（基于进入已删除的时间）

        以 deleted_at（进入已删除的时间，精确到分、秒为 00）加保留天数得到清理时刻，
        与当前时间相减得到剩余时长（精确到分，秒显示为 :00）。

        参数：
            rec<dict>：单条邮件记录（需含 deleted_at / send_time）

        返回：
            str：如「2天 05:00:00」；无时间或异常返回「-」
        """
        try:
            # 优先用进入已删除的时间；旧数据无 deleted_at 时退化为发送时间
            anchor = (rec.get('deleted_at') or '').strip() or (rec.get('send_time') or '').strip()
            if not anchor:
                return '-'
            # 解析进入时间（兼容到秒/到分两种格式）
            anchor_dt = datetime.strptime(anchor, '%Y-%m-%d %H:%M:%S')
            clear_dt = anchor_dt + timedelta(days=DELETE_RETENTION_DAYS)
            now = datetime.now()
            remain = clear_dt - now
            # 已到期或已过期：提示立即清理
            if remain.total_seconds() <= 0:
                return '已到期，将自动清理'
            # 剩余时长精确到分：去掉秒，秒显示为 00
            total_min = int(remain.total_seconds()) // 60
            days = total_min // (24 * 60)
            hours = (total_min % (24 * 60)) // 60
            minutes = total_min % 60
            return '%d天 %02d:%02d:00' % (days, hours, minutes)
        except (ValueError, TypeError):
            return '-'

    def _isAlmostGone(self, rec):
        """判断某邮件是否临近清理（剩余不足 1 天，用于暖色高亮）

        参数：
            rec<dict>：单条邮件记录

        返回：
            bool：True 表示剩余不足 1 天，即将被自动清理
        """
        try:
            anchor = (rec.get('deleted_at') or '').strip() or (rec.get('send_time') or '').strip()
            if not anchor:
                return False
            anchor_dt = datetime.strptime(anchor, '%Y-%m-%d %H:%M:%S')
            remain = (anchor_dt + timedelta(days=DELETE_RETENTION_DAYS)) - datetime.now()
            return 0 < remain.total_seconds() <= 24 * 3600
        except (ValueError, TypeError):
            return False

    def showTable(self, show, placeholder_text='该分类暂无邮件（占位）。'):
        """切换显示：表格 / 占位提示"""
        if show:
            self.table_container.setCurrentWidget(self.table)
        else:
            self.placeholder.setText(placeholder_text)
            self.table_container.setCurrentWidget(self.placeholder)

    # --------- UI 动作（删除 / 再次编辑 / 右键菜单 / 占位） ---------
    def _checkedRowIndices(self):
        """收集勾选复选框或行选中的行号集合（用于删除时汇总待处理行）"""
        indices = set()
        for r_idx in range(self.table.rowCount()):
            chk = self.table.item(r_idx, 0)
            if chk is not None and chk.checkState() == Qt.CheckState.Checked:
                indices.add(r_idx)
        # 若未勾选任何行，退回使用选中行
        if not indices:
            for idx in self.table.selectionModel().selectedRows():
                indices.add(idx.row())
        return sorted(indices)

    def _splitRows(self, indices):
        """把选中行号拆分为 邮件id列表 与 任务id列表（删除/恢复前收集）

        参数：
            indices<list<int>>：行号列表（来自复选框或行选中）

        返回：
            mail_ids<list<int>>：邮件记录 id
            task_ids<list<int>>：高级模式任务 id；首个任务 id 额外返回用于单任务操作
        """
        mail_ids, task_ids = [], []
        first_task_id = None
        for i in indices:
            if i >= len(self.rows):
                continue
            row = self.rows[i]
            if row['type'] == 'mail':
                mail_ids.append(row['mail'].get('id'))
            else:
                tid = row['task'].get('id')
                task_ids.append(tid)
                if first_task_id is None:
                    first_task_id = tid
        return mail_ids, task_ids, first_task_id

    def _rowDisplayText(self, i):
        """取某行的展示名（邮件用主题，任务用任务名），供删除确认提示使用

        参数：
            i<int>：行号
        返回：
            str：展示文本；行号越界返回空串
        """
        if i >= len(self.rows):
            return ''
        row = self.rows[i]
        if row['type'] == 'mail':
            return row['mail'].get('title', '')
        return row['task'].get('name', '高级模式任务')

    def onPermanentDelete(self):
        """彻底删除：永久删除选中邮件/任务；删除前弹窗二次确认（不可恢复）"""
        indices = self._checkedRowIndices()
        if not indices:
            QMessageBox.information(self, '提示', '请先勾选或选中需要彻底删除的邮件。')
            return
        mail_ids, task_ids, _ = self._splitRows(indices)
        if not mail_ids and not task_ids:
            return
        first_text = self._rowDisplayText(indices[0])
        n = len(mail_ids) + len(task_ids)
        tip = '将永久删除 %d 项' % n if n > 1 else '将永久删除该项'
        if first_text:
            tip += '\n「%s」' % first_text
        tip += '\n\n永久删除后不可恢复，确定继续吗？'
        result = QMessageBox.question(
            self, '彻底删除确认', tip,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if result != QMessageBox.StandardButton.Yes:
            return
        # 邮件与任务分别回调 MainWindow 物理删除并刷新
        if mail_ids:
            self.on_click_delete(mail_ids, permanent=True)
        if task_ids:
            self.on_click_task_delete(task_ids, permanent=True)

    def onToolbarDelete(self):
        """工具条删除：把选中行拆分为邮件与任务，分别移入「已删除」"""
        indices = self._checkedRowIndices()
        if not indices:
            QMessageBox.information(self, '提示', '请先勾选或选中需要删除的行。')
            return
        mail_ids, task_ids, _ = self._splitRows(indices)
        if mail_ids:
            self.on_click_delete(mail_ids)
        if task_ids:
            self.on_click_task_delete(task_ids, permanent=False)

    def onToolbarEdit(self):
        """工具条再次编辑：邮件→再次编辑；任务→按分类处理（草稿=恢复编辑，已发送/已删除=进入详情）"""
        indices = self._checkedRowIndices()
        if not indices:
            sel = self.table.currentRow()
            indices = [sel] if sel >= 0 else []
        if not indices:
            QMessageBox.information(self, '提示', '请先选中一行。')
            return
        first = indices[0]
        if first >= len(self.rows):
            return
        row = self.rows[first]
        if row['type'] == 'task':
            if self.on_open_task:
                self.on_open_task(row['task'].get('id'))
            return
        # 邮件行：回传记录 id 发起再次编辑
        if self.on_click_edit:
            self.on_click_edit(row['mail'].get('id'))

    def onTableContextMenu(self, pos):
        """列表右键菜单：邮件→再次编辑/删除；任务→按分类提供 打开/恢复编辑/删除/彻底删除/恢复"""
        item = self.table.itemAt(pos)
        if item is None:
            return
        row = item.row()
        if row >= len(self.rows):
            return
        row_data = self.rows[row]
        menu = QMenu(self)

        if row_data['type'] == 'task':
            # 任务文件夹右键：草稿=「恢复编辑」，已删除=「进入查看/恢复/彻底删除」，已发送=「查看任务邮件」
            task = row_data['task']
            cat = task.get('category')
            tid = task.get('id')
            act_delete = act_restore = act_perm = None
            act_open = QAction('恢复编辑（进入高级模式）' if cat == CATEGORY_DRAFT else '查看任务邮件', self)
            menu.addAction(act_open)
            if cat == CATEGORY_DELETED:
                act_restore = QAction('恢复', self)
                act_perm = QAction('彻底删除', self)
                menu.addAction(act_restore)
                menu.addAction(act_perm)
            else:
                act_delete = QAction('删除', self)
                menu.addAction(act_delete)
            chosen = menu.exec(self.table.viewport().mapToGlobal(pos))
            if chosen is None:
                return
            if chosen == act_open and self.on_open_task:
                self.on_open_task(tid)
            elif chosen == act_delete and self.on_click_task_delete:
                self.on_click_task_delete([tid], permanent=False)
            elif chosen == act_restore and self.on_click_task_restore:
                self.on_click_task_restore(tid)
            elif chosen == act_perm and self.on_click_task_delete:
                self.on_click_task_delete([tid], permanent=True)
            return

        # 邮件行右键（任务详情视图内同样适用）
        mail_id = row_data['mail'].get('id')
        act_edit = QAction('再次编辑', self)
        act_delete = QAction('删除', self)
        menu.addAction(act_edit)
        menu.addAction(act_delete)
        chosen = menu.exec(self.table.viewport().mapToGlobal(pos))
        if chosen is None:
            return
        if chosen == act_edit and self.on_click_edit:
            self.on_click_edit(mail_id)
        elif chosen == act_delete and self.on_click_delete:
            self.on_click_delete([mail_id])

    def onTableDoubleClickEdit(self, model_index):
        """列表双击：邮件→只读查看；任务文件夹→由 MainWindow 按分类路由（草稿恢复编辑/其余进详情）"""
        row = model_index.row()
        if row >= len(self.rows):
            return
        row_data = self.rows[row]
        if row_data['type'] == 'task':
            if self.on_open_task:
                self.on_open_task(row_data['task'].get('id'))
            return
        # 邮件行 → 查看
        if self.on_click_view:
            self.on_click_view(row_data['mail'].get('id'))

    @staticmethod
    def _placeholderForward():
        """转发按钮（占位）：不实现功能逻辑，仅弹窗"""
        QMessageBox.information(None, '提示', '转发功能开发中。')

    @staticmethod
    def _placeholderMore():
        """更多按钮（占位）：不实现功能逻辑，仅弹窗"""
        QMessageBox.information(None, '提示', '更多功能开发中。')


class RecipientPickerDialog(QDialog):
    """收件人选择器（群发功能）：选择收件人 Excel/TXT 文件并让用户勾选收件人

    UI 能力：
      - 「选择文件」按钮弹出文件对话框（*.xlsx / *.xls / *.txt）
      - 收件人以可勾选的列表展示，支持 全选 / 取消全选
      - 按首字母排序（A-Z，忽略大小写）
      - 关键字搜索：输入字符即时筛选出包含该字符的邮箱
      - 底部显示「已勾选 N / 共 M」
      - 确定后返回勾选的收件人邮箱列表

    说明：文件识别、排序、过滤等业务逻辑均收敛在 activity.recipient_bulk，
          本类仅负责界面交互与调用。
    """

    def __init__(self, parent=None):
        """初始化收件人选择器对话框"""
        super().__init__(parent)
        self.setWindowTitle('收件人选择（群发）')
        self.resize(420, 520)
        # 当前已识别出的全部收件人列表（字符串邮箱，保持解析顺序）
        self.all_recipients = []
        # 已勾选的邮箱集合（跨排序/筛选切换时保持勾选状态）
        self.selected_emails = set()
        # 当前排序开关：True=按首字母；False=原始顺序
        self.sort_by_letter = False

        v = QVBoxLayout(self)
        v.setSpacing(8)

        # 顶部说明 + 「选择文件」按钮
        top = QHBoxLayout()
        tip = QLabel('选择收件人 Excel 或 TXT 文件（文件内只含收件人邮箱）')
        tip.setStyleSheet('color:#666;')
        top.addWidget(tip)
        top.addStretch()
        pick_btn = QPushButton('选择文件')
        pick_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;border:0;'
                               'padding:5px 14px;border-radius:4px;}')
        pick_btn.clicked.connect(self.onPickFile)
        top.addWidget(pick_btn)
        v.addLayout(top)

        # 排序 + 搜索工具栏
        tool_row = QHBoxLayout()
        sort_label = QLabel('排序:')
        tool_row.addWidget(sort_label)
        self.sort_combo = QComboBox()
        self.sort_combo.addItem('原始顺序', False)
        self.sort_combo.addItem('按首字母 A-Z', True)
        self.sort_combo.currentIndexChanged.connect(self._onSortChanged)
        tool_row.addWidget(self.sort_combo)
        tool_row.addSpacing(8)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText('输入字符筛选邮箱…')
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.textChanged.connect(self._onSearchChanged)
        tool_row.addWidget(self.search_edit, 1)
        v.addLayout(tool_row)

        # 收件人可勾选列表
        self.recipient_list = QListWidget()
        self.recipient_list.itemChanged.connect(self._onItemToggled)
        v.addWidget(self.recipient_list)

        # 全选 / 取消全选
        op_row = QHBoxLayout()
        self.select_all_btn = QPushButton('全选')
        self.select_all_btn.clicked.connect(lambda: self._setAllChecked(True))
        self.clear_all_btn = QPushButton('取消全选')
        self.clear_all_btn.clicked.connect(lambda: self._setAllChecked(False))
        op_row.addWidget(self.select_all_btn)
        op_row.addWidget(self.clear_all_btn)
        op_row.addStretch()
        self.count_label = QLabel('已勾选 0 / 共 0')
        op_row.addWidget(self.count_label)
        v.addLayout(op_row)

        # 确定 / 取消
        bottom = QHBoxLayout()
        bottom.addStretch()
        confirm_btn = QPushButton('确定')
        confirm_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;border:0;'
                                  'padding:6px 20px;border-radius:4px;}')
        confirm_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(confirm_btn)
        bottom.addWidget(cancel_btn)
        v.addLayout(bottom)

        # 启用列表为空时的按钮态保护
        self._updateButtons()
        self._updateCount()

    # ---------------- 文件选择（解析逻辑为占位） ----------------

    def onPickFile(self):
        """弹出文件对话框选择收件人文件（Excel/TXT），随后解析并展示收件人

        解析由 activity.recipient_bulk 提供，识别出的收件人填充为可勾选列表。
        """
        path, _ = QFileDialog.getOpenFileName(
            self, '选择收件人文件', '',
            '收件人文件 (*.xlsx *.xls *.txt);;所有文件 (*)')
        if not path:
            return
        # 解析收件人（业务逻辑在 activity 层）；文件类型不支持、解析依赖缺失等均在
        # 此处兜底捕获并提示，避免在 modal 对话框事件循环里抛裸异常导致程序崩溃
        try:
            recipients = self._loadRecipientsFromFile(path)
        except (ValueError, ImportError) as exc:
            QMessageBox.warning(self, '解析失败', str(exc))
            return
        self.setRecipients(recipients)

    def _loadRecipientsFromFile(self, path):
        """从文件读取并识别收件人（业务逻辑委托给 activity.recipient_bulk 实现）

        参数：
            path<str>：选中的收件人文件路径（xlsx/xls/txt）

        返回：
            list<str>：识别出的收件人邮箱列表（去重、保持顺序）
        """
        # 委托 activity 层解析；文件类型不支持等异常由调用方（onPickFile）统一处理
        return parseRecipientFile(path)

    def setRecipients(self, recipients):
        """用识别出的收件人填充列表（默认全部勾选），并重置排序/搜索为原始

        参数：
            recipients<list<str>>：收件人邮箱列表
        """
        self.all_recipients = list(recipients or [])
        # 初次载入：默认全选（清空旧的勾选记录并全部勾选）
        self.selected_emails = set(self.all_recipients)
        # 重置排序与搜索状态
        self.sort_by_letter = False
        self._setComboSilent(0)
        self.search_edit.blockSignals(True)
        self.search_edit.clear()
        self.search_edit.blockSignals(False)
        self._renderList()

    def _setComboSilent(self, index):
        """不触发回调地切换排序下拉的选中项

        参数：
            index<int>：目标下拉索引
        """
        self.sort_combo.blockSignals(True)
        self.sort_combo.setCurrentIndex(index)
        self.sort_combo.blockSignals(False)

    # ---------------- 排序 / 搜索 ----------------

    def _onSortChanged(self, index):
        """排序下拉变化：记录排序开关后按当前状态重渲染列表"""
        self.sort_by_letter = bool(self.sort_combo.itemData(index))
        self._renderList()

    def _onSearchChanged(self, text):
        """搜索框文本变化：按关键字即时筛选并重渲染列表"""
        self._renderList()

    def _renderList(self):
        """按 排序 + 搜索 当前状态重渲染可勾选列表（保持已勾选状态）

        数据流：all_recipients →(activity.filterRecipients 按关键字)→(activity.sortRecipients 按字母)→ 展示
        """
        # 屏蔽 itemChanged 避免重填过程触发计数刷新（计数在填完后统一更新）
        self.recipient_list.itemChanged.disconnect(self._onItemToggled)
        self.recipient_list.clear()

        # 先按关键字过滤（业务逻辑在 activity），再按需排序（业务逻辑在 activity）
        filtered = filterRecipients(self.all_recipients, self.search_edit.text())
        ordered = sortRecipients(filtered, self.sort_by_letter)
        for email in ordered:
            item = QListWidgetItem(email)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(
                Qt.CheckState.Checked if email in self.selected_emails
                else Qt.CheckState.Unchecked)
            self._addRecipientItem(item)

        self.recipient_list.itemChanged.connect(self._onItemToggled)
        self._updateCount()
        self._updateButtons()

    def _addRecipientItem(self, item):
        """把单个收件人条目加入列表（供不同重填路径复用）"""
        self.recipient_list.addItem(item)

    # ---------------- 勾选状态管理 ----------------

    def _onItemToggled(self, item):
        """列表项勾选状态变化：同步到勾选集合并刷新计数"""
        email = item.text()
        if item.checkState() == Qt.CheckState.Checked:
            self.selected_emails.add(email)
        else:
            self.selected_emails.discard(email)
        self._updateCount()

    def _setAllChecked(self, checked):
        """全选或取消全选（更新勾选集合并对当前列表逐项设置勾选态）
        注意：仅作用于当前可见（筛选后）的列表项。

        参数：
            checked<bool>：True 全选，False 取消全选
        """
        for i in range(self.recipient_list.count()):
            item = self.recipient_list.item(i)
            if checked:
                self.selected_emails.add(item.text())
                item.setCheckState(Qt.CheckState.Checked)
            else:
                self.selected_emails.discard(item.text())
                item.setCheckState(Qt.CheckState.Unchecked)
        self._updateCount()

    def _checkedRecipients(self):
        """返回当前全部已勾选收件人（所有已勾选集合，含被筛选隐藏的）"""
        return sorted(self.selected_emails, key=lambda e: e.lower())

    def getSelectedRecipients(self):
        """供外部读取最终勾选的收件人（确定后调用）

        返回：
            list<str>：勾选的收件人邮箱列表
        """
        return self._checkedRecipients()

    def _updateCount(self):
        """刷新「已勾选 / 共」计数文案

        显示值：已勾选数 / 当前(筛选后可见)列表条数，便于用户感知筛选效果。
        """
        total = self.recipient_list.count()
        checked = len(self.selected_emails)
        self.count_label.setText('已勾选 %d / 共 %d' % (checked, total))

    def _updateButtons(self):
        """按列表是否为空启用/禁用 全选 与 取消全选 按钮"""
        has_items = self.recipient_list.count() > 0
        self.select_all_btn.setEnabled(has_items)
        self.clear_all_btn.setEnabled(has_items)


class RecipientManageDialog(QDialog):
    """查看/管理所有收件人对话框：应对收件人过多时输入框显示不全的问题

    与输入框维护「同一个」收件人集合：
      - 打开时从输入框文本即时拆分（activity.splitRecipients），保证列表与输入框一致；
      - 支持排序、搜索、全选/取消、逐项勾选取消；
      - 非法片段（用户手误删除字符导致）在列表中标红展示；
      - 确定后把仍勾选的结果回写输入框（activity 层拼接），从而：
          输入框里删掉的内容不会残留到列表；列表里取消勾选的内容也不会残留到输入框。

    业务逻辑（拆分/校验/排序/过滤/拼接）均收敛在 activity.recipient_bulk。
    """

    def __init__(self, parent=None):
        """初始化查看/管理收件人对话框"""
        super().__init__(parent)
        self.setWindowTitle('查看所有收件人')
        self.resize(440, 520)
        # 输入框拆分出的全部片段（含非法项，保持顺序）
        self.all_tokens = []
        # 非法片段集合（标红用）：手误导致的不完整邮箱
        self.invalid_set = set()
        # 仍勾选的片段集合（跨排序/筛选保持）
        self.selected_items = set()

        v = QVBoxLayout(self)
        v.setSpacing(8)

        # 顶部提示
        tip = QLabel('以下为收件人输入框中的所有收件人，可勾选/取消，确定后回写输入框。')
        tip.setStyleSheet('color:#666;')
        tip.setWordWrap(True)
        v.addWidget(tip)

        # 排序 + 搜索工具栏
        tool_row = QHBoxLayout()
        self.sort_combo = QComboBox()
        self.sort_combo.addItem('原始顺序', False)
        self.sort_combo.addItem('按首字母 A-Z', True)
        self.sort_combo.currentIndexChanged.connect(self._renderList)
        tool_row.addWidget(QLabel('排序:'))
        tool_row.addWidget(self.sort_combo)
        tool_row.addSpacing(8)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText('输入字符筛选邮箱…')
        self.search_edit.setClearButtonEnabled(True)
        self.search_edit.textChanged.connect(self._renderList)
        tool_row.addWidget(self.search_edit, 1)
        v.addLayout(tool_row)

        # 可勾选列表
        self.recipient_list = QListWidget()
        self.recipient_list.itemChanged.connect(self._onItemToggled)
        v.addWidget(self.recipient_list)

        # 全选 / 取消全选
        op_row = QHBoxLayout()
        self.select_all_btn = QPushButton('全选')
        self.select_all_btn.clicked.connect(lambda: self._setAllChecked(True))
        self.clear_all_btn = QPushButton('取消全选')
        self.clear_all_btn.clicked.connect(lambda: self._setAllChecked(False))
        op_row.addWidget(self.select_all_btn)
        op_row.addWidget(self.clear_all_btn)
        op_row.addStretch()
        self.count_label = QLabel('已勾选 0 / 共 0')
        op_row.addWidget(self.count_label)
        v.addLayout(op_row)

        # 确定 / 取消
        bottom = QHBoxLayout()
        bottom.addStretch()
        confirm_btn = QPushButton('确定')
        confirm_btn.setStyleSheet('QPushButton{background:#1b7bf2;color:#fff;border:0;'
                                  'padding:6px 20px;border-radius:4px;}')
        confirm_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(confirm_btn)
        bottom.addWidget(cancel_btn)
        v.addLayout(bottom)

        self._updateButtons()

    # ---------------- 数据填充 —— 与输入框保持同一集合 ----------------

    def loadFromText(self, text):
        """从输入框文本载入收件人（即时拆分）并默认全选，标记非法项

        参数：
            text<str>：输入框当前文本
        """
        self.all_tokens = splitRecipients(text)
        valid, invalid = validateRecipients(self.all_tokens)
        self.invalid_set = set(invalid)
        # 默认全选（包含非法项也默认勾选，便于用户自行取消）
        self.selected_items = set(self.all_tokens)
        # 重置排序/搜索
        self.sort_combo.blockSignals(True)
        self.sort_combo.setCurrentIndex(0)
        self.sort_combo.blockSignals(False)
        self.search_edit.blockSignals(True)
        self.search_edit.clear()
        self.search_edit.blockSignals(False)
        self._renderList()

    # ---------------- 展示 ----------------

    def _renderList(self):
        """按 排序 + 搜索 重渲染列表（非法项标红，保持勾选状态）"""
        self.recipient_list.itemChanged.disconnect(self._onItemToggled)
        self.recipient_list.clear()

        keyword = self.search_edit.text()
        kw = keyword.strip().lower()
        # 过滤 + 排序（业务逻辑在 activity）
        pool = [t for t in self.all_tokens if t.lower().find(kw) >= 0] if kw else list(self.all_tokens)
        if self.sort_combo.currentData():
            pool = sortRecipients(pool, by_letter=True)
        for tok in pool:
            item = QListWidgetItem(tok)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if tok in self.selected_items
                               else Qt.CheckState.Unchecked)
            # 非法片段标红提示
            if tok in self.invalid_set:
                item.setForeground(QColor('#d93025'))  # 红色
            self.recipient_list.addItem(item)

        self.recipient_list.itemChanged.connect(self._onItemToggled)
        self._updateCount()
        self._updateButtons()

    # ---------------- 勾选状态管理 ----------------

    def _onItemToggled(self, item):
        """勾选变化：同步勾选集合，刷新计数"""
        tok = item.text()
        if item.checkState() == Qt.CheckState.Checked:
            self.selected_items.add(tok)
        else:
            self.selected_items.discard(tok)
        self._updateCount()

    def _setAllChecked(self, checked):
        """全选/取消全选（作用于当前筛选后可见项）"""
        for i in range(self.recipient_list.count()):
            item = self.recipient_list.item(i)
            if checked:
                self.selected_items.add(item.text())
                item.setCheckState(Qt.CheckState.Checked)
            else:
                self.selected_items.discard(item.text())
                item.setCheckState(Qt.CheckState.Unchecked)
        self._updateCount()

    def getCheckedText(self):
        """返回确定时回写输入框用的拼接文本（按当前勾选、去重、分号分隔）

        返回：
            str：回写文本；无勾选项时为空串
        """
        # 按原始顺序保留勾选项；用 activity.mergeRecipients 风格拼接（分号）
        return '; '.join(t for t in self.all_tokens if t in self.selected_items)

    def hasInvalid(self):
        """是否存在非法邮箱（供打开时弹窗提示）

        返回：
            list<str>：非法片段列表
        """
        return sorted(self.invalid_set, key=lambda s: s.lower())

    def _updateCount(self):
        """刷新计数文案（已勾选/当前可见）"""
        total = self.recipient_list.count()
        checked = len(self.selected_items)
        self.count_label.setText('已勾选 %d / 共 %d' % (checked, total))

    def _updateButtons(self):
        """列表为空时禁用全选/取消全选"""
        has_items = self.recipient_list.count() > 0
        self.select_all_btn.setEnabled(has_items)
        self.clear_all_btn.setEnabled(has_items)


class NormalPage(QWidget):
    """普通模式发信页：仿 QQ 邮箱网页版写邮件界面

    包含：
      顶部操作栏（返回、发送、预览、附件菜单、设置菜单）
      收件人（可点抄送/密送展开，分别发送）
      主题行
      两行富文本工具栏 + 正文大输入区
      左下发件人显示
    """

    def __init__(self, db, on_back):
        super().__init__()
        # 数据库引用（发送成功写「已发送」、保存草稿写「草稿箱」时使用）
        self.db = db
        # 返回首页的回调函数
        self.on_back = on_back
        # 当前正在编辑的邮件记录 id（再次编辑进入时非空；新建写信为 None）
        self.editing_id = None
        # 当前编辑来源分类（draft/sent/deleted）：写已发送时据此判断「更新原记录 vs 新增」
        self.edit_category = None
        # 恢复回调（查看已删除时点击「恢复」触发，由 MainWindow 注入）
        self.restore_callback = None
        # 查看模式时需记住的记录 id（供恢复回传）
        self.view_record_id = None
        # 附件完整路径列表（列表控件中只显示文件名）
        self.attachment_paths = []
        # html 内嵌图片映射 {content_id: 文件路径}（正文以 cid:content_id 引用，
        # 发送时交给 LEmail 按映射构建 related，确保引用与 Content-ID 一致）
        self.inline_image_paths = {}
        # 已占用的 content-id 名称集合（保证多个图片的 cid 唯一）
        self._used_image_names = set()
        self.buildUi()

    def buildUi(self):
        """按照 QQ 邮箱网页版布局构建发信页整体结构"""
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 12)
        root.setSpacing(6)

        # 1. 顶部操作栏（仿照截图最上方一排：返回 | 发送 预览 附件▼ 设置▼）
        root.addLayout(self.buildTopBar())

        # 2. 收件人区（收件人 + 右侧抄送/密送/分别发送可点展开）
        root.addLayout(self.buildRecipientArea())

        # 3. 主题行
        root.addLayout(self.buildTitleRow())

        # 先创建正文编辑器（工具栏槽函数需要引用它）
        self.body_editor = QTextEdit()
        self.body_editor.setAcceptRichText(True)
        self.body_editor.setPlaceholderText('输入正文')

        # 4. 富文本工具栏（两行：第一行是插入型工具，第二行是格式型工具）
        root.addWidget(self.buildToolbarRow1())
        root.addWidget(self.buildToolbarRow2())

        # 5. 正文编辑器（拉伸填充剩余空间）
        root.addWidget(self.body_editor, 1)

        # 6. 附件区（默认折叠，点击「附件」添加后展开）
        self.attachment_area = self.buildAttachmentArea()
        self.attachment_area.setVisible(False)
        root.addWidget(self.attachment_area)

        # 7. 底部：左下发件人
        root.addLayout(self.buildSenderRow())

    # ---------------- 顶部操作栏 ----------------

    def buildTopBar(self):
        """构建顶部操作栏：返回 / 发送 / 预览 / 附件▼ / 发信设置▼

        返回：
            QHBoxLayout：顶部按钮行布局
        """
        row = QHBoxLayout()
        row.setSpacing(8)
        # 返回按钮
        self.back_btn = QPushButton('返回')
        self.back_btn.clicked.connect(self.on_back)
        row.addWidget(self.back_btn)

        # 存草稿按钮：仅把当前填写内容写入「草稿箱」，不发信（编辑草稿/查看模式隐藏）
        draft_btn = QPushButton('存草稿')
        draft_btn.setMinimumHeight(34)
        draft_btn.clicked.connect(self.onSaveDraft)
        row.addWidget(draft_btn)
        # 供 setWriteMode/setEditMode/setEditDraftMode 等控制显隐：存草稿按钮
        self.draft_btn = draft_btn

        # 恢复按钮：仅「已删除」只读查看时显示，点击恢复回草稿箱
        self.restore_btn = QPushButton('恢复')
        self.restore_btn.setMinimumHeight(34)
        # 恢复按钮的槽函数由 MainWindow 通过 setRestoreCallback 注入（需携带记录 id）
        self.restore_btn.setVisible(False)
        row.addWidget(self.restore_btn)

        row.addSpacing(12)

        # 发送按钮（主操作）
        send_btn = QPushButton('发送')
        send_btn.setMinimumHeight(34)
        send_btn.setMinimumWidth(90)
        send_btn.clicked.connect(self.onSend)
        row.addWidget(send_btn)
        # 供各模式切换显隐：发送按钮
        self.send_btn = send_btn
        # 预览按钮
        preview_btn = QPushButton('预览')
        preview_btn.setMinimumHeight(34)
        preview_btn.clicked.connect(self.onPreview)
        row.addWidget(preview_btn)
        # 供各模式切换显隐：预览按钮
        self.preview_btn = preview_btn

        # 附件▼ 按钮（下拉菜单：添加附件 / 删除附件）
        attach_menu_btn = QToolButton()
        attach_menu_btn.setText('附件▼')
        attach_menu_btn.setMinimumHeight(34)
        attach_menu_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        # 供查看模式隐藏：附件下拉按钮
        self.attach_menu_btn = attach_menu_btn
        attach_menu = QMenu(self)
        # 菜单项：添加附件
        add_attach_action = QAction('添加附件…', self)
        add_attach_action.triggered.connect(self.onAddAttachment)
        attach_menu.addAction(add_attach_action)
        # 菜单项：删除选中附件
        remove_attach_action = QAction('删除选中附件', self)
        remove_attach_action.triggered.connect(self.onRemoveAttachment)
        attach_menu.addAction(remove_attach_action)
        # 菜单项：清空所有附件
        clear_attach_action = QAction('清空全部附件', self)
        clear_attach_action.triggered.connect(self.onClearAttachments)
        attach_menu.addAction(clear_attach_action)
        attach_menu_btn.setMenu(attach_menu)
        row.addWidget(attach_menu_btn)

        # 发信设置▼ 按钮（下拉菜单：回信地址 / 退信地址 / 分别发送等）
        setting_menu_btn = QToolButton()
        setting_menu_btn.setText('发信设置▼')
        setting_menu_btn.setMinimumHeight(34)
        setting_menu_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        # 供查看模式隐藏：设置下拉按钮
        self.setting_menu_btn = setting_menu_btn
        setting_menu = QMenu(self)
        # 回信地址（对话框输入）
        reply_to_action = QAction('设置回信地址…', self)
        reply_to_action.triggered.connect(self.onSetReplyTo)
        setting_menu.addAction(reply_to_action)
        # 退信地址（对话框输入）
        return_email_action = QAction('设置退信地址…', self)
        return_email_action.triggered.connect(self.onSetReturnEmail)
        setting_menu.addAction(return_email_action)
        setting_menu.addSeparator()
        # 「分别发送」复选（作为可勾选菜单项）
        self.separate_send_action = QAction('分别发送（每人独立一封）', self)
        self.separate_send_action.setCheckable(True)
        self.separate_send_action.setChecked(False)
        setting_menu.addAction(self.separate_send_action)
        setting_menu_btn.setMenu(setting_menu)
        row.addWidget(setting_menu_btn)

        row.addStretch()
        return row

    # ---------------- 收件人区 ----------------

    def buildRecipientArea(self):
        """构建收件人区：左「收件人 ⊕」输入框，右侧「抄送 密送 分别发送」可点切换显隐

        返回：
            QVBoxLayout：收件人区整体布局
        """
        wrap_layout = QVBoxLayout()
        wrap_layout.setSpacing(4)

        # 第一行：收件人标签 + 输入框 + 右侧（抄送/密送/分别发送 链接）
        row1 = QHBoxLayout()
        # 收件人标签（带加号，蓝色视觉）
        to_label = QLabel('收件人  ⊕')
        to_label.setStyleSheet('color:#1a73e8;')
        row1.addWidget(to_label)
        # 收件人输入框
        self.to_edit = QLineEdit()
        self.to_edit.setPlaceholderText('多个收件人用逗号或分号分隔')
        self.to_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        row1.addWidget(self.to_edit, 1)
        # 群发收件人选择按钮：打开收件人选择器（批量勾选后可回填到收件人框）
        self.pick_recipient_btn = QPushButton('收件人选择')
        self.pick_recipient_btn.setToolTip('从 Excel/TXT 批量导入收件人并勾选')
        self.pick_recipient_btn.setStyleSheet('QPushButton{background:#f5f7fa;color:#1a73e8;'
                                              'border:1px solid #d0d7de;padding:4px 10px;'
                                              'border-radius:4px;}'
                                              'QPushButton:hover{background:#e6efff;}')
        self.pick_recipient_btn.clicked.connect(self.onPickRecipients)
        row1.addWidget(self.pick_recipient_btn)
        # 查看所有收件人按钮：收件人过多输入框显示不全时，打开管理对话框查看/勾选
        self.manage_recipient_btn = QPushButton('查看所有收件人')
        self.manage_recipient_btn.setToolTip('收件人过多时查看/勾选管理，确定后回写输入框')
        self.manage_recipient_btn.setStyleSheet('QPushButton{background:#f5f7fa;color:#1a73e8;'
                                                'border:1px solid #d0d7de;padding:4px 10px;'
                                                'border-radius:4px;}'
                                                'QPushButton:hover{background:#e6efff;}')
        self.manage_recipient_btn.clicked.connect(self.onManageRecipients)
        row1.addWidget(self.manage_recipient_btn)
        # 右侧可点击「抄送」「密送」「分别发送」标签
        # 抄送/密送：常规模式支持「抄送」「密送」可点展开输入行；
        # 高级模式因逐封独立发送、一次仅发一封，本无 Cc/Bcc 概念，
        # 由 _cc_bcc_supported=False 走 else 分支，隐藏入口并用说明标签提示，后续有需求再扩展。
        if getattr(self, '_cc_bcc_supported', True):
            # —— 常规模式：可点「抄送」「密送」展开对应输入行 ——
            self.cc_link = QLabel('<a href="#" style="text-decoration:none;color:#666;">抄送</a>')
            self.cc_link.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            self.cc_link.linkActivated.connect(self.toggleCcArea)
            row1.addWidget(self.cc_link)
            self.bcc_link = QLabel('<a href="#" style="text-decoration:none;color:#666;">密送</a>')
            self.bcc_link.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            self.bcc_link.linkActivated.connect(self.toggleBccArea)
            row1.addWidget(self.bcc_link)
        else:
            # —— 高级模式：不提供抄送/密送，用灰色说明标签提示原因 ——
            self.cc_link = None
            self.bcc_link = None
            self.cc_bcc_hint = QLabel('抄送/密送暂不支持（逐封独立发送）')
            self.cc_bcc_hint.setStyleSheet('color:#9aa4b2;font-size:12px;')
            self.cc_bcc_hint.setToolTip(
                '高级模式按收件人逐封独立发送，同一邮件仅发给当前收件人，'
                '因此不提供抄送（Cc）与密送（Bcc）。如确有需要，可在后续版本扩展。')
            row1.addWidget(self.cc_bcc_hint)
        # 分别发送勾选框
        self.separate_checkbox = QCheckBox('分别发送')
        self.separate_checkbox.stateChanged.connect(lambda s: self.separate_send_action.setChecked(bool(s)))
        row1.addWidget(self.separate_checkbox)
        wrap_layout.addLayout(row1)

        # 抄送行（默认隐藏，点「抄送」展开）
        self.cc_row_widget = QWidget()
        cc_row = QHBoxLayout(self.cc_row_widget)
        cc_row.setContentsMargins(0, 0, 0, 0)
        cc_label = QLabel('抄  送')
        cc_label.setMinimumWidth(70)
        cc_row.addWidget(cc_label)
        self.cc_edit = QLineEdit()
        self.cc_edit.setPlaceholderText('可空，多个用逗号/分号分隔')
        cc_row.addWidget(self.cc_edit, 1)
        self.cc_row_widget.setVisible(False)
        wrap_layout.addWidget(self.cc_row_widget)

        # 密送行（默认隐藏，点「密送」展开）
        self.bcc_row_widget = QWidget()
        bcc_row = QHBoxLayout(self.bcc_row_widget)
        bcc_row.setContentsMargins(0, 0, 0, 0)
        bcc_label = QLabel('密  送')
        bcc_label.setMinimumWidth(70)
        bcc_row.addWidget(bcc_label)
        self.bcc_edit = QLineEdit()
        self.bcc_edit.setPlaceholderText('可空，多个用逗号/分号分隔')
        bcc_row.addWidget(self.bcc_edit, 1)
        self.bcc_row_widget.setVisible(False)
        wrap_layout.addWidget(self.bcc_row_widget)

        # 回信 / 退信地址行（默认隐藏，在设置菜单中设定后可见）
        self.reply_row_widget = QWidget()
        reply_row = QHBoxLayout(self.reply_row_widget)
        reply_row.setContentsMargins(0, 0, 0, 0)
        reply_label = QLabel('回  信')
        reply_label.setMinimumWidth(70)
        reply_row.addWidget(reply_label)
        self.reply_edit = QLineEdit()
        self.reply_edit.setPlaceholderText('收件人回复时发送到这里')
        reply_row.addWidget(self.reply_edit, 1)
        self.reply_row_widget.setVisible(False)
        wrap_layout.addWidget(self.reply_row_widget)

        self.return_row_widget = QWidget()
        return_row = QHBoxLayout(self.return_row_widget)
        return_row.setContentsMargins(0, 0, 0, 0)
        return_label = QLabel('退  信')
        return_label.setMinimumWidth(70)
        return_row.addWidget(return_label)
        self.return_edit = QLineEdit()
        self.return_edit.setPlaceholderText('投递失败时退信到这里')
        return_row.addWidget(self.return_edit, 1)
        self.return_row_widget.setVisible(False)
        wrap_layout.addWidget(self.return_row_widget)

        return wrap_layout

    def onPickRecipients(self):
        """打开收件人选择器，确定后把勾选的收件人回填到收件人输入框

        选择器支持从 Excel/TXT 批量导入（解析在 activity 层）；
        确定后把勾选结果委托 activity.mergeRecipients 合并回填（去重、滤空）。
        """
        dlg = RecipientPickerDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        selected = dlg.getSelectedRecipients()
        if not selected:
            return
        # 合并业务委托 activity 层：保留已有 + 新增勾选（去重、滤空），分号分隔
        self.to_edit.setText(mergeRecipients(self.to_edit.text(), selected))

    def onManageRecipients(self):
        """打开「查看所有收件人」对话框管理当前收件人，确定后回写输入框

        与输入框维护同一收件人集合：
          - 打开时从输入框文本即时拆分（activity.splitRecipients），保证列表一致；
          - 存在非法邮箱时弹窗提示并标红；
          - 确定后仅把仍勾选的结果回写输入框，从而列表取消项不会残留、
            输入框删除项也不会残留到列表。
        """
        text = self.to_edit.text().strip()
        if not text:
            QMessageBox.information(self, '查看所有收件人', '收件人框当前为空，无需管理。')
            return
        dlg = RecipientManageDialog(self)
        dlg.loadFromText(text)
        # 检测到非法邮箱时弹窗提示用户（输入框可能因手误删出非法字符）
        invalid = dlg.hasInvalid()
        if invalid:
            QMessageBox.warning(
                self, '存在非法邮箱',
                '检测到 %d 个非法收件人（已标红）：\n%s\n\n确定后仍会保留它们，'
                '请在列表取消勾选或回到输入框修正。'
                % (len(invalid), '\n'.join('- ' + s for s in invalid)))
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        # 回写仍勾选的结果（单一数据源仍是输入框，避免双向不一致）
        result = dlg.getCheckedText()
        self.to_edit.setText(result)

    def toggleCcArea(self):
        """切换抄送输入行的显示/隐藏"""
        visible = self.cc_row_widget.isVisible()
        self.cc_row_widget.setVisible(not visible)
        # 展开后把焦点放到输入框，方便直接输入
        if not visible:
            self.cc_edit.setFocus()

    def toggleBccArea(self):
        """切换密送输入行的显示/隐藏"""
        visible = self.bcc_row_widget.isVisible()
        self.bcc_row_widget.setVisible(not visible)
        if not visible:
            self.bcc_edit.setFocus()

    def onSetReplyTo(self):
        """通过「发信设置」菜单打开回信地址输入对话框，填入后展开回信行"""
        text, ok = self.inputDialogSingleLine('设置回信地址', '收件人回复时发送到该邮箱：')
        if ok and text.strip():
            self.reply_edit.setText(text.strip())
            self.reply_row_widget.setVisible(True)

    def onSetReturnEmail(self):
        """通过「发信设置」菜单打开退信地址输入对话框，填入后展开退信行"""
        text, ok = self.inputDialogSingleLine('设置退信地址', '投递失败时退信到该邮箱：')
        if ok and text.strip():
            self.return_edit.setText(text.strip())
            self.return_row_widget.setVisible(True)

    def inputDialogSingleLine(self, title, prompt):
        """简易单行输入对话框（避免额外依赖 QInputDialog 时不可用的情况，这里使用 PyQt6 自带 QInputDialog）

        参数：
            title<str>：对话框标题
            prompt<str>：提示文字

        返回：
            tuple(str, bool)：(用户输入文本, 是否确认)
        """
        # 局部导入避免顶层依赖
        from PyQt6.QtWidgets import QInputDialog
        result, ok = QInputDialog.getText(self, title, prompt)
        return result, ok

    # ---------------- 主题行 ----------------

    def buildTitleRow(self):
        """构建主题行：「主 题」标签 + 输入框（整行宽度）

        返回：
            QHBoxLayout：主题行布局
        """
        row = QHBoxLayout()
        label = QLabel('主  题')
        label.setMinimumWidth(70)
        row.addWidget(label)
        self.title_edit = MarkableLineEdit()   # 支持变量段局部高亮
        self.title_edit.setPlaceholderText('邮件主题')
        row.addWidget(self.title_edit, 1)
        return row

    # ---------------- 富文本工具栏 ----------------

    def buildToolbarRow1(self):
        """工具栏第一行：撤销/重做 / 图片 / 插入▼ / 签名▼ … （仿照截图第一排：AI润色/图片/插入/导入文档/日程/更多/签名）

        返回：
            QWidget：工具栏组件
        """
        bar = QWidget()
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # 撤销
        undo_btn = QToolButton()
        undo_btn.setText('↺ 撤销')
        undo_btn.clicked.connect(self.body_editor.undo)
        layout.addWidget(undo_btn)
        # 重做
        redo_btn = QToolButton()
        redo_btn.setText('↻ 重做')
        redo_btn.clicked.connect(self.body_editor.redo)
        layout.addWidget(redo_btn)

        layout.addSpacing(8)

        # 图片按钮
        image_btn = QToolButton()
        image_btn.setText('图片')
        image_btn.clicked.connect(self.onInsertImage)
        layout.addWidget(image_btn)

        # 超链接按钮（便捷插入可下载的链接文字，用于大附件走网盘分享的场景）
        link_btn = QToolButton()
        link_btn.setText('超链接')
        link_btn.clicked.connect(self.onInsertLink)
        layout.addWidget(link_btn)

        # 插入▼（水平线 / HTML 片段）
        insert_menu_btn = QToolButton()
        insert_menu_btn.setText('插入▼')
        insert_menu_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        insert_menu = QMenu(self)
        hr_action = QAction('水平线', self)
        hr_action.triggered.connect(self.onInsertHr)
        insert_menu.addAction(hr_action)
        link_action = QAction('超链接…', self)
        link_action.triggered.connect(self.onInsertLink)
        insert_menu.addAction(link_action)
        insert_menu_btn.setMenu(insert_menu)
        layout.addWidget(insert_menu_btn)

        # 签名▼
        sign_menu_btn = QToolButton()
        sign_menu_btn.setText('签名▼')
        sign_menu_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        sign_menu = QMenu(self)
        default_sign_action = QAction('默认签名', self)
        default_sign_action.triggered.connect(self.onInsertDefaultSignature)
        sign_menu.addAction(default_sign_action)
        clear_sign_action = QAction('清空默认签名', self)
        clear_sign_action.triggered.connect(self.onClearSignature)
        sign_menu.addAction(clear_sign_action)
        sign_menu_btn.setMenu(sign_menu)
        layout.addWidget(sign_menu_btn)

        layout.addStretch()
        return bar

    def buildToolbarRow2(self):
        """工具栏第二行：字体 / 字号 / 粗体 斜体 下划线 删除线 / 字色 / 背景色 / 对齐 / 列表 … （仿照截图第二排：格式刷/字体/字号/B/I/U/S/字色/背景色/对齐/列表…）

        返回：
            QWidget：工具栏组件
        """
        bar = QWidget()
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # 字体下拉
        layout.addWidget(QLabel('字体'))
        self.font_combo = QFontComboBox()
        self.font_combo.setEditable(False)
        self.font_combo.currentFontChanged.connect(self.onFontFamilyChanged)
        # 显式高对比样式：保证下拉框与弹出列表在任何父级配色下均为白底深字、可读清晰
        self.font_combo.setStyleSheet(
            'QComboBox{background:#ffffff;color:#1f2328;border:1px solid #d0d7e2;'
            'border-radius:4px;padding:2px 6px;}'
            'QComboBox QAbstractItemView{background:#ffffff;color:#1f2328;'
            'selection-background-color:#1b7bf2;selection-color:#ffffff;}')
        layout.addWidget(self.font_combo)

        # 字号下拉
        layout.addWidget(QLabel('字号'))
        self.size_combo = QComboBox()
        self.size_combo.addItems(FONT_SIZES)
        self.size_combo.setCurrentText('12')
        self.size_combo.currentTextChanged.connect(self.onFontSizeChanged)
        # 与字体下拉一致的高对比样式（白底深字）
        self.size_combo.setStyleSheet(
            'QComboBox{background:#ffffff;color:#1f2328;border:1px solid #d0d7e2;'
            'border-radius:4px;padding:2px 6px;}'
            'QComboBox QAbstractItemView{background:#ffffff;color:#1f2328;'
            'selection-background-color:#1b7bf2;selection-color:#ffffff;}')
        layout.addWidget(self.size_combo)

        # 粗体（可切换）
        self.bold_btn = QToolButton()
        self.bold_btn.setText('B')
        self.bold_btn.setCheckable(True)
        self.bold_btn.clicked.connect(self.onBold)
        bold_font = self.bold_btn.font()
        bold_font.setBold(True)
        self.bold_btn.setFont(bold_font)
        layout.addWidget(self.bold_btn)
        # 斜体
        self.italic_btn = QToolButton()
        self.italic_btn.setText('I')
        self.italic_btn.setCheckable(True)
        self.italic_btn.clicked.connect(self.onItalic)
        italic_font = self.italic_btn.font()
        italic_font.setItalic(True)
        self.italic_btn.setFont(italic_font)
        layout.addWidget(self.italic_btn)
        # 下划线
        self.underline_btn = QToolButton()
        self.underline_btn.setText('U')
        self.underline_btn.setCheckable(True)
        self.underline_btn.clicked.connect(self.onUnderline)
        underline_font = self.underline_btn.font()
        underline_font.setUnderline(True)
        self.underline_btn.setFont(underline_font)
        layout.addWidget(self.underline_btn)
        # 删除线
        self.strike_btn = QToolButton()
        self.strike_btn.setText('S')
        self.strike_btn.setCheckable(True)
        self.strike_btn.clicked.connect(self.onStrikeOut)
        strike_font = self.strike_btn.font()
        strike_font.setStrikeOut(True)
        self.strike_btn.setFont(strike_font)
        layout.addWidget(self.strike_btn)

        # 字色
        color_btn = QToolButton()
        color_btn.setText('A▼')
        color_btn.clicked.connect(self.onTextColor)
        layout.addWidget(color_btn)

        # 背景色
        bg_color_btn = QToolButton()
        bg_color_btn.setText('✎▼')
        bg_color_btn.clicked.connect(self.onTextBgColor)
        layout.addWidget(bg_color_btn)

        layout.addSpacing(6)

        # 对齐按钮组
        align_left_btn = QToolButton()
        align_left_btn.setText('左')
        align_left_btn.clicked.connect(self.onAlignLeft)
        layout.addWidget(align_left_btn)
        align_center_btn = QToolButton()
        align_center_btn.setText('中')
        align_center_btn.clicked.connect(self.onAlignCenter)
        layout.addWidget(align_center_btn)
        align_right_btn = QToolButton()
        align_right_btn.setText('右')
        align_right_btn.clicked.connect(self.onAlignRight)
        layout.addWidget(align_right_btn)

        layout.addSpacing(6)

        # 项目符号 / 编号列表
        bullet_btn = QToolButton()
        bullet_btn.setText('• 列表')
        bullet_btn.clicked.connect(self.onBulletList)
        layout.addWidget(bullet_btn)
        numbered_btn = QToolButton()
        numbered_btn.setText('1. 编号')
        numbered_btn.clicked.connect(self.onNumberedList)
        layout.addWidget(numbered_btn)

        layout.addStretch()
        return bar

    # ---------------- 附件区 ----------------

    def buildAttachmentArea(self):
        """构建附件区（列表 + 标签标题）

        返回：
            QWidget：附件区组件
        """
        area = QWidget()
        layout = QVBoxLayout(area)
        layout.setContentsMargins(0, 0, 0, 0)
        head = QHBoxLayout()
        head.addWidget(QLabel('附件：'))
        head.addStretch()
        count_label = QLabel()
        self.attachment_count_label = count_label
        head.addWidget(count_label)
        layout.addLayout(head)
        self.attachment_list = QListWidget()
        self.attachment_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.attachment_list.setMaximumHeight(100)
        layout.addWidget(self.attachment_list)
        return area

    def updateAttachmentHeader(self):
        """更新附件区标题的数量显示，并根据是否有附件切换显示/隐藏"""
        count = len(self.attachment_paths)
        # 高级模式配置页未创建 attachment_area 时，仅更新计数标签
        if hasattr(self, 'attachment_count_label') and self.attachment_count_label is not None:
            self.attachment_count_label.setText('共 %d 个' % count)
        # 普通模式附件区在高级/普通页均可能有；高级页不存在 attachment_area 时跳过显隐切换
        if hasattr(self, 'attachment_area'):
            self.attachment_area.setVisible(count > 0)

    # ---------------- 底部发件人行 ----------------

    def buildSenderRow(self):
        """构建底部发件人显示行

        返回：
            QHBoxLayout：发件人行布局
        """
        row = QHBoxLayout()
        label = QLabel('发件人：')
        label.setStyleSheet('color:#666;')
        row.addWidget(label)
        # 发件人昵称（可编辑，模仿 QQ 邮箱「昵称 <邮箱>」）
        self.from_edit = QLineEdit()
        self.from_edit.setPlaceholderText(DEFAULT_FROM_NAME)
        self.from_edit.setMaximumWidth(260)
        row.addWidget(self.from_edit)
        # 邮箱地址（只读，阿里云账号默认发件箱）
        self.from_email_label = QLabel('&lt;%s&gt;' % DEFAULT_FROM_EMAIL)
        self.from_email_label.setStyleSheet('color:#666;')
        row.addWidget(self.from_email_label)
        row.addStretch()
        return row

    # ---------------- 预览 ----------------

    def onPreview(self):
        """预览当前邮件：弹出一个只读对话框展示 HTML 正文"""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle('邮件预览（HTML）')
        msg_box.setIcon(QMessageBox.Icon.Information)
        # 展示主题 + 收件人 + HTML 正文
        header = '<h3>主题：%s</h3>' % (self.title_edit.text() or '(无主题)')
        header += '<p><b>收件人：</b>%s</p>' % (self.to_edit.text() or '(空)')
        if self.cc_edit.isVisible() and self.cc_edit.text():
            header += '<p><b>抄送：</b>%s</p>' % self.cc_edit.text()
        html = self.body_editor.toHtml()
        msg_box.setText(header + '<hr>' + html)
        msg_box.exec()

    # ---------------- 富文本编辑槽函数 ----------------

    def onFontFamilyChanged(self, font):
        """字体下拉变更：应用到正文当前选择及后续输入"""
        self.body_editor.setFontFamily(font.family())

    def onFontSizeChanged(self, text):
        """字号下拉变更：应用到正文当前选择及后续输入"""
        if text:
            self.body_editor.setFontPointSize(float(text))

    def onBold(self):
        """切换加粗状态（根据当前字体权重判断）"""
        bold_value = QFont.Weight.Bold.value
        if self.body_editor.fontWeight() >= bold_value:
            self.body_editor.setFontWeight(QFont.Weight.Normal.value)
            self.bold_btn.setChecked(False)
        else:
            self.body_editor.setFontWeight(bold_value)
            self.bold_btn.setChecked(True)

    def onItalic(self):
        """切换斜体状态"""
        target = not self.body_editor.fontItalic()
        self.body_editor.setFontItalic(target)
        self.italic_btn.setChecked(target)

    def onUnderline(self):
        """切换下划线状态"""
        target = not self.body_editor.fontUnderline()
        self.body_editor.setFontUnderline(target)
        self.underline_btn.setChecked(target)

    def onStrikeOut(self):
        """切换删除线状态（通过合并 QTextCharFormat 实现）"""
        cursor = self.body_editor.textCursor()
        fmt = cursor.charFormat()
        fmt.setFontStrikeOut(not fmt.fontStrikeOut())
        cursor.mergeCharFormat(fmt)
        self.body_editor.setTextCursor(cursor)
        self.strike_btn.setChecked(fmt.fontStrikeOut())

    def onTextColor(self):
        """弹出色盘设置文字颜色"""
        color = QColorDialog.getColor(parent=self)
        if color.isValid():
            self.body_editor.setTextColor(color)

    def onTextBgColor(self):
        """弹出色盘设置文字背景色（高亮色）"""
        color = QColorDialog.getColor(parent=self)
        if color.isValid():
            cursor = self.body_editor.textCursor()
            fmt = cursor.charFormat()
            fmt.setBackground(color)
            cursor.mergeCharFormat(fmt)
            self.body_editor.setTextCursor(cursor)

    def onAlignLeft(self):
        """设置段落左对齐"""
        self.body_editor.setAlignment(Qt.AlignmentFlag.AlignLeft)

    def onAlignCenter(self):
        """设置段落居中对齐"""
        self.body_editor.setAlignment(Qt.AlignmentFlag.AlignCenter)

    def onAlignRight(self):
        """设置段落右对齐"""
        self.body_editor.setAlignment(Qt.AlignmentFlag.AlignRight)

    def onBulletList(self):
        """在光标所在段落插入项目符号列表"""
        cursor = self.body_editor.textCursor()
        fmt = QTextListFormat()
        fmt.setStyle(QTextListFormat.Style.ListDisc)
        cursor.createList(fmt)

    def onNumberedList(self):
        """在光标所在段落插入编号列表"""
        cursor = self.body_editor.textCursor()
        fmt = QTextListFormat()
        fmt.setStyle(QTextListFormat.Style.ListDecimal)
        cursor.createList(fmt)

    def onInsertImage(self):
        """在正文光标处插入图片，以 cid:文件名 引用并注册为邮件内嵌图片

        说明：不能使用 QTextCursor.insertImage(QImage) 直接插入，因为 Qt
        会将其序列化为 base64 的 data: URI 写入 HTML，多数邮件客户端不解析；
        也会导致 LEmail 无法走 related 结构。故此处统一用「文件名即 cid」
        的方式（与 MainEmail.buildRelated 的 content-id 规则一致）：
        编辑器内通过 addResource 显示，发送时 toHtml() 输出 <img src="cid:文件名">。
        """
        path, _ = QFileDialog.getOpenFileName(
            self, '选择图片', '', '图片文件 (*.png *.jpg *.jpeg *.gif *.bmp)')
        if not path:
            return
        image = QImage(path)
        if image.isNull():
            QMessageBox.warning(self, '提示', '无法读取图片：' + path)
            return
        # 用文件名作为 content-id；同名冲突时追加序号保证唯一
        content_id = self.uniqueImageName(path)
        # 注册图片显示资源，使编辑器内能看到图片（key 为 cid:文件名）
        document = self.body_editor.document()
        document.addResource(QTextDocument.ResourceType.ImageResource,
                             QUrl('cid:' + content_id), image)
        # 以 cid 方式插入图片（toHtml() 会输出 src="cid:文件名"）
        image_format = QTextImageFormat()
        image_format.setName('cid:' + content_id)
        cursor = self.body_editor.textCursor()
        cursor.insertImage(image_format)
        # 记录图片文件路径（以 content_id 为键），发送时交给 LEmail 完成 related 内嵌
        self.inline_image_paths[content_id] = path

    def uniqueImageName(self, path):
        """为图片文件生成唯一的 content-id 名称（重名时追加序号）

        参数：
            path<str>：图片文件路径

        返回：
            str：唯一的 content-id 名称（不含 cid: 前缀）
        """
        base_name = os.path.basename(path)
        name = base_name
        counter = 1
        # 若名称已被占用，改为 base_序号.后缀
        while name in self._used_image_names:
            root, ext = os.path.splitext(base_name)
            name = '%s_%d%s' % (root, counter, ext)
            counter += 1
        self._used_image_names.add(name)
        return name

    def onInsertHr(self):
        """插入水平线"""
        self.body_editor.textCursor().insertHtml('<hr>')

    def onInsertLink(self):
        """插入超链接（弹窗输入 url 与显示文字）

        说明：适合将大附件上传到网盘后，把可下载分享链接插入正文；
        收件人点击"显示文字"即可下载，不受 15MB 附件限额影响。
        """
        from PyQt6.QtWidgets import QInputDialog
        url, ok = QInputDialog.getText(
            self, '插入超链接',
            '请输入可下载链接 URL\n（大附件可先上传到网盘获取分享链接）：')
        if not ok or not url.strip():
            return
        text, ok = QInputDialog.getText(self, '插入超链接', '请输入显示文字（留空则用链接）：')
        if not ok:
            return
        # 若没填显示文字，用 url 作为显示文字
        display = text.strip() or url.strip()
        self.body_editor.textCursor().insertHtml('<a href="%s">%s</a>' % (url.strip(), display))

    def onInsertDefaultSignature(self):
        """插入默认签名：署名 + 发件邮箱 + 一句话"""
        html = ('<br><hr><p>-- <br>'
                '发件人：<b>%s</b><br>'
                '邮箱：%s<br>'
                '（来自 Lhack 邮箱助手）</p>') % (
            self.from_edit.text().strip() or DEFAULT_FROM_NAME,
            DEFAULT_FROM_EMAIL)
        cursor = self.body_editor.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        cursor.insertHtml(html)
        self.body_editor.setTextCursor(cursor)

    def onClearSignature(self):
        """清空签名只是提示：用户手动删除"""
        QMessageBox.information(self, '提示', '签名已插入，若要清空请手动选中删除。')

    # ---------------- 附件槽函数 ----------------

    def onAddAttachment(self):
        """弹出文件多选对话框添加附件（含总大小预检，超阿里云 15MB 上限提示并跳过）"""
        files, _ = QFileDialog.getOpenFileNames(self, '选择附件')
        # 先汇总当前已添加附件的原始字节数（供添加新文件时做上限预判）
        current_bytes = sum(os.path.getsize(p) for p in self.attachment_paths if os.path.exists(p))
        for file_path in files:
            # 去重后加入列表，并显示文件名
            if file_path not in self.attachment_paths:
                # 附件数量达到官方上限 100 时拒绝继续添加
                if len(self.attachment_paths) >= 100:
                    QMessageBox.warning(
                        self, '附件数量超限',
                        '附件数量已达 100 个（阿里云邮件推送单封上限），\n不能再添加附件。')
                    break
                # 把待添加文件计入后校验总大小（base64 膨胀），超 15MB 上限则拒绝添加
                add_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                if int((current_bytes + add_size) * 1.35) > 15 * 1024 * 1024:
                    QMessageBox.warning(
                        self, '附件过大',
                        '添加「%s」后附件总大小可能超过阿里云邮件推送的 15MB 上限，\n'
                        '该附件未添加。\n'
                        '提示：不支持发送超大附件，可将该文件上传到网盘/云盘，\n'
                        '获取分享链接后点正文「超链接」按钮插入，收件人点击即可下载。'
                        % os.path.basename(file_path))
                    continue
                self.attachment_paths.append(file_path)
                self.attachment_list.addItem(os.path.basename(file_path))
                # 累加已加入的成功附件字节数
                current_bytes += add_size
        self.updateAttachmentHeader()

    def onRemoveAttachment(self):
        """删除选中的附件"""
        row = self.attachment_list.currentRow()
        if row < 0:
            QMessageBox.information(self, '提示', '请先选中要删除的附件。')
            return
        self.attachment_list.takeItem(row)
        self.attachment_paths.pop(row)
        self.updateAttachmentHeader()

    def onClearAttachments(self):
        """清空所有附件"""
        if not self.attachment_paths:
            return
        self.attachment_paths.clear()
        self.attachment_list.clear()
        self.updateAttachmentHeader()

    # ---------------- 发送 ----------------

    def showLoading(self):
        """显示「正在发送中...」遮罩（懒加载创建，首次点击发送时生成）"""
        if not hasattr(self, 'loading_overlay') or self.loading_overlay is None:
            # 遮罩挂在顶层窗口上，可覆盖整个窗口（含状态栏区域）
            self.loading_overlay = LoadingOverlay(self.window())
        self.loading_overlay.showLoading()

    def hideLoading(self):
        """隐藏发送中遮罩"""
        if hasattr(self, 'loading_overlay') and self.loading_overlay is not None:
            self.loading_overlay.hideLoading()

    def onSendFinished(self, result):
        """发送工作线程完成回调：隐藏遮罩；成功后写入已发送并返回列表主页"""
        # 隐藏发送中遮罩
        self.hideLoading()
        if result:
            # 发送成功：先把本次邮件写入数据库「已发送」，再提示并返回
            self._storeRecord(CATEGORY_SENT)
            QMessageBox.information(self, '发送成功', '邮件已发送成功！')
            # 成功后不再留在发信页：回调切回主窗口的列表页（返回按钮同样调用 on_back）
            self.on_back()
        else:
            QMessageBox.warning(self, '发送失败', '邮件发送失败，请检查网络或 SMTP 配置。')

    def onSend(self):
        """收集页面参数，交给 NormalMode 实装发送（普通模式功能实装入口）"""
        log.info('普通模式发起发送，收件人 %d 人', len(parseEmails(self.to_edit.text())))
        # 收件人为必填项
        to_list = parseEmails(self.to_edit.text())
        if not to_list:
            QMessageBox.warning(self, '提示', '收件人不能为空。')
            return
        # 抄送/密送仅在展开且有内容时取用
        cc_list = parseEmails(self.cc_edit.text()) if self.cc_row_widget.isVisible() else []
        bcc_list = parseEmails(self.bcc_edit.text()) if self.bcc_row_widget.isVisible() else []
        # 回信/退信地址：只有设置了内容才使用
        reply_to = self.reply_edit.text().strip()
        return_email = self.return_edit.text().strip()
        # 汇总各参数，生成用于发送的配置
        kwargs = dict(
            n_name=self.from_edit.text().strip() or DEFAULT_FROM_NAME,
            to_list=to_list,
            cc_list=cc_list or None,
            bcc_list=bcc_list or None,
            reply_to=reply_to,
            return_email=return_email,
            email_title=self.title_edit.text().strip(),
            html_text=self.body_editor.toHtml(),
            attachment_paths=list(self.attachment_paths) if self.attachment_paths else None,
            inline_image_paths=dict(self.inline_image_paths) if self.inline_image_paths else None,
        )
        # 实例化普通模式并注入配置，随后交由 NormalMode 调用 LEmail 发送
        normal_mode = NormalMode()
        # 收件人非空时配置必成功，此处配置失败仅兜底提示
        if not normal_mode.setConfig(**kwargs):
            QMessageBox.warning(self, '提示', '邮件配置失败，请检查收件人。')
            return
        # 发送前先校验收件人数/附件数量：超过官方上限则提示并中止
        over_limits, limit_msg = normal_mode.checkLimits()
        if over_limits:
            QMessageBox.warning(self, '超出限制', limit_msg)
            return
        # 发送前预检总大小：超过阿里云 15MB 上限则提示用户改用网盘+超链接，不发信
        over_limit, est_size, max_size = normal_mode.checkSize()
        if over_limit:
            QMessageBox.warning(
                self, '附件过大',
                '邮件总大小约 %.1f MB，超过阿里云邮件推送的 %.1f MB 上限，%s\n\n'
                '不支持发送超大附件。请：\n'
                '1. 将大文件上传到网盘/云盘，获取可下载分享链接；\n'
                '2. 在正文点「超链接」按钮，把链接插入邮件；\n'
                '3. 收件人点击即可下载。' % (
                    est_size / 1024 / 1024, max_size / 1024 / 1024, '无法发送。'))
            return
        # 显示「正在发送中...」遮罩提示用户（转圈动画在后台线程发送期间持续旋转）
        self.showLoading()
        # 创建发送工作线程，将实际发送放入后台，避免阻塞遮罩动画
        self.send_worker = SendWorker(normal_mode)
        self.send_worker.finished_signal.connect(self.onSendFinished)
        self.send_worker.start()

    # ---------------- 数据库关联：存草稿 / 已发送 / 再次编辑回填 ----------------

    def rememberRecordId(self, mail_id):
        """记录当前编辑的邮件记录 id（再次编辑进入时由 MainWindow 调用）

        参数：
            mail_id<int>|None：当前记录 id；新建写信传 None
        备注：同时把 edit_category 保留旧值，仅改 id（避免覆盖来源分类）。
        """
        self.editing_id = mail_id

    def onSaveDraft(self):
        """存草稿：把当前填写内容写入数据库「草稿箱」，不发送

        说明：存草稿仅在书写新邮件或再次编辑「已发送」时提供；编辑草稿/查看模式不显示该按钮，
        故此处一律新建一条草稿，避免覆盖他人草稿或改动已发送记录。
        """
        payload = self.collectPayload()
        # 新建草稿（存草稿按钮不在草稿编辑/查看模式显示，故不存在「更新既有草稿」场景）
        self.db.insertMail(CATEGORY_DRAFT, **payload)
        QMessageBox.information(self, '保存成功', '邮件已存入草稿箱。')
        # 返回列表主页并切到草稿箱，让用户看到新保存的草稿
        self.on_back()

    def _storeRecord(self, category):
        """把当前页面内容写入数据库指定分类（发送成功写已发送）

        参数：
            category<str>：目标分类常量（CATEGORY_SENT / CATEGORY_DRAFT 等）
        备注：从草稿/已发送再次编辑后发送，更新原记录到已发送；新写信则新增一条已发送。
        """
        payload = self.collectPayload()
        if self.editing_id is not None and self.edit_category in (CATEGORY_DRAFT, CATEGORY_SENT):
            # 来自草稿/已发送的再次编辑：更新原记录并改分类为已发送
            self.db.updateMail(self.editing_id, category=category, **{k: v for k, v in payload.items()})
        else:
            # 新建写信：新增一条已发送记录
            self.db.insertMail(category, **payload)

    def collectPayload(self):
        """收集发信页当前填写的全部字段，生成一条邮件记录 dict（供写库：已发送/草稿）

        返回：
            dict：包含 title/recipient/from_name/to_list/cc_list/bcc_list/reply_to/
                  return_email/html_text/attachment_paths/inline_image_paths
        """
        to_list = parseEmails(self.to_edit.text())
        # 抄送/密送仅在展开且有内容时取用
        cc_list = parseEmails(self.cc_edit.text()) if self.cc_row_widget.isVisible() else []
        bcc_list = parseEmails(self.bcc_edit.text()) if self.bcc_row_widget.isVisible() else []
        # 收件人摘要：用于列表页「收件人/发件人」列展示
        recipient = '; '.join(to_list)
        return dict(
            from_name=self.from_edit.text().strip() or DEFAULT_FROM_NAME,
            to_list=to_list,
            cc_list=cc_list,
            bcc_list=bcc_list,
            reply_to=self.reply_edit.text().strip(),
            return_email=self.return_edit.text().strip(),
            title=self.title_edit.text().strip(),
            recipient=recipient,
            html_text=self.body_editor.toHtml(),
            attachment_paths=list(self.attachment_paths) if self.attachment_paths else [],
            inline_image_paths=dict(self.inline_image_paths) if self.inline_image_paths else {},
        )

    def fillFromRecord(self, record):
        """用数据库邮件记录回填发信页各输入（再次编辑进入时调用）

        参数：
            record<dict>：一条数据库邮件记录
        """
        # 先把表单清空（含隐藏展开行/清附件），避免上次内容残留
        self.clearForm()
        # 记录来源分类，供写已发送时「更新原记录 vs 新增」判断
        self.edit_category = record.get('category')
        # 发件人昵称 / 收件人 / 抄送 / 密送
        self.from_edit.setText(record.get('from_name', ''))
        self.to_edit.setText('; '.join(record.get('to_list') or []))
        cc_list = record.get('cc_list') or []
        if cc_list:
            self.cc_row_widget.setVisible(True)
            self.cc_edit.setText('; '.join(cc_list))
        bcc_list = record.get('bcc_list') or []
        if bcc_list:
            self.bcc_row_widget.setVisible(True)
            self.bcc_edit.setText('; '.join(bcc_list))
        # 回信 / 退信地址（有值则展开对应行）
        if record.get('reply_to'):
            self.reply_row_widget.setVisible(True)
        if record.get('return_email'):
            self.return_row_widget.setVisible(True)
        # 主题与正文（HTML）
        self.title_edit.setText(record.get('title', ''))
        self.body_editor.setHtml(record.get('html_text', '') or '')
        # 附件路径列表与内嵌图片映射
        self.attachment_paths = list(record.get('attachment_paths') or [])
        self.attachment_list.clear()
        for p in self.attachment_paths:
            self.attachment_list.addItem(os.path.basename(p))
        self.inline_image_paths = dict(record.get('inline_image_paths') or {})
        self._used_image_names = set(self.inline_image_paths.keys())
        # 有附件则展开附件区；无则自动隐藏（updateAttachmentHeader 内部处理）
        self.updateAttachmentHeader()

    def setRestoreCallback(self, callback):
        """注入「恢复」按钮回调（由 MainWindow 提供，携带当前查看的记录 id）

        参数：
            callback<callable(mail_id)>：恢复处理函数
        """
        self.restore_callback = callback
        self.restore_btn.clicked.connect(self._onRestoreClicked) if callback else None

    def _onRestoreClicked(self):
        """点「恢复」：调用注入的恢复回调，回传当前查看的记录 id"""
        if self.restore_callback is not None:
            self.restore_callback(self.view_record_id)

    # ---------------- 页面状态模式 ----------------

    def _setFormReadOnly(self, readonly):
        """统一设置表单输入控件的只读状态

        参数：
            readonly<bool>：True=全部只读禁用；False=可编辑
        """
        for edit in (self.to_edit, self.cc_edit, self.bcc_edit,
                     self.reply_edit, self.return_edit, self.title_edit):
            edit.setReadOnly(readonly)
        self.body_editor.setReadOnly(readonly)
        # 附件/内嵌图片相关按钮与菜单在只读时隐藏（无法编辑附件）
        self.attach_menu_btn.setVisible(not readonly)
        self.setting_menu_btn.setVisible(not readonly)

    def clearForm(self):
        """清空写信表单（含展开行/附件/图片），用于写信初始化或再次编辑回填前重置"""
        # 文本输入清空
        self.from_edit.setText('')
        self.to_edit.setText('')
        self.cc_edit.setText('')
        self.bcc_edit.setText('')
        self.reply_edit.setText('')
        self.return_edit.setText('')
        self.title_edit.setText('')
        self.body_editor.clear()
        # 隐藏展开行（抄送/密送/回信/退信）
        self.cc_row_widget.setVisible(False)
        self.bcc_row_widget.setVisible(False)
        self.reply_row_widget.setVisible(False)
        self.return_row_widget.setVisible(False)
        # 清空附件与内嵌图片
        self.attachment_paths = []
        self.attachment_list.clear()
        self.inline_image_paths = {}
        self._used_image_names = set()
        self.updateAttachmentHeader()
        # 重置编辑/查看状态
        self.editing_id = None
        self.edit_category = None
        self.view_record_id = None

    def setWriteMode(self):
        """写信模式：所有编辑可用，显示发送/存草稿/附件/设置，隐藏恢复"""
        self._setFormReadOnly(False)
        # 显示发送/存草稿/预览；隐藏恢复
        self.send_btn.setVisible(True)
        self.draft_btn.setVisible(True)
        self.preview_btn.setVisible(True)
        self.restore_btn.setVisible(False)

    def setEditMode(self):
        """再次编辑模式：可发送/预览；草稿箱不提供存草稿；隐藏恢复"""
        self._setFormReadOnly(False)
        self.send_btn.setVisible(True)
        self.preview_btn.setVisible(True)
        # 草稿箱编辑不提供「存草稿」（再次保存草稿无意义）
        self.draft_btn.setVisible(self.edit_category != CATEGORY_DRAFT)
        self.restore_btn.setVisible(False)

    def setViewMode(self, restorable=False):
        """只读查看模式：所有编辑禁用；仅显示返回/恢复（已删除），隐藏发送/存草稿/附件/设置

        参数：
            restorable<bool>：True=已删除可恢复，显示「恢复」按钮
        """
        self._setFormReadOnly(True)
        # 编辑类按钮全部隐藏；恢复仅已删除查看时显示
        self.send_btn.setVisible(False)
        self.draft_btn.setVisible(False)
        self.preview_btn.setVisible(False)
        self.restore_btn.setVisible(restorable)
        # 恢复需记录当前查看的记录 id（onRequestView 已在 fillFromRecord 前设置）


class MarkableLineEdit(QLineEdit):
    """单行输入框（用于主题/昵称）

    变量仅用 $$...$$ 符号标记，不改变文字背景颜色。此类保留原类名与位置，
    保持普通 QLineEdit 外观，避免对既有引用造成影响。
    """
    # 变量正则：与 AdvancedPage.VARIABLE_PATTERN 一致，匹配 $$...$$
    VAR_PATTERN = re.compile(r'\$\$(.*?)\$\$')


class AdvancedPage(NormalPage):
    """高级模式发信页：完整具备普通模式发信页的全部写信能力，并叠加「变量标记」

    继承 NormalPage，复用全部普通模式写信功能：
      顶部操作栏（返回/存草稿/恢复/发送/预览/附件▼/发信设置▼）、
      收件人区（抄送/密送可展开、分别发送、收件人选择导入、查看所有收件人）、
      主题行、两行富文本工具栏、正文编辑器、附件区、发件人栏。
    在此基础上额外提供「变量标记」（Payload Positions）：
      在主题/正文中选中文字 → 点击「标记为变量」→ 包裹为 $变量名$；
      右侧变量面板实时解析并列出所有 $变量$，可删除选中/清空所有。
    右侧变量面板布局保持不变；$变量$ 后续发送时按导入数据的列替换。
    """

    # 变量包裹正则：匹配 $...$ 格式（用 $$ 包裹变量，避免花括号与正文内容冲突）
    VARIABLE_PATTERN = re.compile(r'\$\$(.*?)\$\$')
    # 预览页分段生成：每批生成 PREVIEW_BATCH_SIZE 封邮件，避免大批量一次性生成卡死界面
    PREVIEW_BATCH_SIZE = 5
    # 抄送/密送支持开关：高级模式按收件人逐封独立发送、一次仅发一封，本无 Cc/Bcc 概念，
    # 故置 False 隐藏「抄送」「密送」入口（对照常规模式默认开启）；后续有需求再开启并扩展。
    _cc_bcc_supported = False

    # ---------------- 任务机制 ----------------
    def startNewTask(self):
        """开启新一轮发送任务：重置任务名（自增默认名称）、清空上次任务的全部编辑状态

        每次从模式选择页进入高级模式即视为开启新任务，页面必须为空白（无任何残留内容）。
        由 MainWindow.showHighModePlaceholder 回调。

        说明：恢复草稿任务走 _restoreTaskConfig 还原，不会经过本方法，故可安全全量清空。
        """
        # 默认任务名：高级模式任务x，x = 历史任务总数 + 1（跨已发送/草稿箱累计）
        default_name = '高级模式任务%d' % (self.db.countTasks() + 1)
        if hasattr(self, 'task_name_edit'):
            self.task_name_edit.setText(default_name)
        # 预览批量生成邮件数：进入新任务时按用户设置刷新（原为类属性写死 5）
        self.PREVIEW_BATCH_SIZE = int(settings.get('preview_batch_size'))
        # 当前编辑中任务的数据库 id（草稿恢复进入时为该草稿任务 id，新任务为 None）
        self._current_task_id = None
        # ---- 清空上一任务残留的编辑内容（避免新任务打开时带旧数据）----
        def _clear_inputattr(attr_name):
            """清空指定输入控件文本；控件不存在则忽略"""
            widget = getattr(self, attr_name, None)
            if widget is not None:
                widget.setText('')
        for name in ('to_edit', 'cc_edit', 'bcc_edit', 'reply_edit', 'return_edit',
                     'nickname_edit', 'title_edit'):
            _clear_inputattr(name)
        if hasattr(self, 'body_editor'):
            self.body_editor.clear()
        # 抄送/密送展开行收起
        for name in ('cc_row_widget', 'bcc_row_widget'):
            if hasattr(self, name):
                getattr(self, name).setVisible(False)
        # 附件：固定路径与文件夹列表清空
        self.attachment_paths = []
        if hasattr(self, 'attachment_list'):
            self.attachment_list.clear()
        self.attach_folders = []
        if hasattr(self, 'attach_folder_list'):
            self.attach_folder_list.clear()
        self.inline_image_paths = {}
        # ---- 重置负载绑定 / 预览 / 变量 ----
        self._load_map = {}
        self._marker_keys = []
        self._marker_names = {}
        self._file_row_counts = {}
        self._preview_mails = []
        self._preview_all_generated = False
        self._preview_rows_total = 0
        self._load_state_initialized = False
        self._markers_loaded = False
        # 回到配置页第1步，并刷新右侧变量面板（此时应为空）
        if hasattr(self, 'step_stack'):
            self.step_stack.setCurrentIndex(0)
            self._refreshIndicator(1)
            self._refreshFooter(1)
        if hasattr(self, '_updateVariableList'):
            self._updateVariableList()

    def _currentTaskName(self):
        """取当前任务名（用户自定义或默认值）

        返回：
            str：任务名
        """
        if hasattr(self, 'task_name_edit') and self.task_name_edit.text().strip():
            return self.task_name_edit.text().strip()
        return '高级模式任务%d' % (self.db.countTasks() + 1)

    def _serializeTaskConfig(self):
        """序列化当前高级模式全部可复现配置供存草稿/发送落库

        覆盖全部输入控件、负载绑定、附件文件夹与已生成预览邮件，保证可完整复现。

        返回：
            dict：可 JSON 序列化的配置快照
        """
        cfg = {}
        # ---- 输入控件（高级模式配置页可能缺少某字段时自动跳过，避免 AttributeError）----
        cfg['to_text'] = self.to_edit.text() if hasattr(self, 'to_edit') else ''
        cfg['cc_text'] = self.cc_edit.text() if hasattr(self, 'cc_edit') else ''
        cfg['bcc_text'] = self.bcc_edit.text() if hasattr(self, 'bcc_edit') else ''
        cfg['from_text'] = self.from_edit.text() if hasattr(self, 'from_edit') else ''
        cfg['reply_text'] = self.reply_edit.text() if hasattr(self, 'reply_edit') else ''
        cfg['return_text'] = self.return_edit.text() if hasattr(self, 'return_edit') else ''
        cfg['nickname_text'] = self.nickname_edit.text() if hasattr(self, 'nickname_edit') else ''
        cfg['title_text'] = self.title_edit.text() if hasattr(self, 'title_edit') else ''
        cfg['body_html'] = self.body_editor.toHtml() if hasattr(self, 'body_editor') else ''
        # ---- 抄送/密送行展开状态 ----
        cfg['cc_visible'] = bool(getattr(self, 'cc_row_widget', None)
                                 and self.cc_row_widget.isVisible())
        cfg['bcc_visible'] = bool(getattr(self, 'bcc_row_widget', None)
                                  and self.bcc_row_widget.isVisible())
        # ---- 附件：固定模式（附件路径列表） vs 不固定模式（文件夹列表）----
        cfg['attach_fixed_mode'] = self.attach_fixed_radio.isChecked() \
            if hasattr(self, 'attach_fixed_radio') else True
        cfg['attachment_paths'] = list(getattr(self, 'attachment_paths', []) or [])
        cfg['attach_folders'] = list(getattr(self, 'attach_folders', []) or [])
        # ---- 内嵌图片映射 ----
        cfg['inline_image_paths'] = dict(getattr(self, 'inline_image_paths', {}) or {})
        # ---- 负载绑定 ----
        cfg['load_map'] = dict(getattr(self, '_load_map', {}) or {})
        cfg['marker_keys'] = list(getattr(self, '_marker_keys', []) or [])
        cfg['marker_names'] = dict(getattr(self, '_marker_names', {}) or {})
        # ---- 已生成预览邮件（含用户编辑后的最终内容）----
        cfg['preview_mails'] = list(getattr(self, '_preview_mails', []) or [])
        cfg['preview_all_generated'] = getattr(self, '_preview_all_generated', False)
        cfg['preview_rows_total'] = getattr(self, '_preview_rows_total', 0)
        return cfg

    def _restoreTaskConfig(self, cfg):
        """用序列化配置恢复高级模式页面（供从草稿箱点击任务恢复编辑）

        参数：
            cfg<dict>：_serializeTaskConfig 产生的配置快照
        """
        if not cfg:
            return
        # ---- 输入控件（逐个用 hasattr 保护，旧/待补字段缺失不致崩溃）----
        if hasattr(self, 'to_edit') and cfg.get('to_text') is not None:
            self.to_edit.setText(cfg.get('to_text', ''))
        if hasattr(self, 'cc_edit') and cfg.get('cc_text') is not None:
            self.cc_edit.setText(cfg.get('cc_text', ''))
        if hasattr(self, 'bcc_edit') and cfg.get('bcc_text') is not None:
            self.bcc_edit.setText(cfg.get('bcc_text', ''))
        if hasattr(self, 'from_edit'):
            self.from_edit.setText(cfg.get('from_text', ''))
        if hasattr(self, 'reply_edit'):
            self.reply_edit.setText(cfg.get('reply_text', ''))
        if hasattr(self, 'return_edit'):
            self.return_edit.setText(cfg.get('return_text', ''))
        if hasattr(self, 'nickname_edit'):
            self.nickname_edit.setText(cfg.get('nickname_text', ''))
        if hasattr(self, 'title_edit'):
            self.title_edit.setText(cfg.get('title_text', ''))
        # 正文富文本：为空则清空
        if cfg.get('body_html') and hasattr(self, 'body_editor'):
            self.body_editor.setHtml(cfg.get('body_html', ''))
        # ---- 抄送/密送行展开 ----
        if hasattr(self, 'cc_row_widget') and cfg.get('cc_visible'):
            self.cc_row_widget.setVisible(True)
        if hasattr(self, 'bcc_row_widget') and cfg.get('bcc_visible'):
            self.bcc_row_widget.setVisible(True)
        # ---- 附件模式 ----
        fixed = cfg.get('attach_fixed_mode', True)
        if hasattr(self, 'attach_fixed_radio'):
            self.attach_fixed_radio.setChecked(fixed)
            if hasattr(self, 'attach_variable_radio'):
                self.attach_variable_radio.setChecked(not fixed)
        # 固定附件：写回 attachment_paths 与列表控件
        self.attachment_paths = list(cfg.get('attachment_paths', []) or [])
        if hasattr(self, 'attachment_list'):
            self.attachment_list.clear()
            for p in self.attachment_paths:
                self.attachment_list.addItem(os.path.basename(p) if p else p)
        # 不固定附件：写回文件夹列表
        self.attach_folders = list(cfg.get('attach_folders', []) or [])
        if hasattr(self, 'attach_folder_list'):
            self.attach_folder_list.clear()
            for f in self.attach_folders:
                self.attach_folder_list.addItem(f)
        # ---- 内嵌图片映射 ----
        self.inline_image_paths = dict(cfg.get('inline_image_paths', {}) or {})
        # ---- 负载绑定 ----
        self._load_map = dict(cfg.get('load_map', {}) or {})
        self._marker_keys = list(cfg.get('marker_keys', []) or [])
        self._marker_names = dict(cfg.get('marker_names', {}) or {})
        # ---- 已生成预览邮件 ----
        self._preview_mails = list(cfg.get('preview_mails', []) or [])
        self._preview_all_generated = cfg.get('preview_all_generated', False)
        self._preview_rows_total = cfg.get('preview_rows_total', 0)
        # 负载状态已就绪：此后进入负载/预览页不再重新初始化
        self._load_state_initialized = True
        self._markers_loaded = True
        # 回到配置页并刷新变量列表
        if hasattr(self, 'step_stack'):
            self.step_stack.setCurrentIndex(0)
            self._refreshIndicator(1)
            self._refreshFooter(1)
        self._updateVariableList()

    def onSaveDraft(self):
        """（高级模式重写）存草稿：把本任务全部可复现配置写入「草稿箱」任务

        覆盖父类普通模式存草稿逻辑：高级模式以「任务」为单位保存，
        记录任务名与完整配置快照，恢复时能复现配置/负载/预览等全部过程。
        """
        name = self._currentTaskName()
        name = name or '高级模式任务%d' % (self.db.countTasks() + 1)
        cfg = self._serializeTaskConfig()
        # 新任务写一条草稿任务；若当前是草稿恢复编辑的任务则更新原任务（不新增）
        if getattr(self, '_current_task_id', None) is not None \
                and self.db.getTask(self._current_task_id) is not None:
            self.db.updateTask(self._current_task_id, name=name, config=cfg,
                               category=CATEGORY_DRAFT, mail_count=0)
        else:
            self._current_task_id = self.db.insertTask(
                CATEGORY_DRAFT, name=name, config=cfg, mail_count=0)
        QMessageBox.information(self, '保存成功', '任务「%s」已存入草稿箱，可到草稿箱恢复继续编辑。' % name)
        self.on_back()

    def buildUi(self):
        """构建高级模式「配置页」整体布局

        配置页为四步向导（配置→负载→预览→发信）的第 1 步。
        左侧纵向堆叠各配置区（顶部操作栏 / 四步进度 / 收件人 / 昵称 / 主题 / 附件 / 正文）；
        右侧为常驻「变量标记」面板（选中文字 → 蓝色「标记选中为变量」按钮 → 包裹为 $变量名$）。
        底部提供「下一步」跳转到负载配给页。本轮只搭 UI 骨架，功能逻辑待后续接入。
        """
        # 页面整体纵向布局
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 12)
        root.setSpacing(10)
        self.setStyleSheet(CONFIG_STYLE)

        # 1. 顶部操作栏（返回 / 存草稿 / 预览等；发送/附件/设置按钮此页隐藏，发送在第4步）
        root.addLayout(self.buildTopBar())
        self.back_btn.setVisible(True)
        if getattr(self, 'send_btn', None):
            self.send_btn.setVisible(False)          # 配置页不发送
        if getattr(self, 'attach_menu_btn', None):
            self.attach_menu_btn.setVisible(False)    # 附件在下方分区配置
        if getattr(self, 'setting_menu_btn', None):
            self.setting_menu_btn.setVisible(False)   # 设置移至后续页

        # 2. 四步进度指示器（存引用，切步时动态更新高亮）
        self._indicator_host = self.buildStepIndicator(1)
        root.addWidget(self._indicator_host)

        # 2.1 任务信息条：任务名（可自定义，默认「高级模式任务x」）+ 说明
        task_row = QHBoxLayout()
        task_row.setSpacing(8)
        task_lab = QLabel('任务名称：')
        task_lab.setStyleSheet('color:#374151; font-weight:600;')
        task_row.addWidget(task_lab)
        # 任务名输入框（默认名「高级模式任务x」，进入页面时由 _initTaskName 生成）
        self.task_name_edit = QLineEdit()
        self.task_name_edit.setPlaceholderText('高级模式任务x')
        self.task_name_edit.setMaximumWidth(340)
        self.task_name_edit.setStyleSheet(
            'QLineEdit{padding:5px 8px; border:1px solid #d0d7e2; border-radius:4px; '
            'background:#fff; font-size:13px; background-color:#fffef0;}')
        task_row.addWidget(self.task_name_edit, 0)
        task_tip = QLabel('「存草稿」可保存本任务全部配置，之后可到草稿箱恢复继续编辑。')
        task_tip.setObjectName('yellowTip')
        task_tip.setWordWrap(True)
        task_row.addWidget(task_tip, 1)
        root.addLayout(task_row)

        # 3. 主体：左=配置区，右=变量标记面板（用分割器，可自由拖动分隔大小）
        self.body_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.body_splitter.setChildrenCollapsible(False)
        self.body_splitter.setHandleWidth(6)

        # ---- 左侧配置区（放入滚动容器，避免窗口过小时内容被挤压出界）----
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        left_scroll.setStyleSheet('background:transparent;')
        left_host = QWidget()
        lv = QVBoxLayout(left_host)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(10)

        # 3.1 收件人区（收件人是负载对齐的行主键）
        lv.addWidget(self._makeSection('收件人（行主键）',
                                      '每一行收件人对应一组负载，第 k 个收件人取各组负载的第 k 条数据',
                                      lambda: self._wrapLayout1(self.buildRecipientArea())))

        # 3.2 昵称配置（可标记变量，多标记按出现顺序命名）
        self.nickname_edit = MarkableLineEdit()   # 支持变量段局部高亮
        self.nickname_edit.setPlaceholderText('发件人昵称，可选中后标记为变量（可不填）')
        lv.addWidget(self._makeSection('昵称', '支持标记变量，例如 $昵称1$、$昵称2$…',
                                       lambda: self._wrapInput(self.nickname_edit)))

        # 3.3 主题配置（可标记变量，多标记按出现顺序命名）
        lv.addWidget(self._makeSection('主题', '支持标记变量，例如 $主题1$、$主题2$…',
                                       lambda: self._wrapLayout1(self.buildTitleRow())))

        # 3.4 正文配置（富文本 + 标记变量；置于附件之前，占据更大空间）
        self.body_editor = QTextEdit()
        self.body_editor.setAcceptRichText(True)
        self.body_editor.setPlaceholderText('输入正文，选中文字后点右侧「标记选中为变量」')
        self.body_editor.setMinimumHeight(300)   # 正文编辑区更高，容纳更多内容
        self.body_editor.textChanged.connect(self._updateVariableList)
        body_section = self._makeSection('正文', '支持标记变量，仅替换文本内容，不支持图像/媒体',
                                         lambda: self._wrapPlain(self.body_editor))
        # 正文下方：富文本工具栏 + 变量操作行
        body_col = QVBoxLayout()
        body_col.setSpacing(2)
        body_col.addWidget(self.buildToolbarRow1())
        body_col.addWidget(self.buildToolbarRow2())
        body_col.addWidget(self.body_editor)
        body_btn_row = QHBoxLayout()
        remove_body_btn = QPushButton('删除光标处变量')
        remove_body_btn.setToolTip('删除光标所在位置的 $变量名$')
        remove_body_btn.clicked.connect(lambda: self._removeVariableAtCursor(self.body_editor))
        body_btn_row.addWidget(remove_body_btn)
        body_btn_row.addStretch()
        body_col.addLayout(body_btn_row)
        body_section.layout().addLayout(body_col)
        lv.addWidget(body_section)

        # 3.5 附件配置（固定 / 不固定双模式；置于正文之后，占更少空间）
        lv.addWidget(self.buildAttachmentConfig())

        lv.addStretch()
        left_scroll.setWidget(left_host)
        self.body_splitter.addWidget(left_scroll)   # 左侧：配置区

        # ---- 右侧：变量标记面板 ----
        self.body_splitter.addWidget(self._buildVariablePanel())
        # 初始宽度：左 3 / 右 2，比例自由拖拽
        self.body_splitter.setStretchFactor(0, 3)
        self.body_splitter.setStretchFactor(1, 2)
        self.body_splitter.setSizes([600, 300])

        # 3. 步骤栈：第1步配置页 → 第2步负载配给 → 第3步绑定/对应 → 第4步预览 → 第5步发信
        self.step_stack = QStackedWidget()
        root.addWidget(self.step_stack, 1)

        # 3.1 第 1 步：配置页（左侧配置区 + 右侧变量面板 = body_splitter）
        config_page = QWidget()
        config_page.setObjectName('stepPage')
        cp_layout = QVBoxLayout(config_page)
        cp_layout.setContentsMargins(0, 0, 0, 0)
        cp_layout.addWidget(self.body_splitter)
        self.step_stack.addWidget(config_page)

        # 3.2 第 2 步：负载配给页
        self.step_stack.addWidget(self._buildLoadStep())

        # 3.3 第 3 步：绑定关系与对应关系页
        self.step_stack.addWidget(self._buildRelationStep())

        # 3.4 第 4 步：预览页（按负载分段生成邮件并列表展示）
        self.step_stack.addWidget(self._buildPreviewStep())

        # 3.5 第 5 步：发信页（逐封发送 + 状态展示 + 失败重发）
        self.step_stack.addWidget(self._buildSendStep())

        # 默认停留在配置页
        self.step_stack.setCurrentIndex(0)

        # 4. 底部导航（上一步/下一步，随步骤动态更新可用状态）
        root.addLayout(self.buildStepFooter())

        # 主题/昵称输入变化时刷新变量列表
        self.title_edit.textChanged.connect(self._updateVariableList)
        if hasattr(self, 'nickname_edit'):
            self.nickname_edit.textChanged.connect(self._updateVariableList)
        self._updateVariableList()

        # ---- 关键修复：连接 selectionChanged 信号以记录最后选中的控件 ----
        # 解决 QLineEdit 失焦后 selectedText() 被清除导致判断错误的问题
        if hasattr(self, 'nickname_edit'):
            self.nickname_edit.selectionChanged.connect(self._onSelectionChanged)
        self.title_edit.selectionChanged.connect(self._onSelectionChanged)
        self.body_editor.selectionChanged.connect(self._onSelectionChanged)
        # 初始化记录变量
        self._last_selected_widget = None
        self._last_selected_text = ''
        self._last_selection_start = -1
        self._last_selection_length = 0

    # ---------------- 配置页子构件 ----------------

    def _wrapLayout1(self, layout):
        """把一个布局装入单 Widget（供 Section 复用点连接）"""
        w = QWidget()
        w.setLayout(layout)
        w.layout().setContentsMargins(0, 0, 0, 0)
        return w

    def _wrapInput(self, edit):
        """把单行输入框装入 Widget（供 Section 复用点连接）"""
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(edit, 1)
        return w

    def _wrapPlain(self, edit):
        """把正文编辑器装入 Widget（供 Section 复用点连接）"""
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(edit, 1)
        return w

    def _makeSection(self, title, hint, content_factory):
        """生成一个分组视觉区块：标题 + 提示 + 内容区

        参数：
            title<str>：区块标题
            hint<str>：区块说明文字
            content_factory<callable>：返回可加入布局的控件/布局
        """
        frame = QFrame()
        frame.setObjectName('configSection')
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(6)
        title_label = QLabel(title)
        title_label.setObjectName('sectionTitle')
        lay.addWidget(title_label)
        if hint:
            hint_label = QLabel(hint)
            hint_label.setObjectName('sectionHint')
            hint_label.setWordWrap(True)
            lay.addWidget(hint_label)
        lay.addWidget(content_factory())
        return frame

    def buildStepIndicator(self, current_step):
        """构建五步向导进度指示器

        参数：
            current_step<int>：当前所在步骤（1..5），配置=1 负载=2 绑定/对应=3 预览=4 发信=5

        返回：
            QWidget：进度条组件
        """
        steps = [('1', '邮件配置'), ('2', '负载配给'), ('3', '绑定/对应'),
                 ('4', '预览'), ('5', '发信')]
        host = QWidget()
        lay = QHBoxLayout(host)
        lay.setSpacing(6)
        lay.setContentsMargins(0, 0, 0, 0)
        self._step_pills = []          # 记录五颗胶囊，切步时用 _refreshIndicator 更新高亮
        for i, (num, name) in enumerate(steps):
            step_index = i + 1
            pill = QLabel()
            if step_index < current_step:
                pill.setText('%s  %s' % (num, name))
                pill.setObjectName('stepDone')
            elif step_index == current_step:
                pill.setText('%s  %s' % (num, name))
                pill.setObjectName('stepActive')
            else:
                pill.setText('%s  %s' % (num, name))
                pill.setObjectName('stepTodo')
            pill.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._step_pills.append(pill)
            lay.addWidget(pill, 1)
        return host

    def _refreshIndicator(self, current_step):
        """切换步骤后更新进度指示器当前高亮

        参数：
            current_step<int>：当前步骤 1..5
        """
        for i, pill in enumerate(self._step_pills):
            idx = i + 1
            obj = 'stepDone' if idx < current_step else \
                ('stepActive' if idx == current_step else 'stepTodo')
            if pill.objectName() != obj:
                pill.setObjectName(obj)
                # 强制重新应用样式（objectName 变更后需 unpolish/polish 才能刷新)
                pill.style().unpolish(pill)
                pill.style().polish(pill)

    def buildStepFooter(self):
        """构建底部导航条：左「上一步」（第1步禁用）+ 右「下一步」

        返回：
            QHBoxLayout：底部导航布局
        """
        row = QHBoxLayout()
        row.setSpacing(8)
        # 上一步：配置页为第 1 步，禁用；进入后续步骤后启用
        self.prev_step_btn = QPushButton('← 上一步')
        self.prev_step_btn.setEnabled(False)
        self.prev_step_btn.clicked.connect(self._onGoPrevStep)
        row.addWidget(self.prev_step_btn)
        row.addStretch()
        # 下一步：跳转到负载配给页（第1步→第2步）
        self.next_step_btn = QPushButton('下一步 →')
        self.next_step_btn.setObjectName('primaryBtn')
        self.next_step_btn.setMinimumWidth(120)
        self.next_step_btn.setMinimumHeight(36)
        self.next_step_btn.clicked.connect(self._onGoNextStep)
        row.addWidget(self.next_step_btn)
        return row

    def _refreshFooter(self, current_step):
        """切步后更新底部按钮状态

        参数：
            current_step<int>：当前步骤 1..5
        """
        # 第1步不能回退；其余可回退
        self.prev_step_btn.setEnabled(current_step > 1)
        # 第5步（发信页）没有“下一步”，而是“发送”
        if current_step == 5:
            self.next_step_btn.setText('发送 →')
        else:
            self.next_step_btn.setText('下一步 →')

    def _onGoNextStep(self):
        """「下一步」点击：进入下一向导步骤，并在进入负载页前收集变量生成负载列表；
        从负载页进入后续页前校验全部标记已绑定文件"""
        # 发送中禁止切换：强制停留在发信页
        if self._isSending():
            QMessageBox.information(self, '提示', '发送正在进行中，请等待完成后再切换。')
            return
        cur = self.step_stack.currentIndex()
        # 由第1步（配置页）进入第2步（负载配给页）：
        # 首次构建负载列表；之后再返回时只增不减地并入配置页新增的标记（不清除既有绑定）
        if cur == 0:
            if not self._load_state_initialized:
                self._rebuildLoadList()
                self._load_state_initialized = True
            else:
                self._syncLoadList()
        # 第2步→第3步：校验全部标记已绑定，未绑定则弹窗提示阻止前进
        if cur == 1:
            total = len(self._marker_keys)
            if total == 0:
                QMessageBox.warning(self, '配置不完整',
                                    '当前没有检测到任何标记。\n请返回配置页，在昵称/主题/正文中标记变量。')
                return
            unbound = [self._marker_names.get(k, k) for k in self._marker_keys
                       if k not in self._load_map or not self._load_map[k]]
            if unbound:
                QMessageBox.warning(self, '配置不完整',
                                    '以下 %d 个标记尚未绑定数据文件：\n\n%s\n\n'
                                    '请在「标记」下拉中选择对应标记并点击「选择文件…」完成绑定。'
                                    % (len(unbound), '\n'.join('  · ' + n for n in unbound)))
                return
        # 第3步→第4步（进入预览页）：首次进入时初始化邮件批次
        if cur == 2:
            self._initPreviewIfNeeded()
        # 第4步→第5步（进入发信页）：首次进入才构建发送列表，切走再回不重置
        if cur == 3:
            self._ensureSendTableBuilt()
        # 第5步（发信页）：按钮为「发送」，点击时触发批量发送而非继续前进
        if cur == 4:
            self._onStartBatchSend()
            return
        # 越界保护
        if cur + 1 >= self.step_stack.count():
            return
        nxt = cur + 1
        self.step_stack.setCurrentIndex(nxt)
        self._refreshIndicator(nxt + 1)
        self._refreshFooter(nxt + 1)

    def _onGoPrevStep(self):
        """「上一步」点击：返回上一向导步骤"""
        # 发送中禁止切换：强制停留在发信页
        if self._isSending():
            QMessageBox.information(self, '提示', '发送正在进行中，请等待完成后再切换。')
            return
        cur = self.step_stack.currentIndex()
        if cur <= 0:
            return
        prev = cur - 1
        self.step_stack.setCurrentIndex(prev)
        self._refreshIndicator(prev + 1)
        self._refreshFooter(prev + 1)

    def _buildSendStep(self):
        """构建第 5 步「发信」页 UI：逐步展示每个收件人邮件的发送状态

        布局：
          ① 顶部：页面标题 + 发送前的提示文案
          ② 发送目标列表：QTableWidget，列 = 状态 / 收件人 / 主题 / 附件 / 结果说明
          ③ 底部：发送汇总状态 + 重发失败按钮（仅在存在失败邮件时可用）
        返回：
            QWidget：发信页
        """
        page = QWidget()
        page.setObjectName('stepPage')
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        # ---- ① 标题 + 提示 ----
        tl = QLabel('发送')
        tl.setObjectName('loadSectionTitle')
        root.addWidget(tl)
        hint = QLabel('点击底部「发送 →」即逐封发送以下邮件。\n'
                      '每封之间自动随机间隔，避免被判定为垃圾邮件；失败的邮件将标记状态，发送后可重发。')
        hint.setObjectName('loadHint')
        hint.setWordWrap(True)
        root.addWidget(hint)

        # ---- ② 发送目标列表 ----
        self.send_table = QTableWidget(0, 7)
        self.send_table.setObjectName('sendTable')
        self.send_table.setHorizontalHeaderLabels(
            ['状态', '收件人邮箱', '主题', '附件', '结果', '发送时间', '操作'])
        self.send_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.send_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.send_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.send_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        self.send_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)
        self.send_table.horizontalHeader().setSectionResizeMode(
            5, QHeaderView.ResizeMode.ResizeToContents)
        self.send_table.horizontalHeader().setSectionResizeMode(
            6, QHeaderView.ResizeMode.ResizeToContents)
        self.send_table.verticalHeader().setVisible(False)
        root.addWidget(self.send_table, 1)

        # ---- ③ 底部状态 + 重发 ----
        foot = QHBoxLayout()
        self.send_summary_label = QLabel('尚未发送')
        self.send_summary_label.setObjectName('sendHint')
        foot.addWidget(self.send_summary_label)
        foot.addStretch()
        self.retry_failed_btn = QPushButton('重发失败邮件')
        self.retry_failed_btn.setEnabled(False)
        self.retry_failed_btn.clicked.connect(self._onRetryFailed)
        foot.addWidget(self.retry_failed_btn)
        root.addLayout(foot)

        # 初始化发送控制状态（跨步骤往返均不重置）
        self._send_worker = None          # 当前发送线程
        self._send_status = []            # 每封的状态：空闲/发送中/成功/失败
        self._send_times = []             # 每封的发送时间（HH:MM:SS，发送后记录）
        self._send_ready_items = []       # 待发送项（含单封 NormalMode 构造所需字段）
        self._send_success_count = 0
        self._send_fail_count = 0
        self._send_prepared = False       # 是否已为当前预览批次构建过表（保证切换不重置）
        return page

    def _buildSendItems(self):
        """根据预览批次生成「待发送」项列表，供发信页逐封发送

        发送内容使用第 4 步生成并经用户编辑后的 _preview_mails：
        每封对应一个收件人，主题/正文/附件取自该封预览数据；每封独立构造 NormalMode 参数。
        返回：
            list<dict>：发送项（含 to_list、email_title、html_text、attachment_paths 等字段）
        """
        # 发件人/回信/退信/内嵌图等公共配置自配置页取值
        reply_to = self.reply_edit.text().strip()
        return_email = self.return_edit.text().strip()
        inline_images = dict(self.inline_image_paths) if getattr(self, 'inline_image_paths', None) else None
        cc_list = None
        bcc_list = None
        items = []
        for mail in self._preview_mails:
            # 每封发件人昵称优先用该封配置页的发件人昵称（已是变量替换后的值）；
            # 未填则回退底部 from_edit，仍空用默认占位名
            per_sender = (mail.get('nickname') or '').strip() \
                or self.from_edit.text().strip() or DEFAULT_FROM_NAME
            items.append(dict(
                n_name=per_sender,
                to_list=[mail['email']],
                email_title=mail['title'],
                html_text=mail['html'],
                attachment_paths=list(mail.get('attachments') or []),
                inline_image_paths=inline_images,
                cc_list=cc_list,
                bcc_list=bcc_list,
                reply_to=reply_to,
                return_email=return_email,
            ))
        return items

    def _ensureSendTableBuilt(self):
        """首次进入发信页时构建发送列表（切换步骤不重建，保留已发送状态）

        仅当 _send_prepared 为 False（尚未为当前预览批次构建）时才初始化；
        此后切走再回会保留已有的 成功/失败/发送时间 等固定状态。
        """
        if self._send_prepared:
            return
        mails = self._preview_mails
        self._send_prepared = True
        self.send_table.setRowCount(0)
        self.send_table.setRowCount(len(mails))
        # 状态数组：只有第一次构建时初始化为「待发送」，此后仅由发送回调改写
        if len(self._send_status) != len(mails):
            self._send_status = ['空闲'] * len(mails)
        if len(self._send_times) != len(mails):
            self._send_times = [''] * len(mails)
        for i, mail in enumerate(mails):
            # ---- 状态列 ----
            cur_state = self._getDisplayState(self._send_status[i])
            status_item = QTableWidgetItem(cur_state)
            self.send_table.setItem(i, 0, status_item)
            # ---- 收件人、主题、附件列 ----
            self.send_table.setItem(i, 1, QTableWidgetItem(mail['email']))
            self.send_table.setItem(i, 2, QTableWidgetItem(mail['title'] or '(无主题)'))
            attach_text = '%d 个附件' % len(mail.get('attachments') or []) \
                if mail.get('attachments') else '无附件'
            self.send_table.setItem(i, 3, QTableWidgetItem(attach_text))
            # ---- 结果、发送时间列 ----
            self.send_table.setItem(i, 4, QTableWidgetItem(''))
            self.send_table.setItem(i, 5, QTableWidgetItem(self._send_times[i]))
            # ---- 操作列：逐封重发按钮（默认禁用，失败后启用）----
            retry_btn = QPushButton('重发')
            retry_btn.setEnabled(False)
            retry_btn.clicked.connect(lambda _=False, row=i: self._onRetryOne(row))
            self.send_table.setCellWidget(i, 6, self._wrapRetryBtn(retry_btn, False))
            # 若此前已失败，恢复禁用态的重发按钮为可点
            if self._send_status[i] == '失败':
                self._setRowRetryEnabled(i, True)
        self._send_success_count = 0
        self._send_fail_count = 0
        self._updateSendSummary()

    @staticmethod
    def _wrapRetryBtn(btn, enabled):
        """把逐封重发按钮包进一个容器（便于居中且不随行高度拉伸）"""
        box = QWidget()
        lay = QHBoxLayout(box)
        lay.setContentsMargins(6, 2, 6, 2)
        lay.addWidget(btn)
        box.btn = btn
        box.btn_enabled = enabled
        return box

    def _getDisplayState(self, state):
        """把内部状态词转成界面展示文案

        参数：
            state<str>：内部状态（空闲/发送中/成功/失败）
        返回：
            str：界面展示文案
        """
        return {
            '空闲': '待发送',
            '发送中': '正在发送…',
            '成功': '成功',
            '失败': '失败',
        }.get(state, '待发送')

    def _setRowStatus(self, row, state, color=None):
        """设置某行状态列文案与颜色，并同步内部状态

        参数：
            row<int>：行号
            state<str>：内部状态词
            color<QColor|None>：状态文字颜色（为空则用默认灰）
        """
        self._send_status[row] = state
        item = QTableWidgetItem(self._getDisplayState(state))
        if color is not None:
            item.setForeground(color)
        self.send_table.setItem(row, 0, item)

    def _setRowRetryEnabled(self, row, enabled):
        """启用/禁用某行逐封重发按钮

        失败时按钮呈红色「重新发送」样式并允许点击，其余状态恢复默认灰态。
        参数：
            row<int>：行号
            enabled<bool>：是否启用（True=失败可重发）
        """
        box = self.send_table.cellWidget(row, 6)
        if box is None or getattr(box, 'btn', None) is None:
            return
        btn = box.btn
        btn.setEnabled(enabled)
        if enabled:
            # 失败时：红色警示 + 文案「重新发送」+ 恢复置灰的禁用态
            btn.setText('重新发送')
            btn.setStyleSheet(
                'QPushButton{background:#dc2626;color:#fff;border:0;'
                'border-radius:4px;padding:4px 10px;}'
                'QPushButton:hover{background:#b91c1c;}')
        else:
            btn.setText('重发')
            btn.setStyleSheet('')

    def _onRetryOne(self, row):
        """点某行「重发」：仅重发该行一封邮件"""
        if row < 0 or row >= len(self._send_ready_items):
            return
        if self._isSending():
            QMessageBox.information(self, '提示', '发送正在进行中，请稍候…')
            return
        item = self._send_ready_items[row]
        # 置为发送中
        self._setRowStatus(row, '发送中', QColor('#2563eb'))
        self.send_table.setItem(row, 4, QTableWidgetItem(''))
        self._send_fail_count = max(0, self._send_fail_count - 1)
        self._updateSendSummary()
        # 启动单封重发线程
        self.retry_failed_btn.setEnabled(False)
        self._setFooterSendState(True)
        self._send_worker = BatchSendWorker([item])
        self._send_worker.mail_done.connect(
            lambda sub, ok, msg, g=row: self._onMailSentFinished(g, ok, msg))
        self._send_worker.all_done.connect(self._onAllMailSent)
        self._send_worker.start()

    def _updateSendSummary(self):
        """刷新发信页底部汇总状态文本"""
        total = len(self._preview_mails)
        self.send_summary_label.setText(
            '共 %d 封 · 成功 %d · 失败 %d%s' % (
                total, self._send_success_count, self._send_fail_count,
                '' if total else ''))

    def _onStartBatchSend(self):
        """「发送」点击：构造待发送项、构建发送列表（首次）、启动批量发送线程

        发送前弹出「开始发送邮件，请耐心等待！」提示；
        发送过程中各邮件的状态实时更新。全部完成后状态固定，切走再回不重置。
        """
        if not self._preview_mails:
            QMessageBox.warning(self, '提示', '没有可发送的邮件，请先完成预览生成。')
            return
        # 防止重复点击启动多个发送线程
        if self._isSending():
            QMessageBox.information(self, '提示', '发送正在进行中，请稍候…')
            return
        self._ensureSendTableBuilt()
        # 若本次预览已发送过邮件（存在成功/失败状态），再次点击发送时二次确认，避免误重复发送
        if any(s in ('成功', '失败') for s in self._send_status):
            reply = QMessageBox.question(
                self, '发送确认',
                '检测到已发送过邮件，再次发送将重复发送这些邮件。\n\n是否继续发送？',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return  # 用户选择取消，不发送
        # 构造发送项
        self._setFooterSendState(True)
        self._send_ready_items = self._buildSendItems()
        # 弹提示
        QMessageBox.information(self, '开始发送', '开始发送邮件，请耐心等待！')
        # 启动后台批量发送线程
        self._send_worker = BatchSendWorker(self._send_ready_items)
        self._send_worker.mail_started.connect(self._onMailStarted)
        self._send_worker.mail_done.connect(self._onMailSentFinished)
        self._send_worker.all_done.connect(self._onAllMailSent)
        self._send_worker.start()

    def _onMailStarted(self, idx):
        """某行邮件开始发送回调：把状态改为「正在发送…」

        参数：
            idx<int>：行号
        """
        if idx < 0 or idx >= self.send_table.rowCount():
            return
        self._setRowStatus(idx, '发送中', QColor('#2563eb'))
        self.send_table.setItem(idx, 4, QTableWidgetItem(''))

    def _onMailSentFinished(self, idx, success, msg):
        """单封发送完成回调：更新状态列、结果列、发送时间列与逐封重发按钮

        参数：
            idx<int>：行号
            success<bool>：是否成功
            msg<str>：结果说明
        """
        if idx < 0 or idx >= self.send_table.rowCount():
            return
        # 更新状态与结果（成功/失败均固定，此后切走再回不再变动）
        state = '成功' if success else '失败'
        color = QColor('#16a34a') if success else QColor('#dc2626')
        self._setRowStatus(idx, state, color)
        self.send_table.setItem(idx, 4, QTableWidgetItem(msg))
        # 记录发送时间
        self._send_times[idx] = datetime.now().strftime('%H:%M:%S')
        self.send_table.setItem(idx, 5, QTableWidgetItem(self._send_times[idx]))
        # 失败时启用该行逐封重发按钮
        self._setRowRetryEnabled(idx, not success)
        # 计数
        if success:
            self._send_success_count += 1
        else:
            self._send_fail_count += 1
        self._updateSendSummary()

    def _recordSentTask(self, ok_cnt):
        """发送成功后把本次高级模式任务及其下成功邮件写入数据库「已发送」

        若全部失败（ok_cnt==0）不写任务记录；任务名取用户自定义或默认值。
        任务配置保存为可复现快照，任务下每封成功邮件单独落库（关联 task_id）。
        """
        if ok_cnt <= 0:
            return
        name = self._currentTaskName()
        # 取成功发送的邮件（状态为「成功」）
        sent_mails = [mail for i, mail in enumerate(self._preview_mails)
                      if i < len(self._send_status) and self._send_status[i] == '成功']
        if not sent_mails:
            return
        cfg = self._serializeTaskConfig()
        # 若当前是从草稿恢复的任务且已存在草稿任务，发送后将其改到「已发送」
        existing_id = getattr(self, '_current_task_id', None)
        if existing_id is not None and self.db.getTask(existing_id) is not None:
            task_id = existing_id
            self.db.updateTask(task_id, name=name, config=cfg,
                               category=CATEGORY_SENT, mail_count=len(sent_mails))
        else:
            task_id = self.db.insertTask(
                CATEGORY_SENT, name=name, config=cfg, mail_count=len(sent_mails))
        self._current_task_id = task_id
        # 逐封写「已发送」（收件人/主题/正文各不相同）
        for mail in sent_mails:
            self.db.insertMail(
                CATEGORY_SENT,
                title=mail.get('title') or '',
                recipient=mail.get('email') or '',
                from_name=mail.get('nickname') or '',
                to_list=[mail.get('email') or ''],
                cc_list=[], bcc_list=[],
                reply_to=self.reply_edit.text().strip(),
                return_email=self.return_edit.text().strip(),
                html_text=mail.get('html') or '',
                attachment_paths=list(mail.get('attachments') or []),
                task_id=task_id)

    def _onAllMailSent(self, ok_cnt, fail_cnt):
        """全部发送完成回调：落库已发送任务、恢复按钮、提示结果、控制批量重发可用性

        成功发送后把本任务及成功邮件写入数据库「已发送」，保证列表页可见。
        参数：
            ok_cnt<int>：成功数
            fail_cnt<int>：失败数
        """
        self._setFooterSendState(False)
        # 存在失败时启用「重发失败邮件」
        self.retry_failed_btn.setEnabled(fail_cnt > 0)
        # 本次任务写入「已发送」（至少一封成功）
        self._recordSentTask(ok_cnt)
        if fail_cnt == 0:
            QMessageBox.information(self, '全部发送成功',
                                    '共 %d 封邮件已全部发送成功！' % ok_cnt)
        else:
            QMessageBox.warning(self, '部分发送失败',
                                '成功 %d 封，失败 %d 封。可点击「重发失败邮件」重试失败的邮件。'
                                % (ok_cnt, fail_cnt))
        # 全部发送成功后清空本次预览/发送内存缓存：收件人/正文等个人数据不残留于内存；
        # 若有失败则保留内存待「重发失败邮件」。历史已由 _recordSentTask 加密落库，可从列表回看。
        if fail_cnt == 0:
            self._preview_mails = []
            self._send_ready_items = []
            self._send_status = []

    def _onRetryFailed(self):
        """「重发失败邮件」点击：仅对状态为失败的邮件再次批量发送"""
        # 收集失败项及其在完成态列表中的索引（记入重发映射，供回调定位行号）
        failed_items = []
        failed_global_rows = []
        for i, item in enumerate(self._send_ready_items):
            if self._send_status[i] == '失败':
                failed_global_rows.append(i)
                failed_items.append(item)
        if not failed_items:
            self.retry_failed_btn.setEnabled(False)
            return
        # 记录「重发项序号 -> 全局行号」映射，供重发回调写回正确行
        self._retry_row_map = failed_global_rows
        # 清理旧计数（重发后按新结果重新统计）
        self._send_fail_count = 0
        self._send_success_count = 0
        self._updateSendSummary()
        # 启动重发线程
        self._setFooterSendState(True)
        self._send_worker = BatchSendWorker(failed_items)
        self._send_worker.mail_started.connect(self._onRetryMailStarted)
        self._send_worker.mail_done.connect(self._onRetryMailFinished)
        self._send_worker.all_done.connect(self._onRetryAllFinished)
        self._send_worker.start()

    def _onRetryMailStarted(self, sub_idx):
        """重发时某封开始发送回调：把对应原行置为「正在发送…」

        参数：
            sub_idx<int>：失败项内的下标
        """
        row_map = getattr(self, '_retry_row_map', [])
        if sub_idx < 0 or sub_idx >= len(row_map):
            return
        global_row = row_map[sub_idx]
        if global_row < 0 or global_row >= self.send_table.rowCount():
            return
        self._setRowStatus(global_row, '发送中', QColor('#2563eb'))
        self.send_table.setItem(global_row, 4, QTableWidgetItem(''))

    def _onRetryMailFinished(self, sub_idx, success, msg):
        """重发时的单封完成回调：把结果写回原行

        因重发线程只携带失败项，sub_idx 为失败项内的序号；
        通过 _retry_row_map 映射回其在 send_ready_items 中的全局行号。
        参数：
            sub_idx<int>：失败项内的下标
            success<bool>：是否成功
            msg<str>：结果说明
        """
        row_map = getattr(self, '_retry_row_map', [])
        if sub_idx < 0 or sub_idx >= len(row_map):
            return
        global_row = row_map[sub_idx]
        if global_row < 0 or global_row >= self.send_table.rowCount():
            return
        state = '成功' if success else '失败'
        color = QColor('#16a34a') if success else QColor('#dc2626')
        self._setRowStatus(global_row, state, color)
        self.send_table.setItem(global_row, 4, QTableWidgetItem(msg))
        # 重发后记录最新发送时间
        self._send_times[global_row] = datetime.now().strftime('%H:%M:%S')
        self.send_table.setItem(global_row, 5, QTableWidgetItem(self._send_times[global_row]))
        # 重发仍失败时保留/启用该行逐封重发按钮，成功后禁用
        self._setRowRetryEnabled(global_row, not success)
        if success:
            self._send_success_count += 1
        else:
            self._send_fail_count += 1
        self._updateSendSummary()

    def _onRetryAllFinished(self, ok_cnt, fail_cnt):
        """重发全部完成后回调：恢复按钮并再次提示"""
        self._setFooterSendState(False)
        self.retry_failed_btn.setEnabled(fail_cnt > 0)
        if fail_cnt == 0:
            QMessageBox.information(self, '重发完成', '失败的邮件已全部重发成功！')

    def _isSending(self):
        """判断当前是否处于批量发送中（发送线程存在且在运行）

        返回：
            bool：True=发送中
        """
        return self._send_worker is not None and self._send_worker.isRunning()

    def _setFooterSendState(self, sending):
        """切换发信页底部导航按钮的交互状态（发送中全部禁用，结束恢复）

        发送期间强制停留在发信页：同时禁用「上一步」与「发送」，阻止任何步骤切换。
        参数：
            sending<bool>：True=发送中（禁用），False=可操作
        """
        self.prev_step_btn.setEnabled(not sending)
        self.next_step_btn.setEnabled(not sending)
        self.next_step_btn.setText('发送中…' if sending else '发送 →')

    def _buildLoadStep(self):
        """构建第 2 步「负载配给」页 UI（Burp Suite Intruder Payload 风格）

        布局：
          ① 顶部：页面标题 + 使用提示
          ② 标记选择区：下拉列表（列出全部标记位置）+ 配置进度
          ③ Payload设置区：左侧按钮列（选择文件/重新加载/移除绑定）+ 右侧文件内容预览
        返回：
            QWidget：负载配给页（绑定关系/对应关系在其后的第 3 步展示）
        """
        page = QWidget()
        page.setObjectName('stepPage')
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        # ---- ① 页面标题 + 提示 ----
        title_row = QHBoxLayout()
        title = QLabel('负载配给（Payload）')
        title.setObjectName('loadSectionTitle')
        title_row.addWidget(title)
        title_row.addStretch()
        root.addLayout(title_row)

        hint = QLabel('在「标记」下拉选择一个标记 → 点击「选择文件」绑定数据文件。'
                      '默认按「第 k 个收件人 = 每处标记的第 k 条数据」对齐。'
                      '请先在数据文件中排好顺序；每一处标记都必须绑定一个文件。')
        hint.setObjectName('loadHint')
        hint.setWordWrap(True)
        root.addWidget(hint)

        # ---- ② 标记选择区 ----
        mark_frame = QFrame()
        mark_frame.setObjectName('configSection')
        mark_lay = QVBoxLayout(mark_frame)
        mark_lay.setContentsMargins(12, 10, 12, 10)
        mark_lay.setSpacing(6)

        mark_title = QLabel('标记选择')
        mark_title.setObjectName('loadSubTitle')
        mark_lay.addWidget(mark_title)

        mark_row = QHBoxLayout()
        mark_row.setSpacing(10)
        mark_lbl = QLabel('标记：')
        mark_lbl.setFixedWidth(50)
        mark_row.addWidget(mark_lbl)
        # 下拉列表：动态填充全部标记位置
        self.mark_combo = QComboBox()
        self.mark_combo.setObjectName('loadMarkCombo')
        self.mark_combo.setMinimumWidth(280)
        self.mark_combo.currentIndexChanged.connect(self._onMarkComboChanged)
        mark_row.addWidget(self.mark_combo, 1)
        # 配置进度标签
        self.load_progress_label = QLabel('未检测到标记')
        self.load_progress_label.setObjectName('loadStatusPending')
        mark_row.addWidget(self.load_progress_label)
        mark_lay.addLayout(mark_row)

        root.addWidget(mark_frame)

        # ---- ③ Payload设置区（Burp 风格：左按钮列 + 右预览）----
        payload_frame = QFrame()
        payload_frame.setObjectName('loadPayloadFrame')
        payload_lay = QHBoxLayout(payload_frame)
        payload_lay.setContentsMargins(10, 10, 10, 10)
        payload_lay.setSpacing(8)

        # 左侧按钮列（Burp Payload 设置左列按钮风格）
        btn_col = QVBoxLayout()
        btn_col.setSpacing(6)
        self.pick_file_btn = QPushButton('选择文件…')
        self.pick_file_btn.setObjectName('loadPrimaryBtn')
        self.pick_file_btn.setMinimumWidth(90)
        self.pick_file_btn.clicked.connect(self._onPickCurrentMarkFile)
        btn_col.addWidget(self.pick_file_btn)

        reload_btn = QPushButton('重新加载')
        reload_btn.setObjectName('loadActionBtn')
        reload_btn.setMinimumWidth(90)
        reload_btn.clicked.connect(self._onReloadCurrentFile)
        btn_col.addWidget(reload_btn)

        self.remove_bind_btn = QPushButton('移除绑定')
        self.remove_bind_btn.setObjectName('loadActionBtn')
        self.remove_bind_btn.setMinimumWidth(90)
        self.remove_bind_btn.clicked.connect(self._onRemoveCurrentBinding)
        btn_col.addWidget(self.remove_bind_btn)

        btn_col.addStretch()
        payload_lay.addLayout(btn_col)

        # 右侧：文件内容预览（QListWidget，等宽字体）
        preview_col = QVBoxLayout()
        preview_col.setSpacing(4)
        preview_title = QLabel('文件内容预览（前 100 行）')
        preview_title.setObjectName('loadSubTitle')
        preview_col.addWidget(preview_title)

        self.load_preview_list = QListWidget()
        self.load_preview_list.setObjectName('loadPreviewList')
        self.load_preview_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.load_preview_list.setMinimumHeight(220)
        preview_col.addWidget(self.load_preview_list, 1)

        # 文件信息行（路径 + 行数）
        self.load_file_info = QLabel('尚未选择文件')
        self.load_file_info.setObjectName('loadFileInfoEmpty')
        preview_col.addWidget(self.load_file_info)

        payload_lay.addLayout(preview_col, 1)
        root.addWidget(payload_frame, 1)

        # 初始化负载映射与标记键列表
        self._load_map = {}          # key(str) -> 文件绝对路径(str)
        self._marker_keys = []       # 按出现顺序的标记 key 列表
        self._marker_names = {}      # key -> 显示名（如 "昵称标记1"）
        self._file_row_counts = {}   # key -> 文件行数(int)，供表格显示
        # 负载状态初始化标志：只在首次进入负载页时构建/刷新列表；
        # 之后停留在负载页期间（含往返其他步骤）均保留已配置内容
        self._load_state_initialized = False

        return page

    def _buildRelationStep(self):
        """构建第 3 步「绑定关系与对应关系」页 UI

        用页签收纳两张表：
          - 绑定规则：全体 标记→文件 一览（未绑定标红）
          - 对应关系展望：行 = 第 k 收件人（含邮箱号），列 = 各标记位置
        返回：
            QWidget：绑定关系与对应关系页
        """
        page = QWidget()
        page.setObjectName('stepPage')
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        # 页面标题
        title = QLabel('绑定关系与对应关系')
        title.setObjectName('loadSectionTitle')
        root.addWidget(title)

        # 用页签收纳两张表，避免同屏过度拥挤，也便于来回切换
        self.load_tabs = QTabWidget()
        self.load_tabs.setObjectName('loadTabs')
        root.addWidget(self.load_tabs, 1)

        # 页签一：绑定规则（全体 标记→文件 一览，支持清除）
        binding_host = QWidget()
        binding_lay = QVBoxLayout(binding_host)
        binding_lay.setContentsMargins(4, 6, 4, 4)
        binding_lay.setSpacing(4)
        binding_hint = QLabel('该页用于确认“每个标记对应哪个数据文件”；未绑定的标记标红。')
        binding_hint.setObjectName('sectionHint')
        binding_lay.addWidget(binding_hint)
        self.binding_table = QTableWidget(0, 5)
        self.binding_table.setObjectName('loadBindingTable')
        self.binding_table.setHorizontalHeaderLabels(['#', '标记位置', '绑定文件', '行数', '操作'])
        self.binding_table.horizontalHeader().setStretchLastSection(True)
        self.binding_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.binding_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.binding_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.binding_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.binding_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.binding_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.binding_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        binding_lay.addWidget(self.binding_table, 1)
        self.load_tabs.addTab(binding_host, '绑定规则')

        # 页签二：对应关系展望（行对齐矩阵：第 k 收件人 = 各标记第 k 条数据）
        corr_host = QWidget()
        corr_lay = QVBoxLayout(corr_host)
        corr_lay.setContentsMargins(4, 6, 4, 4)
        corr_lay.setSpacing(4)
        corr_hint = QLabel('每一行 = 第 k 条数据（第 k 个收件人，含邮箱号）；每一列 = 一个标记位置。'
                           '横向对齐：收件人A → 昵称1第1条 → 昵称2第1条 → 正文1第1条 → … → '
                           '附件文件夹①第1个 附件文件夹②第1个…）。')
        corr_hint.setObjectName('sectionHint')
        corr_hint.setWordWrap(True)
        corr_lay.addWidget(corr_hint)
        # 列 = 收件人/邮箱号 + 各标记位置 + 各附件文件夹；横向展示逐行对齐矩阵
        self.corr_table = QTableWidget(0, 0)
        self.corr_table.setObjectName('loadBindingTable')
        self.corr_table.horizontalHeader().setDefaultSectionSize(160)
        self.corr_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.corr_table.verticalHeader().setVisible(True)
        self.corr_table.verticalHeader().setDefaultSectionSize(26)
        self.corr_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.corr_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        corr_lay.addWidget(self.corr_table, 1)
        self.load_tabs.addTab(corr_host, '对应关系展望')

        return page

    # ---------------- 第 4 步：预览页 ----------------

    def _buildPreviewStep(self):
        """构建第 4 步「邮件预览」页 UI

        布局：
          - 顶部状态行：已生成 x / 总数 y + 「继续生成 n 封」按钮
          - 下方邮件列表：一封邮件一行（收件人/主题/昵称/附件），双击查看详情
        返回：
            QWidget：预览页
        """
        page = QWidget()
        page.setObjectName('stepPage')
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        # ---- 顶部：标题 + 状态 ----
        title = QLabel('邮件预览')
        title.setObjectName('loadSectionTitle')
        root.addWidget(title)

        status_row = QHBoxLayout()
        status_hint = QLabel('每行一封邮件；双击可点开查看并编辑。')
        status_hint.setObjectName('sectionHint')
        status_row.addWidget(status_hint)
        status_row.addStretch()
        # 生成进度文本
        self.preview_status_label = QLabel('尚未生成')
        self.preview_status_label.setObjectName('loadStatusPending')
        status_row.addWidget(self.preview_status_label)
        # 继续生成按钮（初始隐藏，总数量 > 当前已生成时显示）
        self.gen_more_btn = QPushButton('继续生成 %d 封' % self.PREVIEW_BATCH_SIZE)
        self.gen_more_btn.setObjectName('loadPrimaryBtn')
        self.gen_more_btn.clicked.connect(self._onGenerateMorePreview)
        self.gen_more_btn.setVisible(False)
        status_row.addWidget(self.gen_more_btn)
        root.addLayout(status_row)

        # ---- 邮件列表：列 = 新邮件标记(绿点) / 序号 / 收件人 / 主题 / 昵称 / 附件 ----
        self.preview_table = QTableWidget(0, 5)
        self.preview_table.setObjectName('previewTable')
        self.preview_table.setHorizontalHeaderLabels(
            ['', '收件人邮箱', '邮件主题', '昵称', '附件信息'])
        self.preview_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.preview_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents)
        self.preview_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.preview_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.ResizeToContents)
        self.preview_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)
        # 序号列：默认启用，若需删除行可动态调整
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        # 双击行 → 打开该邮件详情编辑对话框
        self.preview_table.cellDoubleClicked.connect(self._onPreviewRowDoubleClick)
        root.addWidget(self.preview_table, 1)

        # 预览批次状态
        self._preview_mails = []      # 已生成的邮件 dict 列表
        self._preview_generated = 0   # 已生成的封数
        self._preview_total = 0       # 邮件总封数
        self._preview_initialized = False   # 预览批次是否已初始化（切换页面不重复重置）

        return page

    def _generateOneMail(self, row):
        """按负载对齐规则生成第 row 封邮件的完整内容（纯数据，不触碰 UI）

        对应关系：第 row 个收件人 ↔ 每个已绑定标记文件的第 row 条数据 ⊇
        各附件文件夹第 row 个文件。文本模板取自配置页的昵称/主题/正文，
        把其中 $$变量$$ 逐一替换为对应行的取值。
        参数：
            row<int>：0 起始的行号
        返回：
            dict：{'email','title','nickname','html','attachments','row'}
        """
        # 收件人：取第 row 个邮箱
        email_list = parseEmails(self.to_edit.text()) if hasattr(self, 'to_edit') else []
        email = email_list[row] if row < len(email_list) else ''

        # 读取每个已绑定标记对应行的原始文本数据（供变量替换）
        row_values = {}               # key -> 该行文本
        for k in self._marker_keys:
            path = self._load_map.get(k)
            if not path:
                row_values[k] = ''
                continue
            rows = self._readDataRows(path)
            row_values[k] = rows[row] if row < len(rows) else ''

        # 各模板字段文本
        nickname_tpl = getattr(self, 'nickname_edit', None)
        nickname_tpl = nickname_tpl.text() if nickname_tpl else ''
        title_tpl = self.title_edit.text()
        body_tpl = self.body_editor.toHtml()   # 正文为 HTML 富文本

        # 按「位置」替换：负载配给对每个位置（来源+序号）绑定一个数据文件，
        # 故模板中某来源的第 i 个 $$变量$$ 使用该来源 key（来源:i）绑定文件的第 row 行取值。
        nickname = self._replaceTemplate(nickname_tpl, '昵称', row_values)
        title = self._replaceTemplate(title_tpl, '主题', row_values)
        html = self._replaceTemplate(body_tpl, '正文', row_values)

        # 附件：固定附件模式用配置页统一附件；不固定模式取各文件夹第 row 个文件
        attachments = []           # list of 文件绝对路径
        if self.attach_fixed_radio.isChecked():
            attachments = list(getattr(self, 'attachment_paths', []) or [])
        else:
            folders = getattr(self, 'attach_folders', []) or []
            for folder in folders:
                names = self._scanFolderFiles(folder)
                if row < len(names):
                    attachments.append(os.path.join(folder, names[row]))
                    # 缺文件就跳过该文件夹（第 row 收件人无第 row 个文件）

        return {'email': email, 'title': title, 'nickname': nickname,
                'html': html, 'attachments': attachments, 'row': row}

    def _replaceTemplate(self, text, src_name, row_values):
        """中某来源模板中的每个 $$变量$$ 替换为该来源对应位置第 row 行的取值

        某来源（昵称/主题/正文）中第 i 个 $$变量$$ 对应该来源 key「来源:i」，
        用其绑定文件当前行的取值替换；未绑定或空时保留原占位符。
        参数：
            text<str>：该来源模板文本
            src_name<str>：来源标识（昵称/主题/正文），与 _collectMarkers 的 key 前缀一致
            row_values<dict>：key -> 该位置当前行取值
        返回：
            str：替换后的文本
        """
        counter = [0]    # 用列表模拟可变计数器（某来源内第几个变量）
        def repl(m):
            counter[0] += 1
            val = row_values.get('%s:%d' % (src_name, counter[0]))
            return val if val else m.group(0)
        return self.VARIABLE_PATTERN.sub(repl, text)

    def _onPreviewRowDoubleClick(self, row, _col):
        """双击某个邮件行 → 打开详情查看/编辑对话框

        编辑后的内容写回 self._preview_mails[row]（仅预览批次，不回写负载文件）。
        参数：
            row<int>：被双击的行号
            _col<int>：列号（忽略）
        """
        if row < 0 or row >= len(self._preview_mails):
            return
        mail = self._preview_mails[row]
        self._openMailDetailDialog(mail)

    def _openMailDetailDialog(self, mail):
        """弹出某封邮件的详情查看/编辑对话框"""
        dlg = PreviewMailDialog(self, mail)
        dlg.exec()

    def _onGenerateMorePreview(self):
        """「继续生成」点击：生成并追加下一批邮件到列表"""
        self._fillPreviewRows(self._preview_generated)

    def _refreshPreviewStatus(self):
        """刷新预览页顶部状态文本与「继续生成」按钮显隐"""
        total = self._preview_total
        cur = self._preview_generated
        if total == 0:
            self.preview_status_label.setText('尚无邮件')
            self.gen_more_btn.setVisible(False)
            return
        self.preview_status_label.setText('已生成 %d / %d' % (cur, total))
        # 未全部生成时才可继续
        self.gen_more_btn.setVisible(cur < total)
        self.gen_more_btn.setEnabled(True)

    def _fillPreviewRows(self, start):
        """生成 [start, start+n) 区间的邮件并追加到列表，更新状态

        参数：
            start<int>：当前已生成的封数，即本批起始行号
        """
        n = self.PREVIEW_BATCH_SIZE
        num = min(n, self._preview_total - start)   # 本批实际生成数量
        # 生成本批邮件数据
        batch = [self._generateOneMail(start + i) for i in range(num)]
        self._preview_mails.extend(batch)

        # 追加到表格
        existing = self.preview_table.rowCount()
        self.preview_table.setRowCount(existing + num)
        # 新生成本批的起始序号（1 起始），用于首列绿点标记
        for i, mail in enumerate(batch):
            r = existing + i
            # 首列：绿色小圆点「●」标记本批新生成的邮件，方便区分
            dot_item = QTableWidgetItem('●')
            dot_item.setForeground(QColor('#34c759'))   # 绿色小点
            dot_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            dot_item.setToolTip('本批新生成的邮件')
            self.preview_table.setItem(r, 0, dot_item)
            email_item = QTableWidgetItem(mail['email'])
            self.preview_table.setItem(r, 1, email_item)
            title_item = QTableWidgetItem(mail['title'] or '(无主题)')
            self.preview_table.setItem(r, 2, title_item)
            nick_item = QTableWidgetItem(mail['nickname'] or '')
            self.preview_table.setItem(r, 3, nick_item)
            # 附件信息：固定模式逐个显示文件名；不固定模式显示文件数与首文件名
            if mail['attachments']:
                attach_text = '%d 个附件' % len(mail['attachments'])
            else:
                attach_text = '无附件'
            attach_item = QTableWidgetItem(attach_text)
            self.preview_table.setItem(r, 4, attach_item)

        self._preview_generated = start + num
        self._refreshPreviewStatus()

    def _refreshPreviewRowFor(self, mail):
        """保存某封邮件后，刷新预览列表中对应行的摘要显示

        参数：
            mail<dict>：含 row 字段的邮件数据
        """
        row = mail.get('row')
        if row is None or row < 0 or row >= self.preview_table.rowCount():
            return
        self.preview_table.item(row, 1).setText(mail.get('email') or '')
        self.preview_table.item(row, 2).setText(mail.get('title') or '(无主题)')
        self.preview_table.item(row, 3).setText(mail.get('nickname') or '')
        attach_text = '%d 个附件' % len(mail.get('attachments') or []) \
            if mail.get('attachments') else '无附件'
        self.preview_table.item(row, 4).setText(attach_text)

    def _initPreviewIfNeeded(self):
        """首次进入预览页时初始化：计算总封数并生成第一段（n 封）"""
        if self._preview_initialized:
            return
        # 总封数 = 收件人个数
        email_list = parseEmails(self.to_edit.text()) if hasattr(self, 'to_edit') else []
        total = len(email_list)
        self._preview_total = total
        self._preview_mails = []
        self._preview_generated = 0
        # 生成第一批
        if total > 0:
            self._fillPreviewRows(0)
        else:
            self._refreshPreviewStatus()
        self._preview_initialized = True

    def _collectMarkers(self):
        """从昵称/主题/正文收集全部标记位置（按出现顺序），构建 key 与显示名映射

        附件文件夹不属于负载标记：附件在配置页已单独配置，此处只收集文本变量标记。
        返回：
            list<tuple>：[(key, display_name), ...]，按出现顺序
        """
        markers = []
        counters = {'昵称': 0, '主题': 0, '正文': 0}
        sources = [('昵称', getattr(self, 'nickname_edit', None)),
                   ('主题', self.title_edit),
                   ('正文', self.body_editor)]
        for src_name, widget in sources:
            if widget is None:
                continue
            text = widget.toPlainText() if isinstance(widget, QTextEdit) else widget.text()
            for m in self.VARIABLE_PATTERN.finditer(text):
                counters[src_name] += 1
                key = '%s:%d' % (src_name, counters[src_name])
                display = '%s标记%d' % (src_name, counters[src_name])
                markers.append((key, display))
        return markers

    def _rebuildLoadList(self):
        """进入负载页时（首次）：收集全部标记 → 填充下拉列表 → 刷新绑定表 → 加载首个标记的预览"""
        # 清空旧数据（仅首次进入时全量构建）
        self._load_map = {}
        self._marker_keys = []
        self._marker_names = {}
        self._file_row_counts = {}

        # 收集标记
        markers = self._collectMarkers()
        self._marker_keys = [k for k, _ in markers]
        self._marker_names = {k: n for k, n in markers}

        # 填充下拉列表
        self._fillMarkerCombo(markers)

        # 刷新绑定表
        self._updateBindingTable()
        # 更新进度
        self._updateProgress()

        # 加载首个标记的预览（如有）
        if markers:
            self._loadMarkerPreview(markers[0][0])
        else:
            self._clearPreview()

    def _syncLoadList(self):
        """增量同步负载页：只增不减——保留既有标记及已绑定文件，仅并入配置页新增的标记

        从配置页新增标记后，再次进入负载页时调用；不会重置/清空既有绑定，
        只把「新出现的标记」并入下拉与绑定表（默认未绑定，待用户配置）。
        """
        markers = self._collectMarkers()
        new_names = {k: n for k, n in markers}
        # 已有标记与绑定（顺序与已绑定文件均保留）
        existing_keys = list(getattr(self, '_marker_keys', []) or [])
        merged_names = dict(getattr(self, '_marker_names', {}) or {})
        merged_names.update(new_names)
        # 只追加本次新增的标记 key（既有 key 一律保留，不做减法）
        seen = set(existing_keys)
        for k, n in markers:
            if k not in seen:
                existing_keys.append(k)
                merged_names[k] = n
                seen.add(k)
        self._marker_keys = existing_keys
        self._marker_names = merged_names
        # 重建下拉：绑定文件保存在 _load_map 中，这里不会丢失
        combo_items = [(k, self._marker_names.get(k, k)) for k in self._marker_keys]
        self._fillMarkerCombo(combo_items)
        self._updateBindingTable()
        self._updateProgress()
        # 有新增标记时，加载首个预览便于查看绑定状态；否则保持当前选中项
        idx = self.mark_combo.currentIndex()
        if idx < 0 and self._marker_keys:
            self._loadMarkerPreview(self._marker_keys[0])

    def _fillMarkerCombo(self, markers):
        """按 (key, name) 列表填充标记下拉框，并联动启用/禁用

        参数：
            markers<list<tuple>>：[(key, display_name), ...]；为空时展示占位项并禁用
        """
        self.mark_combo.blockSignals(True)
        self.mark_combo.clear()
        if not markers:
            self.mark_combo.addItem('（未检测到标记）')
            self.mark_combo.setEnabled(False)
        else:
            self.mark_combo.setEnabled(True)
            for key, name in markers:
                self.mark_combo.addItem(name, key)
        self.mark_combo.blockSignals(False)

    def _onMarkComboChanged(self, index):
        """下拉列表切换标记：加载该标记的文件绑定与预览内容

        参数：
            index<int>：下拉框当前索引
        """
        if index < 0:
            self._clearPreview()
            return
        key = self.mark_combo.itemData(index)
        if key is None:
            self._clearPreview()
            return
        self._loadMarkerPreview(key)

    def _loadMarkerPreview(self, key):
        """为指定标记 key 加载已绑定文件的预览内容到右侧列表

        参数：
            key<str>：标记键（如 '昵称:1' / 'body:2' / 'folder:0'）
        """
        path = self._load_map.get(key)
        if path and os.path.isfile(path):
            self._previewFile(path, key)
        else:
            self._clearPreview()

    def _clearPreview(self):
        """清空右侧预览列表与文件信息（未绑定状态）"""
        self.load_preview_list.clear()
        self.load_file_info.setText('尚未选择文件')
        self.load_file_info.setObjectName('loadFileInfoEmpty')
        self.load_file_info.style().unpolish(self.load_file_info)
        self.load_file_info.style().polish(self.load_file_info)
        self.pick_file_btn.setEnabled(True)

    def _previewFile(self, path, key):
        """读取指定文件的前 100 行内容，填充到预览列表，并更新文件信息

        参数：
            path<str>：文件绝对路径
            key<str>：标记键（用于更新行数记录）
        """
        self.load_preview_list.clear()
        ext = os.path.splitext(path)[1].lower()
        rows = 0
        max_preview = 100
        try:
            if ext in ('.csv', '.txt'):
                # CSV/TXT：逐行读取，跳过无意义的空行/纯空白行，与 _readDataRows 统计一致
                with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                    for line in fh:
                        if not line.strip():
                            continue   # 空白行不计入有效数据行
                        if rows < max_preview:
                            item = QListWidgetItem(line.rstrip('\n\r'))
                            self.load_preview_list.addItem(item)
                        rows += 1
            elif ext in ('.xlsx', '.xls'):
                # Excel：尝试用 openpyxl 读取第一工作表
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    for row_data in ws.iter_rows(values_only=True):
                        joined = '\t'.join([str(c) if c is not None else '' for c in row_data])
                        if not joined.strip():
                            continue   # 整行为空时不计数
                        if rows < max_preview:
                            item = QListWidgetItem(joined)
                            self.load_preview_list.addItem(item)
                        rows += 1
                    wb.close()
                except ImportError:
                    item = QListWidgetItem('（未安装 openpyxl，无法预览 Excel 文件）')
                    self.load_preview_list.addItem(item)
                    rows = 0
            else:
                item = QListWidgetItem('（不支持的文件格式：%s）' % ext)
                self.load_preview_list.addItem(item)
        except Exception as exc:
            item = QListWidgetItem('（读取文件失败：%s）' % exc)
            self.load_preview_list.addItem(item)

        # 更新行数记录
        self._file_row_counts[key] = rows
        # 更新文件信息标签
        file_name = os.path.basename(path)
        if rows > max_preview:
            info_text = '%s  |  共 %d 行（仅预览前 %d 行）' % (file_name, rows, max_preview)
        else:
            info_text = '%s  |  共 %d 行' % (file_name, rows)
        self.load_file_info.setText(info_text)
        self.load_file_info.setObjectName('loadFileInfo')
        self.load_file_info.style().unpolish(self.load_file_info)
        self.load_file_info.style().polish(self.load_file_info)

    def _onPickCurrentMarkFile(self):
        """为当前下拉选中的标记选择数据文件（CSV/TXT/Excel），绑定并刷新预览与绑定表"""
        if not self._marker_keys:
            QMessageBox.information(self, '提示', '当前没有可绑定的标记，请先在配置页标记变量。')
            return
        index = self.mark_combo.currentIndex()
        if index < 0:
            return
        key = self.mark_combo.itemData(index)
        if key is None:
            return
        # 打开文件选择对话框
        path, _ = QFileDialog.getOpenFileName(
            self, '选择负载数据文件（CSV/TXT/Excel）', '', '数据文件 (*.csv *.txt *.xlsx *.xls)')
        if not path:
            return
        # 绑定到当前标记
        self._load_map[key] = path
        # 刷新预览
        self._previewFile(path, key)
        # 刷新绑定表与进度
        self._updateBindingTable()
        self._updateProgress()

    def _onReloadCurrentFile(self):
        """重新加载当前标记绑定文件的预览（文件在外部被修改后刷新）"""
        index = self.mark_combo.currentIndex()
        if index < 0:
            return
        key = self.mark_combo.itemData(index)
        if key is None:
            return
        path = self._load_map.get(key)
        if not path:
            QMessageBox.information(self, '提示', '当前标记尚未绑定文件。')
            return
        self._previewFile(path, key)
        self._updateBindingTable()
        self._updateProgress()

    def _onRemoveCurrentBinding(self):
        """移除当前下拉选中标记的文件绑定，清空预览与绑定表对应行"""
        index = self.mark_combo.currentIndex()
        if index < 0:
            return
        key = self.mark_combo.itemData(index)
        if key is None:
            return
        if key not in self._load_map:
            QMessageBox.information(self, '提示', '当前标记尚未绑定文件。')
            return
        # 确认对话框
        ret = QMessageBox.question(
            self, '确认移除',
            '确定要移除「%s」的文件绑定吗？' % self._marker_names.get(key, key),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ret != QMessageBox.StandardButton.Yes:
            return
        # 移除绑定
        self._load_map.pop(key, None)
        self._file_row_counts.pop(key, None)
        self._clearPreview()
        self._updateBindingTable()
        self._updateProgress()

    def _updateBindingTable(self):
        """刷新绑定规则表：展示全部 标记→文件 绑定，未绑定的标记以红色提示"""
        self.binding_table.setRowCount(0)
        row = 0
        for key in self._marker_keys:
            self.binding_table.insertRow(row)
            # # 列
            idx_item = QTableWidgetItem(str(row + 1))
            idx_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.binding_table.setItem(row, 0, idx_item)
            # 标记位置列
            name_item = QTableWidgetItem(self._marker_names.get(key, key))
            self.binding_table.setItem(row, 1, name_item)
            # 绑定文件列
            path = self._load_map.get(key, '')
            if path:
                file_item = QTableWidgetItem(os.path.basename(path))
                file_item.setToolTip(path)
                file_item.setForeground(QColor('#1f2329'))
            else:
                file_item = QTableWidgetItem('（未绑定）')
                file_item.setForeground(QColor('#d03050'))
            self.binding_table.setItem(row, 2, file_item)
            # 行数列
            count = self._file_row_counts.get(key, 0)
            count_item = QTableWidgetItem(str(count) if path else '-')
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.binding_table.setItem(row, 3, count_item)
            # 操作列：删除按钮
            del_btn = QPushButton('删除')
            del_btn.setObjectName('loadDelBtn')
            del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            del_btn.clicked.connect(lambda _=False, k=key: self._onDeleteBindingRow(k))
            self.binding_table.setCellWidget(row, 4, del_btn)
            row += 1
        # 自适应列宽
        self.binding_table.resizeRowsToContents()

    def _naturalSortKey(self, name):
        """自然排序 key：将文件名拆成「数字段 / 非数字段」交替片段

        数字段转成 int 参与比较，非数字段保留字符串，使「10」「2」这类
        数字命名的文件按数值升序（1,2,3,…,10），而非字符串序（1,10,2,…）。
        参数：
            name<str>：文件名
        返回：
            list：供 sorted(key=...) 使用的比较键
        """
        # 用正则把数字串切分成单独片段，再对数字片段转 int
        return [int(part) if part.isdigit() else part
                for part in re.split(r'(\d+)', name)]

    def _scanFolderFiles(self, folder):
        """扫描附件文件夹中的文件（仅当前层，不递归子目录），按文件名自然排序

        依据之前约定的附件扫描算法：每个文件夹只取直接子文件，按文件名升序排列，
        第 k 个收件人取排序后第 k 个文件；文件数不足则该收件人缺该文件夹。
        参数：
            folder<str>：附件文件夹绝对路径
        返回：
            list<str>：排序后的文件名列表（不含子目录、不含目录项）
        """
        try:
            names = [f for f in os.listdir(folder)
                     if os.path.isfile(os.path.join(folder, f))]
        except OSError:
            return []
        # 自然排序：数字文件名（如 1..10）按数值排，避免「10」被排到「2」前面
        return sorted(names, key=self._naturalSortKey)

    def _updateCorrelationTable(self):
        """刷新「对应关系展望」矩阵：行 = 第 k 条数据（第 k 收件人），列 = 各标记位置 + 各附件文件夹

        列顺序：收件人/邮箱号 → 各文本标记 → 各附件文件夹（按添加顺序）。
        行取所有已绑定文件与各附件文件夹文件数的最大值；附件缺行留空。
        """
        # 清空表格
        self.corr_table.clear()
        self.corr_table.setRowCount(0)
        self.corr_table.setColumnCount(0)

        # 收集所有已绑定文件的标记 key（保持 _marker_keys 的展示顺序）
        bound_keys = [k for k in self._marker_keys if self._load_map.get(k)]
        # 收集附件文件夹（不固定附件模式下配置页添加的文件夹列表）
        folders = list(getattr(self, 'attach_folders', []) or [])

        if not bound_keys and not folders:
            self.corr_table.setColumnCount(1)
            self.corr_table.setRowCount(1)
            self.corr_table.setHorizontalHeaderLabels(['提示'])
            tip_item = QTableWidgetItem('尚未绑定任何文件/附件文件夹，无法预览对应关系。')
            tip_item.setForeground(QColor('#c0c4cc'))
            self.corr_table.setItem(0, 0, tip_item)
            return

        # 列头：序号/收件人 + 邮箱号 + 各标记位置 + 各附件文件夹
        headers = ['第 k 条数据 / 收件人', '邮箱号']
        for k in bound_keys:
            headers.append(self._marker_names.get(k, k))
        for idx in range(len(folders)):
            headers.append('附件文件夹%d' % (idx + 1))
        self.corr_table.setColumnCount(len(headers))
        self.corr_table.setHorizontalHeaderLabels(headers)

        # 读取每个已绑定标记文件的行数据
        col_rows = []          # 每个绑定 key 对应的一组行字符串
        max_row = 0
        for k in bound_keys:
            rows = self._readDataRows(self._load_map[k])
            col_rows.append(rows)
            max_row = max(max_row, len(rows))

        # 扫描每个附件文件夹的文件名列表（仅当前层、按名排序）
        folder_files = []      # 每个文件夹排序后的文件名列表
        for folder in folders:
            names = self._scanFolderFiles(folder)
            folder_files.append(names)
            max_row = max(max_row, len(names))

        if max_row == 0:
            self.corr_table.setRowCount(1)
            tip_item = QTableWidgetItem('绑定文件/附件文件夹均为空，请检查。')
            tip_item.setForeground(QColor('#c0c4cc'))
            self.corr_table.setItem(0, 1, tip_item)
            return

        self.corr_table.setRowCount(max_row)
        # 从收件人输入框解析第 k 个邮箱号（供“邮箱号”列展示，缺行留空）
        email_list = parseEmails(self.to_edit.text()) if hasattr(self, 'to_edit') else []
        for i in range(max_row):
            col = 0
            # 第 1 列：序号/收件人标签
            idx_item = QTableWidgetItem('收件人 %d' % (i + 1))
            idx_item.setForeground(QColor('#595f69'))
            idx_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.corr_table.setItem(i, col, idx_item)
            col += 1
            # 第 2 列：对应邮箱号
            email = email_list[i] if i < len(email_list) else ''
            email_item = QTableWidgetItem(email)
            email_item.setToolTip(email)
            self.corr_table.setItem(i, col, email_item)
            col += 1
            # 各标记该行数据
            for rows in col_rows:
                val = rows[i] if i < len(rows) else ''
                cell = QTableWidgetItem(val)
                cell.setToolTip(val)
                self.corr_table.setItem(i, col, cell)
                col += 1
            # 各附件文件夹：第 k 个文件（仅当前层排序后），缺文件则留空
            for names in folder_files:
                fname = names[i] if i < len(names) else ''
                file_cell = QTableWidgetItem(fname)
                file_cell.setToolTip(fname)
                self.corr_table.setItem(i, col, file_cell)
                col += 1

        # 表头与行高自适应
        self.corr_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.corr_table.verticalHeader().setDefaultSectionSize(26)

    def _readDataRows(self, path):
        """读取数据文件全部有效行（文本行），供对应关系矩阵使用

        自动过滤空白行 / 纯空格行等无意义内容，避免把多余空行误判为收件人。
        CSV/TXT 按行读取；Excel 读取第一工作表每行以制表符连接各列。
        参数：
            path<str>：文件绝对路径
        返回：
            list<str>：所有有效数据行文本（首行表头也计入，作为第 1 条数据）
        """
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in ('.csv', '.txt'):
                with open(path, 'r', encoding='utf-8', errors='replace') as fh:
                    # strip 后仍为空的整行视为无意义内容，直接跳过
                    return [line.rstrip('\n\r') for line in fh if line.strip()]
            elif ext in ('.xlsx', '.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for row_data in ws.iter_rows(values_only=True):
                    joined = '\t'.join([str(c) if c is not None else '' for c in row_data])
                    # 整行为空（所有单元格均为空字符串/None）时跳过
                    if joined.strip():
                        rows.append(joined)
                wb.close()
                return rows
            else:
                return []
        except Exception:
            return []

    def _onDeleteBindingRow(self, key):
        """绑定规则表中删除指定标记的绑定，并刷新预览与表格

        参数：
            key<str>：标记键
        """
        self._load_map.pop(key, None)
        self._file_row_counts.pop(key, None)
        # 如果删除的是当前选中的标记，清空预览
        cur_key = self.mark_combo.currentData()
        if cur_key == key:
            self._clearPreview()
        self._updateBindingTable()
        self._updateProgress()

    def _updateProgress(self):
        """刷新配置进度标签：显示 已绑定数/总数 + 完成状态颜色"""
        total = len(self._marker_keys)
        bound = sum(1 for k in self._marker_keys if k in self._load_map and self._load_map[k])
        if total == 0:
            text = '未检测到标记'
            obj = 'loadStatusPending'
        elif bound == total:
            text = '已配置 %d/%d  ✓ 全部就绪' % (bound, total)
            obj = 'loadStatusOk'
        else:
            text = '已配置 %d/%d' % (bound, total)
            obj = 'loadStatusPending'
        self.load_progress_label.setText(text)
        if self.load_progress_label.objectName() != obj:
            self.load_progress_label.setObjectName(obj)
            self.load_progress_label.style().unpolish(self.load_progress_label)
            self.load_progress_label.style().polish(self.load_progress_label)
        # 同步刷新「对应关系展望」矩阵（所有绑定修改点都会调用本方法）
        self._updateCorrelationTable()

    def _buildVariablePanel(self):
        """构建右侧「变量标记」面板（配置页右侧常驻）

        返回：
            QWidget：变量面板；含蓝色「标记选中为变量」按钮、变量列表、清空按钮
        """
        panel = QWidget()
        panel.setObjectName('varPanel')
        right_panel = QVBoxLayout(panel)
        right_panel.setContentsMargins(14, 12, 14, 12)
        right_panel.setSpacing(8)

        # 面板标题
        head = QLabel('变量标记（Payload）')
        head.setObjectName('varTitle')
        right_panel.addWidget(head)
        tip = QLabel('在主题/正文/昵称中选中文字 → 点下方按钮包裹为 $$变量$$（两个美元符包裹）')
        tip.setObjectName('varHint')
        tip.setWordWrap(True)
        right_panel.addWidget(tip)

        # 蓝色「标记选中为变量」按钮
        mark_btn = QPushButton('标记选中为变量')
        mark_btn.setToolTip('将主题/正文/昵称中选中的文字包裹为 $$变量$$\n'
                            '（自动作用于当前聚焦或含选中文字的输入框）')
        mark_btn.setObjectName('varMarkBtn')
        mark_btn.setMinimumHeight(34)
        mark_btn.clicked.connect(self._markVariableFromPanel)
        right_panel.addWidget(mark_btn)

        # 变量列表
        self.variable_list = QListWidget()
        self.variable_list.setObjectName('varList')
        right_panel.addWidget(self.variable_list, 1)

        # 批量操作行：删除选中 / 清空所有
        var_btn_row = QHBoxLayout()
        remove_btn = QPushButton('删除光标处')
        remove_btn.clicked.connect(self._removeBodyVariable)
        var_btn_row.addWidget(remove_btn)
        self.clear_all_btn = QPushButton('清空所有变量')
        self.clear_all_btn.clicked.connect(self._clearAllVariables)
        var_btn_row.addWidget(self.clear_all_btn)
        right_panel.addLayout(var_btn_row)

        # 说明提示
        self.data_hint_label = QLabel('变量仅在发送时替换为对应负载数据；\n请确保每个变量都已配置负载。')
        self.data_hint_label.setObjectName('varNote')
        self.data_hint_label.setWordWrap(True)
        right_panel.addWidget(self.data_hint_label)

        right_panel.addStretch()
        return panel

    def _removeBodyVariable(self):
        """删除正文中光标处的变量（右侧面板批量操作入口）"""
        self._removeVariableAtCursor(self.body_editor)

    def buildAttachmentConfig(self):
        """构建附件配置区：支持「固定附件 / 不固定附件」双模式切换

        返回：
            QWidget：附件配置组件
        """
        frame = self._makeSection('附件', '固定=所有收件人相同；不固定=第 k 收件人取每个文件夹第 k 个文件', lambda: QWidget())

        # 模式切换行
        mode_row = QHBoxLayout()
        self.attach_fixed_radio = QRadioButton('固定附件')
        self.attach_fixed_radio.setChecked(True)
        self.attach_variable_radio = QRadioButton('不固定附件')
        mode_row.addWidget(self.attach_fixed_radio)
        mode_row.addWidget(self.attach_variable_radio)
        mode_row.addStretch()
        # 放入同一互斥组，确保两个单选按钮只能二选一（圆圈不消失、状态互斥）
        self.attach_mode_group = QButtonGroup(self)
        self.attach_mode_group.addButton(self.attach_fixed_radio, 0)
        self.attach_mode_group.addButton(self.attach_variable_radio, 1)
        # 默认选中「固定」
        self.attach_fixed_radio.setChecked(True)
        mode_row_widget = QWidget()
        mode_row_widget.setLayout(mode_row)
        frame.layout().addWidget(mode_row_widget)

        # 双模式内容堆叠
        self.attach_stack = QStackedWidget()
        frame.layout().addWidget(self.attach_stack)

        # 固定模式：列表 + 计数 + 添加/删除按钮
        fixed_page = QWidget()
        fixed_lay = QVBoxLayout(fixed_page)
        fixed_lay.setContentsMargins(0, 0, 0, 0)
        fixed_head = QHBoxLayout()
        fixed_head.addWidget(QLabel('全部收件人共用以下附件：'))
        fixed_head.addStretch()
        count_label = QLabel()
        self.attachment_count_label = count_label
        fixed_head.addWidget(count_label)
        fixed_lay.addLayout(fixed_head)
        self.attachment_list = QListWidget()
        self.attachment_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.attachment_list.setMaximumHeight(60)   # 附件列表更矮，压缩附件区高度
        fixed_lay.addWidget(self.attachment_list)
        fixed_btn_row = QHBoxLayout()
        add_fixed_btn = QPushButton('添加附件…')
        add_fixed_btn.clicked.connect(self.onAddAttachment)
        fixed_btn_row.addWidget(add_fixed_btn)
        remove_fixed_btn = QPushButton('删除选中附件')
        remove_fixed_btn.clicked.connect(self.onRemoveAttachment)
        fixed_btn_row.addWidget(remove_fixed_btn)
        fixed_btn_row.addStretch()
        fixed_lay.addLayout(fixed_btn_row)

        # 不固定模式：文件夹列表 + 说明（顺序提示黄字）
        variable_page = QWidget()
        variable_lay = QVBoxLayout(variable_page)
        variable_lay.setContentsMargins(0, 0, 0, 0)
        folder_tip = QLabel('⚠ 第 k 收件人取各文件夹第 k 个文件。请先在文件夹内按期望顺序排好文件，'
                            '程序按读取顺序一一对应，后续可在负载页拖拽调整。')
        folder_tip.setObjectName('yellowTip')
        folder_tip.setWordWrap(True)
        variable_lay.addWidget(folder_tip)
        self.attach_folder_list = QListWidget()
        self.attach_folder_list.setMaximumHeight(60)   # 与固定模式对齐，压缩附件区高度
        variable_lay.addWidget(self.attach_folder_list)
        folder_btn_row = QHBoxLayout()
        add_folder_btn = QPushButton('添加文件夹…')
        add_folder_btn.clicked.connect(self.onAddFolder)
        folder_btn_row.addWidget(add_folder_btn)
        folder_btn_row.addStretch()
        variable_lay.addLayout(folder_btn_row)

        self.attach_stack.addWidget(fixed_page)
        self.attach_stack.addWidget(variable_page)

        # 模式切换联动：固定=index0，不固定=index1
        self.attach_fixed_radio.toggled.connect(
            lambda checked: self.attach_stack.setCurrentIndex(0 if checked else 1))
        return frame

    def onAddFolder(self):
        """「不固定附件」添加文件夹：弹出目录选择，将路径加入文件夹列表并保存

        后续负载配给页可读取 self.attach_folders 用于第 k 收件人与每文件夹第 k 个文件的对应。
        """
        folder = QFileDialog.getExistingDirectory(self, '选择附件文件夹（第 k 收件人取其中第 k 个文件）')
        if not folder:
            return
        # 保存选中的文件夹路径，供第2步负载配给页使用
        if not hasattr(self, 'attach_folders'):
            self.attach_folders = []
        self.attach_folders.append(folder)
        self.attach_folder_list.addItem(folder)

    # ---------------- 变量标记逻辑 ----------------

    def _markVariableFromPanel(self):
        """从右侧面板统一触发变量标记：自动判定目标输入框（主题或正文）

        优先级：含选中文字者 > 当前获焦的输入框 > 默认正文编辑器。
        点击按钮会造成焦点转移，故依据「选中文字」而非焦点来判断更可靠。
        """
        target = self._targetVariableWidget()
        if target is not None:
            self._markSelectionAsVariable(target)

    def _targetVariableWidget(self):
        """返回本次应作用的目标输入框（QLineEdit=主题/昵称，QTextEdit=正文）

        核心修复：QLineEdit 失焦时选中状态会被清除，故通过 _last_selected_widget 记录最后一次
        （selectionChanged 信号触发时）有选中内容的控件，避免点击按钮转移焦点后误判。

        返回：
            widget: 目标输入控件；无合适目标时返回 self.body_editor
        """
        # 1. 优先选取含选中文字的输入框（理论上点击按钮瞬间已失效，但先尝试）
        for w in (self.title_edit, self.body_editor,
                  getattr(self, 'nickname_edit', None)):
            if w is None or not hasattr(w, 'selectedText'):
                continue
            if isinstance(w, QLineEdit):
                if w.selectedText():
                    return w
            elif w.textCursor().hasSelection():
                return w

        # 2. 关键修复：检查上一次触发 selectionChanged 的控件（记录了失焦前的选中状态）
        if hasattr(self, '_last_selected_widget') and self._last_selected_widget is not None:
            w = self._last_selected_widget
            if isinstance(w, QLineEdit):
                # 虽然失焦清空了 selectedText，但我们知道它刚才被选中过
                # 此时我们需要通过它的 text() 和一个内部记录的选区来判断，但更简单的是：
                # 如果是 QLineEdit 且它是最后一个选中的，我们就认为它是目标
                return w
            elif isinstance(w, QTextEdit):
                # 对于 QTextEdit，失焦后 hasSelection() 仍可能为 True（PyQt 行为）
                # 但如果为 False，我们也通过 _last_selected_widget 来认定它
                return w

        # 3. 其次取当前获焦的输入框
        focused = self.focusWidget()
        if focused in (self.title_edit, self.body_editor,
                       getattr(self, 'nickname_edit', None)):
            return focused
            
        # 4. 最终 fallback
        return self.body_editor

    def _onSelectionChanged(self):
        """槽函数：当任一文本框有选中内容时，记录该控件及选区信息。

        只在「确有选中文本」时更新记录。QLineEdit 失焦会自动清空选中并再次触发本信号，
        若不做判断会把已记录的文字覆盖为空，导致误插入 $变量名$ 占位符。
        """
        sender = self.sender()
        if sender in (self.title_edit, self.body_editor,
                      getattr(self, 'nickname_edit', None)):
            if isinstance(sender, QLineEdit):
                # 仅当确有选中文字时才记录（失焦清空的信号不覆盖）
                if not sender.selectedText():
                    return
                self._last_selected_widget = sender
                self._last_selection_start = sender.selectionStart()
                self._last_selection_length = sender.selectionLength()
                self._last_selected_text = sender.selectedText()
            elif isinstance(sender, QTextEdit):
                # 仅当确有选中文字时才记录
                cursor = sender.textCursor()
                if not cursor.hasSelection():
                    return
                self._last_selected_widget = sender
                self._last_selected_text = cursor.selectedText()
                self._last_selection_start = cursor.selectionStart()
                self._last_selection_length = cursor.selectionEnd() - cursor.selectionStart()

    def _markSelectionAsVariable(self, widget):
        """将输入框/编辑器中的选中文字包裹为 $变量名$ 并设置背景高亮（类似 Burp Suite 风格）

        参数：
            widget<QLineEdit|QTextEdit>：目标输入控件
        """
        if isinstance(widget, QLineEdit):
            # 核心修复：如果 QLineEdit 失焦导致 selectedText() 返回空，
            # 则使用 _onSelectionChanged 中记录的精确选区信息
            selected = widget.selectedText()
            if not selected and hasattr(self, '_last_selected_widget') and self._last_selected_widget is widget:
                selected = getattr(self, '_last_selected_text', '')
                # 获取记录的精确位置
                start = getattr(self, '_last_selection_start', -1)
                length = getattr(self, '_last_selection_length', 0)
            else:
                start = widget.selectionStart()
                length = len(selected)

            if not selected:
                # 未选中文字时，在光标处插入空变量占位
                pos = widget.cursorPosition()
                text = widget.text()
                new_text = text[:pos] + '$$变量名$$' + text[pos:]
                widget.setText(new_text)
                widget.setCursorPosition(pos + 2)  # 光标移到第1对 $$ 后
            else:
                text = widget.text()
                # 如果 start 无效（例如 -1），则回退到查找
                if start < 0 or length <= 0:
                    start = text.find(selected)
                    if start == -1:
                        start = widget.cursorPosition() - len(selected)
                        if start < 0:
                            start = 0
                end = start + length
                new_text = text[:start] + '$$' + selected + '$$' + text[end:]
                widget.setText(new_text)
        elif isinstance(widget, QTextEdit):
            cursor = widget.textCursor()
            # 变量仅用 $$...$$ 符号标记，不再修改文字背景颜色，
            # 避免富文本背景色残存到预览生成的邮件正文中

            # 关键修复：点击按钮导致 QTextEdit 失焦后 hasSelection() 返回 False，
            # 此时用 _onSelectionChanged 中记录的选区委据恢复选区，再执行包裹逻辑
            if not cursor.hasSelection() and hasattr(self, '_last_selected_widget') and self._last_selected_widget is widget:
                start = getattr(self, '_last_selection_start', -1)
                length = getattr(self, '_last_selection_length', 0)
                if start >= 0 and length > 0:
                    cursor.setPosition(start)
                    cursor.setPosition(start + length, QTextCursor.MoveMode.KeepAnchor)
                    widget.setTextCursor(cursor)  # 恢复编辑器上的选区，保证 hasSelection() 为真

            if not cursor.hasSelection():
                # 无选中内容时：插入 $$变量名$$
                cursor.insertText('$$变量名$$')
            else:
                # 有选中内容时：包裹 $$...$$（不修改背景，仅插入符号）
                selected = cursor.selectedText()
                cursor.beginEditBlock()
                cursor.removeSelectedText()
                cursor.insertText('$$')
                cursor.insertText(selected)
                cursor.insertText('$$')
                cursor.endEditBlock()
        self._updateVariableList()

    def _removeVariableAtCursor(self, widget):
        """删除光标所在位置的 $变量名$ 标记，并清除高亮背景

        参数：
            widget<QLineEdit|QTextEdit>：目标输入控件
        """
        if isinstance(widget, QLineEdit):
            pos = widget.cursorPosition()
            text = widget.text()
            # 查找光标附近的 $...$
            match = self._findVariableAround(text, pos)
            if match:
                start, end = match.span()
                new_text = text[:start] + text[end:]
                widget.setText(new_text)
        elif isinstance(widget, QTextEdit):
            cursor = widget.textCursor()
            pos = cursor.position()
            text = widget.toPlainText()
            match = self._findVariableAround(text, pos)
            if match:
                start, end = match.span()
                cursor.setPosition(start)
                cursor.setPosition(end, QTextCursor.MoveMode.KeepAnchor)
                # 清除选中区域的文本和格式
                default_fmt = QTextCharFormat()
                cursor.setCharFormat(default_fmt)  # 先清除格式
                cursor.removeSelectedText()         # 再删除文本
        self._updateVariableList()

    def _findVariableAround(self, text, pos):
        """查找指定位置附近的 $...$ 匹配

        参数：
            text<str>：完整文本
            pos<int>：光标位置

        返回：
            re.Match|None：匹配对象或 None
        """
        for match in self.VARIABLE_PATTERN.finditer(text):
            start, end = match.span()
            if start <= pos <= end:
                return match
        return None

    def _clearAllVariables(self):
        """清空主题/正文/昵称中的所有 $...$ 标记（保留变量名本身）"""
        # 正文先清除全部字符格式（含荧光笔背景/前景色），避免清空后黄色背景残存
        body_cur = self.body_editor.textCursor()
        body_cur.select(QTextCursor.SelectionType.Document)
        body_cur.setCharFormat(QTextCharFormat())   # 清空全文格式
        self.body_editor.setTextCursor(body_cur)
        # 主题输入框为 title_edit（来自 NormalPage.buildTitleRow）
        self.title_edit.setText(self.VARIABLE_PATTERN.sub(lambda m: m.group(1), self.title_edit.text()))
        self.body_editor.setPlainText(self.VARIABLE_PATTERN.sub(lambda m: m.group(1), self.body_editor.toPlainText()))
        # 昵称输入框（存在时一并清理）
        if hasattr(self, 'nickname_edit') and self.nickname_edit.text():
            self.nickname_edit.setText(
                self.VARIABLE_PATTERN.sub(lambda m: m.group(1), self.nickname_edit.text()))
        self._updateVariableList()

    def _updateVariableList(self):
        """解析主题/正文/昵称中的所有 $变量$ 并更新右侧变量列表

        不去重：同一变量名在多个位置（或同一位置多处）标记都会各占一条，
        并用「来源」前缀予以区分，方便后续负载配置按位置一一对应。
        """
        self.variable_list.clear()
        # 依次收集各来源的变量；每个匹配逐条入列，不做名称去重
        self._collectVariables(self.title_edit.text(), '主题')
        self._collectVariables(self.body_editor.toPlainText(), '正文')
        if hasattr(self, 'nickname_edit'):
            self._collectVariables(self.nickname_edit.text(), '昵称')
        # 触发主题/昵称重绘，使 MarkableLineEdit 的变量段局部高亮实时更新
        self.title_edit.update()
        if hasattr(self, 'nickname_edit'):
            self.nickname_edit.update()

    def _collectVariables(self, text, source):
        """将该来源文本中每个变量匹配逐条加入右侧变量列表（不去重）

        参数：
            text<str>：某输入框当前文本
            source<str>：来源标识（昵称/主题/正文），用于区分同名变量所处位置
        """
        for match in self.VARIABLE_PATTERN.finditer(text):
            var = match.group(1)
            item = QListWidgetItem('%s: $$%s$$' % (source, var))
            self.variable_list.addItem(item)

    # clearForm / setWriteMode / rememberRecordId 等沿用 NormalPage 的实现


class PreviewMailDialog(QDialog):
    """邮件预览查看/编辑对话框（像真正邮件一样呈现）

    默认只读查看一封完整的邮件：顶部邮件头（发件人/收件人/主题/附件摘要），
    正文用 QTextBrowser 渲染 HTML（含内嵌图），附件以列表展示。
    点「编辑」切换到可编辑表单，改完「保存」写回源邮件 dict（仅预览批次）；
    「退出编辑」时若有改动未保存则弹窗提醒。
    """

    def __init__(self, parent, mail):
        """构建详情对话框

        参数：
            parent<AdvancedPage>：父窗体（用于回调写回 mail 与获取发件人信息）
            mail<dict>：待查看/编辑的邮件数据（含 email/title/nickname/html/attachments）
        """
        super().__init__(parent)
        self._parent = parent
        self._mail = mail
        self._editing = False            # 当前是否处于可编辑状态
        self._dirty = False              # 是否有未保存的改动
        self.setWindowTitle('邮件预览 - 收件人 %s' % (mail.get('email') or '未知'))
        self.resize(760, 600)
        self._buildUi(mail)
        self._setEditing(False)          # 初始为只读邮件视图

    # ---- 发件人信息 ----
    def _senderHeader(self):
        """构造发件人展示信息：(昵称, 邮箱)

        发件人昵称优先取本封邮件配置页的发件人昵称（nickname，已做变量替换）；
        未配置（空）时回退到底部 from_edit，仍为空才用默认占位名。
        """
        sender_name = (self._mail.get('nickname') or '').strip()
        if not sender_name:
            sender_name = (self._parent.from_edit.text().strip()
                           if self._parent and hasattr(self._parent, 'from_edit')
                           else '')
        if not sender_name:
            sender_name = DEFAULT_FROM_NAME
        return sender_name, SMTP_USERNAME

    # ---- 只读视图：像真正邮件那样呈现 ----
    def _buildReadView(self):
        """构建只读邮件视图：头部区 + HTML 正文 + 附件列表"""
        page = QWidget()
        v = QVBoxLayout(page)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)

        # 邮件头：发件人 / 收件人
        sender_name, sender_email = self._senderHeader()
        self.read_from_label = QLabel('发件人：%s <%s>'
                                      % (sender_name, sender_email))
        self.read_to_label = QLabel('收件人：%s' % (self._mail.get('email') or ''))
        for lab in (self.read_from_label, self.read_to_label):
            lab.setStyleSheet('color:#374151;')
        v.addWidget(self.read_from_label)
        v.addWidget(self.read_to_label)

        # 主题（更大字号，分隔线）
        self.read_title_label = QLabel(self._mail.get('title') or '(无主题)')
        self.read_title_label.setStyleSheet(
            'font-size:15px;font-weight:600;color:#111;padding:6px 2px 8px 2px;'
            'border-bottom:1px solid #e5e7eb;')
        self.read_title_label.setWordWrap(True)
        v.addWidget(self.read_title_label)

        # 正文：HTML 富文本渲染（如同打开真实邮件）
        self.read_browser = QTextBrowser()
        self.read_browser.setOpenExternalLinks(True)
        self._renderHtml()
        v.addWidget(self.read_browser, 1)

        # 附件：列表展示文件名
        self.read_attach_label = QLabel()
        self.read_attach_label.setStyleSheet('color:#374151;')
        self.read_attach_label.setWordWrap(True)
        v.addWidget(self.read_attach_label)
        self._refreshAttachLabel()
        return page

    def _refreshAttachLabel(self):
        """刷新只读视图附件的显示（每行一个附件文件名）"""
        attach_list = self._mail.get('attachments') or []
        if not attach_list:
            self.read_attach_label.setText('')
            return
        names = [os.path.basename(p) if p else p for p in attach_list]
        self.read_attach_label.setText(
            '附件（%d）：%s' % (len(names), '  '.join(names)))

    def _renderHtml(self):
        """把邮件 HTML 渲染进正文浏览器，并注册内嵌图片资源使 cid 图片显示"""
        # 注册内嵌图：cid:文件名 -> 实际图片路径（来自配置页 inline_image_paths）
        # 资源 key 须带 cid: 前缀（与编辑器插入图片时的注册方式一致），否则图片无法显示
        doc = self.read_browser.document()
        inline = getattr(self._parent, 'inline_image_paths', None) or {}
        for cid, path in inline.items():
            if os.path.isfile(path):
                doc.addResource(
                    QTextDocument.ResourceType.ImageResource,
                    QUrl('cid:' + cid), QImage(path))
        self.read_browser.setHtml(self._mail.get('html') or '')

    # ---- 编辑视图：可编辑表单 ----
    def _buildEditView(self):
        """构建可编辑视图：收件人/主题/昵称/正文表单 + 附件只读展示"""
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        info_form = QFormLayout()
        self.email_edit = QLineEdit(self._mail.get('email') or '')
        self.title_edit = QLineEdit(self._mail.get('title') or '')
        self.nick_edit = QLineEdit(self._mail.get('nickname') or '')
        # 任一字段被修改即标记未保存改动（供退出/关闭时提醒）
        self.email_edit.textChanged.connect(self._onContentChanged)
        self.title_edit.textChanged.connect(self._onContentChanged)
        self.nick_edit.textChanged.connect(self._onContentChanged)
        # 附件编辑：列表 + 选择/移除按钮（保存时写回该封邮件）
        attach_box = QWidget()
        att_lay = QHBoxLayout(attach_box)
        att_lay.setContentsMargins(0, 0, 0, 0)
        self.attach_list = QListWidget()
        self.attach_list.setMaximumHeight(96)
        # 全路径列表与界面列表一一对应（界面只显示文件名）
        self._edit_attachments = list(self._mail.get('attachments') or [])
        self._reloadAttachList()
        att_lay.addWidget(self.attach_list, 1)
        btn_col = QVBoxLayout()
        pick_btn = QPushButton('选择附件…')
        pick_btn.clicked.connect(self._onPickAttachments)
        btn_col.addWidget(pick_btn)
        remove_btn = QPushButton('移除选中')
        remove_btn.clicked.connect(self._onRemoveAttachment)
        btn_col.addWidget(remove_btn)
        btn_col.addStretch()
        att_lay.addLayout(btn_col)
        info_form.addRow('附件：', attach_box)
        lay.addLayout(info_form)

        # 正文（HTML 富文本）
        self.body_editor = QTextEdit()
        self.body_editor.setAcceptRichText(True)
        self.body_editor.setHtml(self._mail.get('html') or '')
        self.body_editor.setMinimumHeight(320)
        self.body_editor.textChanged.connect(self._onContentChanged)
        lay.addWidget(self.body_editor, 1)
        return page

    def _buildUi(self, mail):
        """构建对话框 UI：只读邮件视图 + 可编辑视图（QStackedWidget 切换）+ 底部按钮"""
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(12, 10, 12, 10)

        # 主区：只读视图(0) / 编辑视图(1)，切换即切换查看/编辑
        self.view_stack = QStackedWidget()
        self.view_stack.addWidget(self._buildReadView())
        self.view_stack.addWidget(self._buildEditView())
        lay.addWidget(self.view_stack, 1)

        # ---- 底部按钮 ----
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.edit_toggle_btn = QPushButton('编辑')
        self.edit_toggle_btn.setObjectName('loadPrimaryBtn')
        self.edit_toggle_btn.clicked.connect(self._onToggleEdit)
        btn_row.addWidget(self.edit_toggle_btn)
        self.save_btn = QPushButton('保存')
        self.save_btn.setObjectName('loadPrimaryBtn')
        self.save_btn.clicked.connect(self._onSave)
        btn_row.addWidget(self.save_btn)
        self.exit_edit_btn = QPushButton('退出编辑')
        self.exit_edit_btn.setObjectName('loadActionBtn')
        self.exit_edit_btn.clicked.connect(self._onExitEdit)
        btn_row.addWidget(self.exit_edit_btn)
        self.close_btn = QPushButton('关闭')
        self.close_btn.clicked.connect(self.accept)
        btn_row.addWidget(self.close_btn)
        lay.addLayout(btn_row)

    def _setEditing(self, editing):
        """切换查看/编辑状态：只读时展示真实邮件视图，编辑时切换到表单"""
        self._editing = editing
        # 切换视图显示
        self.view_stack.setCurrentIndex(1 if editing else 0)
        # 编辑/退出编辑 按钮状态切换（保存始终可用）
        self.edit_toggle_btn.setVisible(not editing)
        self.exit_edit_btn.setVisible(editing)

    def _onContentChanged(self):
        """任一编辑字段内容变化时，标记存在未保存改动"""
        # 仅当处于编辑态时记录脏标记（只读态不会触发 textChanged 修改内容）
        if self._editing:
            self._dirty = True

    def _onToggleEdit(self):
        """点「编辑」：进入可编辑状态"""
        self._setEditing(True)

    def _onExitEdit(self):
        """点「退出编辑」：若已有未保存改动则弹窗提醒，否则直接退出编辑态"""
        if self._dirty:
            # 提醒保存：让用户选择 保存并退出 / 放弃更改 / 取消
            choice = QMessageBox.warning(
                self, '未保存的更改',
                '有未保存的修改，退出编辑后这些更改将丢失。\n是否先保存？',
                QMessageBox.StandardButton.Save
                | QMessageBox.StandardButton.Discard
                | QMessageBox.StandardButton.Cancel)
            if choice == QMessageBox.StandardButton.Save:
                self._onSave()
            elif choice == QMessageBox.StandardButton.Cancel:
                return        # 留在编辑态
            # Discard：放弃更改（不清除已编辑内容，但退出编辑态；已编辑内容不回写）
        self._dirty = False
        self._setEditing(False)

    def _reloadAttachList(self):
        """按当前全路径列表刷新附件编辑列表（只显示文件名）"""
        self.attach_list.clear()
        for p in self._edit_attachments:
            self.attach_list.addItem(os.path.basename(p) if p else p)

    def _onPickAttachments(self):
        """点「选择附件…」：弹选择框，把选中的附件加入本封邮件，并标记未保存"""
        paths, _ = QFileDialog.getOpenFileNames(
            self, '选择附件', '', '所有文件 (*.*)')
        if not paths:
            return
        # 避免重复加入同一路径
        existing = set(self._edit_attachments)
        for p in paths:
            if p not in existing:
                self._edit_attachments.append(p)
                existing.add(p)
        self._reloadAttachList()
        self._onContentChanged()

    def _onRemoveAttachment(self):
        """点「移除选中」：从本封邮件移除选中的附件并标记未保存"""
        row = self.attach_list.currentRow()
        if row < 0:
            return
        del self._edit_attachments[row]
        self._reloadAttachList()
        self._onContentChanged()

    def _onSave(self):
        """保存当前编辑内容：写回源邮件 dict（仅预览批次，不回写负载文件）"""
        self._mail['email'] = self.email_edit.text().strip()
        self._mail['title'] = self.title_edit.text().strip()
        self._mail['nickname'] = self.nick_edit.text().strip()
        self._mail['html'] = self.body_editor.toHtml()
        # 附件：写回编辑后的附件全路径列表
        self._mail['attachments'] = list(self._edit_attachments)
        self._dirty = False
        # 更新只读视图内容并回到只读视图，体现「正式邮件预览」
        self.read_from_label.setText('发件人：%s <%s>' % self._senderHeader())
        self.read_to_label.setText('收件人：%s' % (self._mail.get('email') or ''))
        self.read_title_label.setText(self._mail.get('title') or '(无主题)')
        self._renderHtml()
        self._refreshAttachLabel()
        self._setEditing(False)
        # 写回预览批次后，刷新列表对应行的摘要
        self._parent._refreshPreviewRowFor(self._mail)
        QMessageBox.information(self, '已保存', '已保存到本次预览批次。')

    def closeEvent(self, event):
        """点右上角关闭/「关闭」：若有未保存改动则提醒"""
        if self._dirty:
            choice = QMessageBox.warning(
                self, '未保存的更改',
                '有未保存的修改，关闭后将丢失。\n是否先保存？',
                QMessageBox.StandardButton.Save
                | QMessageBox.StandardButton.Discard
                | QMessageBox.StandardButton.Cancel)
            if choice == QMessageBox.StandardButton.Save:
                self._onSave()
            elif choice == QMessageBox.StandardButton.Cancel:
                event.ignore()
                return
        event.accept()
